"""Excel (xlsx) building utilities. Extracted from server.py (Phase A refactor)."""
from io import BytesIO
from typing import Any, Dict, List

from fastapi.responses import StreamingResponse


def build_xlsx(sheets: List[Dict[str, Any]]) -> bytes:
    """Build a multi-sheet XLSX file. Each sheet: {name, headers: [str], rows: [[cell,...]]}."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="0F3A73", end_color="0F3A73", fill_type="solid")
    header_font = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for sheet in sheets:
        ws = wb.create_sheet(title=(sheet.get("name") or "Sheet")[:31])
        ws.sheet_view.rightToLeft = True
        headers = sheet.get("headers") or []
        ws.append(headers)
        for col_idx, _h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
            cell.border = border
        for r in (sheet.get("rows") or []):
            ws.append(r)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = body_font
                c.alignment = centered
                c.border = border
        for col_idx, _h in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = 22
        ws.freeze_panes = "A2"
    if not wb.sheetnames:
        wb.create_sheet("Empty")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def xlsx_response(data: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
