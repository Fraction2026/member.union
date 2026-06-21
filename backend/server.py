from fastapi import FastAPI, APIRouter, Body, Depends, File, Header, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import Any, Dict, List, Optional, Set, Tuple
import uuid
from datetime import datetime, timedelta, timezone
import hashlib
import re
import shutil
import mimetypes
from collections import defaultdict, Counter
from difflib import SequenceMatcher
import json
import time
import asyncio

# OCR libs (pypdf, pytesseract, pdf2image, PIL) are now used inside services/ocr.py.

from services.letters_html import render_letters_html, LETTER_FOOTER_TEXT, LETTER_FOOTER_EMAIL
from services.caseform_html import render_case_research_form_html as _render_case_form_external
from services.excel import build_xlsx as _build_xlsx_external, xlsx_response as _xlsx_response_external
from services.ocr import (
    extract_pdf_text,
    extract_text_for_document,
    first_match,
    date_near_label,
    parse_membership_fields as _parse_membership_fields_external,
)
from services.retirement import (
    DEFAULT_RETIREMENT_SCHEDULE,
    parse_date_value,
    add_years,
    compute_retirement_info,
)
from services.inheritance_calculator import calculate_inheritance


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')
STORAGE_DIR = ROOT_DIR / "storage"
MEMBERS_DIR = STORAGE_DIR / "members"
TMP_DIR = STORAGE_DIR / "tmp"
for directory in (MEMBERS_DIR, TMP_DIR):
    directory.mkdir(parents=True, exist_ok=True)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin123"
DEFAULT_DEPARTMENT_NAME = "مشروع التكافل الاجتماعي"
# DEFAULT_RETIREMENT_SCHEDULE is now imported from services.retirement


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def today_ar() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def build_id() -> str:
    return str(uuid.uuid4())


def hash_password(password: str, salt: Optional[str] = None) -> Dict[str, str]:
    password_salt = salt or uuid.uuid4().hex
    digest = hashlib.sha256(f"{password_salt}:{password}".encode("utf-8")).hexdigest()
    return {"salt": password_salt, "password_hash": digest}


def verify_password(password: str, salt: str, password_hash: str) -> bool:
    return hash_password(password, salt)["password_hash"] == password_hash


def safe_filename(filename: str) -> str:
    suffix = Path(filename or "document.pdf").suffix.lower() or ".pdf"
    return f"{build_id()}{suffix}"


# Template for the one-shot Windows installer/updater script.
# The placeholder __SERVER_URL__ is replaced at request time with the public backend URL.
SETUP_BAT_TEMPLATE = r"""@echo off
title Electronic Archive Setup v2
setlocal enableextensions enabledelayedexpansion

REM ============================================================
REM  Electronic Archive  -  One-shot Installer / Updater  v2
REM  This script installs or updates the application on Windows.
REM ============================================================

cls
echo.
echo ============================================================
echo  Electronic Archive  -  Installer / Updater  (v2)
echo ============================================================
echo.
echo  If you can SEE this message, the script is running fine.
echo  Press any key to start installation ...
echo.
pause

set "SERVER_URL=__SERVER_URL__"
set "INSTALL_DIR=%LOCALAPPDATA%\ElectronicArchive"
set "ZIP_PATH=%TEMP%\ElectronicArchive.zip"
set "PORT=8090"
set "LOGFILE=%TEMP%\electronic_archive_install.log"
echo. > "%LOGFILE%"

echo.
echo  Install dir: %INSTALL_DIR%
echo  Server URL:  %SERVER_URL%
echo  Port:        %PORT%
echo  Log file:    %LOGFILE%
echo.

REM ----- 0. Verify Administrator -----
net session >nul 2>&1
if errorlevel 1 (
    echo [ERROR] You must run this script as Administrator.
    echo Right-click the file and choose "Run as administrator".
    echo.
    pause
    exit /b 1
)
echo [OK] Administrator rights confirmed.

REM ----- 1. Stop any old server -----
echo.
echo [1/8] Stopping any old server ...
taskkill /F /IM pythonw.exe /T >nul 2>&1
echo [OK] Done.

REM ----- 2. Download latest ZIP -----
echo.
echo [2/8] Downloading latest application from server ...
echo       URL: %SERVER_URL%/api/installer/download
powershell -NoProfile -Command "try { [Net.ServicePointManager]::SecurityProtocol = 'Tls12,Tls11,Tls'; Invoke-WebRequest -Uri '%SERVER_URL%/api/installer/download' -OutFile '%ZIP_PATH%' -UseBasicParsing -TimeoutSec 180; exit 0 } catch { Write-Host $_; exit 1 }" 2>>"%LOGFILE%"
if errorlevel 1 (
    echo [ERROR] Failed to download. Check internet connection.
    echo See log file: %LOGFILE%
    echo.
    pause
    exit /b 1
)
for %%I in ("%ZIP_PATH%") do set "ZIP_SIZE=%%~zI"
echo [OK] Downloaded %ZIP_SIZE% bytes.

REM ----- 3. Backup user data and extract -----
echo.
echo [3/8] Extracting files to %INSTALL_DIR% ...
if not exist "%INSTALL_DIR%" mkdir "%INSTALL_DIR%"
if exist "%INSTALL_DIR%\backend\storage" (
    if exist "%TEMP%\ea_storage_backup" rmdir /S /Q "%TEMP%\ea_storage_backup" >nul 2>&1
    xcopy "%INSTALL_DIR%\backend\storage" "%TEMP%\ea_storage_backup" /E /I /Q /Y >nul
)
if exist "%INSTALL_DIR%\backend\.env" (
    copy /Y "%INSTALL_DIR%\backend\.env" "%TEMP%\ea_env_backup" >nul
)
powershell -NoProfile -Command "Expand-Archive -Path '%ZIP_PATH%' -DestinationPath '%INSTALL_DIR%' -Force" 2>>"%LOGFILE%"
if errorlevel 1 (
    echo [ERROR] Failed to extract ZIP.
    echo See log file: %LOGFILE%
    echo.
    pause
    exit /b 1
)
del /Q "%ZIP_PATH%" >nul 2>&1
if exist "%TEMP%\ea_storage_backup" (
    xcopy "%TEMP%\ea_storage_backup" "%INSTALL_DIR%\backend\storage" /E /I /Q /Y >nul
    rmdir /S /Q "%TEMP%\ea_storage_backup" >nul 2>&1
)
if exist "%TEMP%\ea_env_backup" (
    copy /Y "%TEMP%\ea_env_backup" "%INSTALL_DIR%\backend\.env" >nul
    del /Q "%TEMP%\ea_env_backup" >nul 2>&1
)
echo [OK] Files extracted.

REM ----- 4. Python check -----
echo.
echo [4/8] Checking Python ...
where python >nul 2>&1
if errorlevel 1 (
    echo       Installing Python 3.11 ...
    winget install -e --id Python.Python.3.11 --silent --accept-source-agreements --accept-package-agreements >>"%LOGFILE%" 2>&1
    set "PATH=%LOCALAPPDATA%\Programs\Python\Python311;%LOCALAPPDATA%\Programs\Python\Python311\Scripts;%PATH%"
) else (
    echo [OK] Python already installed.
)

REM ----- 5. MongoDB check -----
echo.
echo [5/8] Checking MongoDB ...
where mongod >nul 2>&1
if errorlevel 1 (
    echo       Installing MongoDB ...
    winget install -e --id MongoDB.Server --silent --accept-source-agreements --accept-package-agreements >>"%LOGFILE%" 2>&1
) else (
    echo [OK] MongoDB already installed.
)
sc query MongoDB >nul 2>&1
if not errorlevel 1 (
    net start MongoDB >nul 2>&1
)

REM ----- 6. Tesseract check -----
echo.
echo [6/8] Checking Tesseract OCR ...
where tesseract >nul 2>&1
if errorlevel 1 (
    echo       Installing Tesseract ...
    winget install -e --id UB-Mannheim.TesseractOCR --silent --accept-source-agreements --accept-package-agreements >>"%LOGFILE%" 2>&1
) else (
    echo [OK] Tesseract already installed.
)

REM ----- 7. Install/update Python libs -----
echo.
echo [7/8] Installing Python libraries (may take several minutes) ...
cd /d "%INSTALL_DIR%\backend"
python -m pip install --upgrade pip >>"%LOGFILE%" 2>&1
if exist requirements-local.txt (
    python -m pip install -r requirements-local.txt >>"%LOGFILE%" 2>&1
) else (
    python -m pip install -r requirements.txt >>"%LOGFILE%" 2>&1
)
if errorlevel 1 (
    echo [WARN] Some Python packages failed. See log: %LOGFILE%
)
if not exist .env (
    echo MONGO_URL=mongodb://localhost:27017> .env
    echo DB_NAME=electronic_archive>> .env
    echo CORS_ORIGINS=*>> .env
)
echo [OK] Done.

REM ----- 8. Build frontend -----
echo.
echo [8/8] Building web frontend (3-8 minutes) ...
cd /d "%INSTALL_DIR%\frontend"
where node >nul 2>&1
if errorlevel 1 (
    echo       Installing Node.js ...
    winget install -e --id OpenJS.NodeJS.LTS --silent --accept-source-agreements --accept-package-agreements >>"%LOGFILE%" 2>&1
    set "PATH=%ProgramFiles%\nodejs;%PATH%"
)
set "REACT_APP_BACKEND_URL="
if not exist node_modules (
    call npm install --no-audit --no-fund --loglevel=error >>"%LOGFILE%" 2>&1
)
call npm run build >>"%LOGFILE%" 2>&1
if errorlevel 1 (
    echo [WARN] Frontend build had warnings. See log: %LOGFILE%
)
echo [OK] Done.

REM ----- Firewall -----
echo.
echo Opening firewall port %PORT% ...
netsh advfirewall firewall delete rule name="Electronic Archive Server" >nul 2>&1
netsh advfirewall firewall add rule name="Electronic Archive Server" dir=in action=allow protocol=TCP localport=%PORT% >nul

REM ----- Create silent runner -----
set "RUN_BAT=%INSTALL_DIR%\run_archive.bat"
> "%RUN_BAT%" echo @echo off
>> "%RUN_BAT%" echo cd /d "%INSTALL_DIR%\backend"
>> "%RUN_BAT%" echo start "" /B pythonw -m uvicorn server:app --host 0.0.0.0 --port %PORT%

REM ----- Auto-start with Windows -----
schtasks /Create /F /SC ONSTART /RL HIGHEST /TN "ElectronicArchiveServer" /TR "\"%RUN_BAT%\"" >nul 2>&1

REM ----- Desktop shortcut -----
set "DESKTOP=%USERPROFILE%\Desktop"
powershell -NoProfile -Command "$s=(New-Object -ComObject WScript.Shell).CreateShortcut('%DESKTOP%\Electronic Archive.lnk'); $s.TargetPath='http://localhost:%PORT%'; $s.Save()" >nul 2>&1

REM ----- Start the server now -----
echo.
echo Starting server now ...
start "" /B "%RUN_BAT%"
timeout /t 4 >nul

set "MY_IP=localhost"
for /f "tokens=2 delims=:" %%a in ('ipconfig ^| findstr /R /C:"IPv4"') do (
    set "TMP_IP=%%a"
    set "TMP_IP=!TMP_IP: =!"
    if "!TMP_IP:~0,7!"=="192.168" set "MY_IP=!TMP_IP!"
    if "!TMP_IP:~0,3!"=="10." set "MY_IP=!TMP_IP!"
)

cls
echo.
echo ============================================================
echo   [SUCCESS] Electronic Archive is running.
echo ============================================================
echo.
echo   On THIS computer:           http://localhost:%PORT%
echo   From OTHER devices:         http://!MY_IP!:%PORT%
echo.
echo   Default credentials:
echo      username:  admin
echo      password:  admin123
echo.
echo   * Auto-starts with Windows
echo   * To update later: run this same file again
echo   * Desktop shortcut: "Electronic Archive"
echo   * Install log: %LOGFILE%
echo ============================================================
echo.
start http://localhost:%PORT%
pause
"""



def parse_membership_fields(text: str) -> Dict[str, str]:
    return _parse_membership_fields_external(text, today_ar())


def retirement_defaults() -> List[Dict[str, Any]]:
    return [{**entry, "created_at": now_iso(), "updated_at": now_iso()} for entry in DEFAULT_RETIREMENT_SCHEDULE]


async def get_retirement_schedule_docs() -> List[Dict[str, Any]]:
    schedule = await db.retirement_schedule.find({}, {"_id": 0}).sort("effective_date", 1).to_list(100)
    return schedule or DEFAULT_RETIREMENT_SCHEDULE


async def enrich_member_retirement(member: Dict[str, Any]) -> Dict[str, Any]:
    schedule = await get_retirement_schedule_docs()
    return {**member, **compute_retirement_info(member.get("birth_date", ""), schedule)}


async def enrich_members_retirement(members: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    schedule = await get_retirement_schedule_docs()
    return [{**member, **compute_retirement_info(member.get("birth_date", ""), schedule)} for member in members]


# Define Models
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")  # Ignore MongoDB's _id field
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str


class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    token: str
    user: Dict[str, Any]


class DepartmentIn(BaseModel):
    name: str
    code: str = ""
    description: str = ""
    active: bool = True


class Department(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=build_id)
    name: str
    code: str = ""
    description: str = ""
    active: bool = True
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class RetirementScheduleIn(BaseModel):
    effective_date: str
    retirement_age: int
    description: str = ""


class RetirementSchedule(RetirementScheduleIn):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=build_id)
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class DocumentRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=build_id)
    file_name: str
    original_name: str
    content_type: str
    path: str
    source: str = "upload"
    batch_id: str = ""
    page_start: int = 0
    page_end: int = 0
    extracted_text: str = ""
    extracted_fields: Dict[str, str] = Field(default_factory=dict)
    created_at: str = Field(default_factory=now_iso)


class MemberIn(BaseModel):
    department_id: str
    document_id: Optional[str] = ""
    governorate: str = ""
    union_committee: str = ""
    membership_number: str = ""
    name: str = ""
    national_id: str = ""
    birth_date: str = ""
    subscription_date: str = ""
    address_phone: str = ""
    status: str = "فعال"
    status_date: str = Field(default_factory=today_ar)
    address: str = ""
    beneficiary_name: str = ""


class MemberStatusUpdate(BaseModel):
    status: str
    status_date: Optional[str] = None


class Member(MemberIn):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=build_id)
    document_file_name: str = ""
    document_original_name: str = ""
    document_url: str = ""
    retirement_age: Optional[int] = None
    retirement_date: str = ""
    retirement_due: bool = False
    retirement_label: str = ""
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


FINANCIAL_CATEGORIES = {
    "pension": "المعاش",
    "resignations": "استقالات",
    "dropout": "إسقاط",
    "letters_received": "خطابات مستلمة",
    "letters_sent": "خطابات مرسلة",
    "aid_pending": "إعانات في انتظار الموافقة",
    "aid_disbursed": "إعانات تم صرفها",
}


class CategoryRecordIn(BaseModel):
    department_id: str
    category: str
    document_id: Optional[str] = ""
    name: str = ""
    national_id: str = ""
    membership_number: str = ""
    record_date: str = Field(default_factory=today_ar)
    reference_number: str = ""
    subject: str = ""
    notes: str = ""


class CategoryRecord(CategoryRecordIn):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=build_id)
    document_file_name: str = ""
    document_original_name: str = ""
    document_url: str = ""
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


SUBSCRIPTION_STATUSES = {"تم التحصيل", "تحت التحصيل", "لا يخص مشروع التكافل", "لاغي"}
SUBSCRIPTION_PAYMENT_METHODS = {"دفع الكتروني", "شيك"}
SUBSCRIPTION_NOT_TAKAFUL_TARGETS = {"النقابة العامة", "جهة أخرى"}
COMMITTEE_MONTHLY_RATE = 3.0  # قيمة الاشتراك الشهري لكل عضو


# Some legacy rows ended up with a date (like "2/17/1954") in the
# governorate / committee fields because data was pasted into the wrong
# column. We hide such rows from every aggregation, dropdown and report.
_DATE_LIKE_RE = re.compile(r"^\s*\d{1,4}[/.\-]\d{1,2}[/.\-]\d{1,4}\s*$")


def _is_invalid_taxonomy_value(s: Any) -> bool:
    """True when the value should be hidden from reports / dropdowns."""
    if not s:
        return True
    text = str(s).strip()
    if not text:
        return True
    return bool(_DATE_LIKE_RE.match(text))


class SubscriptionIn(BaseModel):
    department_id: str
    permit_number: str = ""
    amount: float = 0.0
    governorate: str = ""
    union_committee: str = ""
    payment_method: str = "دفع الكتروني"
    electronic_reference: str = ""           # for "دفع الكتروني"
    # "number" → bank/gateway transaction id, enforced unique within the
    # department.  "text" → manual descriptive label (e.g. statement note),
    # explicitly allowed to repeat across rows.
    electronic_reference_kind: str = "number"
    cheque_number: str = ""                  # for "شيك"
    cheque_bank: str = ""
    cheque_date: str = ""
    subscription_month: str = ""             # YYYY-MM
    issued_at: str = Field(default_factory=today_ar)   # تحريراً في (YYYY-MM-DD)
    status: str = "تحت التحصيل"
    not_takaful_target: str = ""             # "النقابة العامة" or "جهة أخرى"
    not_takaful_other: str = ""              # نص حر لو "جهة أخرى"
    notes: str = ""
    is_dues_settlement: bool = False         # True = تسوية مستحقات على اللجنة (يخصم من المستحقات)


class Subscription(SubscriptionIn):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=build_id)
    recorded_by_id: str = ""
    recorded_by_name: str = ""
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class SubscriptionStatusUpdate(BaseModel):
    status: str
    not_takaful_target: Optional[str] = None
    not_takaful_other: Optional[str] = None


# ===== الإعانات (Aids) =====
AID_TYPE_DEATH = "وفاة"
AID_TYPE_DISABILITY = "عجز كلي أو جزئي منهي للخدمة"
STATUS_TO_AID_TYPE = {
    "متوفي": AID_TYPE_DEATH,
    "عجز كلي أو جزئي منهي للخدمة": AID_TYPE_DISABILITY,
}
AID_STATUS_PENDING = "pending"
AID_STATUS_DISBURSED = "disbursed"


class AidDisburseIn(BaseModel):
    cheque_number: str
    cheque_date: str
    cheque_bank: str
    amount: float
    beneficiaries: List[str] = []


class AidRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=build_id)
    department_id: str
    member_id: str
    aid_type: str
    status: str = AID_STATUS_PENDING
    triggered_by_status: str = ""
    triggered_at: str = Field(default_factory=now_iso)
    # member snapshot (denormalized for stable display even if member edited later)
    member_name: str = ""
    member_national_id: str = ""
    member_membership_number: str = ""
    member_governorate: str = ""
    member_union_committee: str = ""
    member_birth_date: str = ""
    member_subscription_date: str = ""
    member_address: str = ""
    member_phone: str = ""
    member_beneficiary_name: str = ""
    member_status_date: str = ""
    # disbursement fields (filled when status = disbursed)
    cheque_number: str = ""
    cheque_date: str = ""
    cheque_bank: str = ""
    amount: float = 0.0
    beneficiaries: List[str] = []
    disbursed_at: str = ""
    disbursed_by_id: str = ""
    disbursed_by_name: str = ""
    # Toggle: include committee-dues note in printed case study form (الاستمارة)
    print_dues_note: bool = True
    created_at: str = Field(default_factory=now_iso)
    updated_at: str = Field(default_factory=now_iso)


class AidBeneficiary(BaseModel):
    """مستحق للإعانة"""
    model_config = ConfigDict(extra="ignore")
    name: str
    relation: str  # زوج، زوجة، أب، أم، ابن، ابنة
    base_share_fraction: str = ""  # الفرض الأصلي قبل الرد (مثل: 1/8، 2/3)
    base_share_arabic: str = ""  # اسم الفرض بالعربية (مثل: ثمن، ثلثان)
    radd_fraction: str = ""  # حصة الرد إن وجدت (مثل: 5/24)
    final_share_fraction: str = ""  # النصيب النهائي بعد الرد (مثل: 7/24)
    percentage: str = ""  # النسبة الرقمية النهائية (مثل: 1/2، 1/4) - للتوافق مع الكود القديم
    percentage_arabic: str = ""  # النسبة العربية (مثل: نصف، ربع)
    share_type: str = ""  # نوع الاستحقاق: فرض، فرض + رد، تعصيب
    share_group_text: str = ""  # النص المشترك للتجميع (مثل: ثلثان فرضًا والباقي ردًا)
    share_group_key: str = ""  # مفتاح التجميع للمستحقين المتشابهين
    amount: float = 0.0  # المبلغ المستحق
    inheritance_type: str = ""  # نوع الاستحقاق: فرض، تعصيب، رد (للتوافق مع الكود القديم)
    explanation: str = ""  # التفسير الشرعي


class AidBeneficiariesData(BaseModel):
    """بيانات المستحقين للإعانة"""
    model_config = ConfigDict(extra="ignore")
    aid_id: str
    total_amount: float
    beneficiaries: List[AidBeneficiary]
    summary_explanation: str = ""  # الملخص النصي الكامل


async def seed_defaults() -> None:
    existing_admin = await db.users.find_one({"username": DEFAULT_ADMIN_USERNAME}, {"_id": 0})
    if not existing_admin:
        password_data = hash_password(DEFAULT_ADMIN_PASSWORD)
        await db.users.insert_one({
            "id": build_id(),
            "username": DEFAULT_ADMIN_USERNAME,
            "display_name": "مدير النظام",
            "role": "super_admin",
            "active": True,
            **password_data,
            "created_at": now_iso(),
        })
    else:
        # Upgrade legacy admin role to super_admin (only the default admin user)
        if existing_admin.get("role") != "super_admin":
            await db.users.update_one(
                {"username": DEFAULT_ADMIN_USERNAME},
                {"$set": {"role": "super_admin", "active": True}},
            )
    existing_department = await db.departments.find_one({"name": DEFAULT_DEPARTMENT_NAME}, {"_id": 0})
    if not existing_department:
        department = Department(
            name=DEFAULT_DEPARTMENT_NAME,
            code="TAKAFUL",
            description="أرشفة استمارات العضوية وربطها بملف PDF لكل عضو",
        ).model_dump()
        await db.departments.insert_one(department)
    retirement_count = await db.retirement_schedule.count_documents({})
    if retirement_count == 0:
        await db.retirement_schedule.insert_many(retirement_defaults())
    # Cleanup legacy scanner-related collections (feature removed)
    for legacy in ("scanner_configs", "scan_jobs", "scanner_bridge_status"):
        try:
            await db.drop_collection(legacy)
        except Exception:  # noqa: BLE001
            pass

    # ---- MongoDB indexes for performance (idempotent) ----
    try:
        await db.members.create_index([("department_id", 1)])
        await db.members.create_index([("department_id", 1), ("governorate", 1), ("union_committee", 1)])
        await db.members.create_index([("department_id", 1), ("name", 1)])
        await db.members.create_index([("department_id", 1), ("national_id", 1)])
        await db.members.create_index([("department_id", 1), ("status", 1)])
        await db.members.create_index([("department_id", 1), ("created_at", -1)])
        await db.subscriptions.create_index([("department_id", 1), ("date", -1)])
        await db.subscriptions.create_index([("member_id", 1)])
        await db.aids.create_index([("department_id", 1), ("status", 1)])
        await db.aids.create_index([("member_id", 1)])
        # Presence: TTL on expires_at_dt so stale heartbeats vanish automatically.
        await db.presence.create_index([("user_id", 1)], unique=True)
        await db.presence.create_index([("expires_at_dt", 1)], expireAfterSeconds=0)
        await db.sessions.create_index([("token", 1)], unique=True)
        await db.sessions.create_index([("user_id", 1)])
        # TTL index: MongoDB auto-deletes a session at its `expires_at_dt` moment.
        await db.sessions.create_index([("expires_at_dt", 1)], expireAfterSeconds=0)
        # One-time backfill: existing sessions only have iso-string `expires_at`.
        # Convert them so the TTL index can clean them up.
        try:
            cursor = db.sessions.find({"expires_at_dt": {"$exists": False}, "expires_at": {"$exists": True}}, {"_id": 0, "id": 1, "expires_at": 1})
            async for doc in cursor:
                try:
                    dt = datetime.fromisoformat(doc["expires_at"])
                    await db.sessions.update_one({"id": doc["id"]}, {"$set": {"expires_at_dt": dt}})
                except Exception:
                    pass
        except Exception:
            pass
        await db.users.create_index([("username", 1)], unique=True)
    except Exception as _idx_err:  # noqa: BLE001
        logging.getLogger(__name__).warning(f"Index creation warning: {_idx_err}")


async def current_user(authorization: Optional[str] = None):
    if not authorization:
        raise HTTPException(status_code=401, detail="تسجيل الدخول مطلوب")
    token = authorization.replace("Bearer", "").strip()
    session = await db.sessions.find_one({"token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="جلسة غير صالحة")
    expires_at = datetime.fromisoformat(session["expires_at"])
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة")
    user = await db.users.find_one({"id": session["user_id"]}, {"_id": 0, "password_hash": 0, "salt": 0})
    if not user:
        raise HTTPException(status_code=401, detail="المستخدم غير موجود")
    return user


async def require_user(authorization: Optional[str] = Header(None)):
    return await current_user(authorization)


async def require_admin(user: Dict[str, Any] = Depends(require_user)):
    if user.get("role") not in {"admin", "super_admin"}:
        raise HTTPException(status_code=403, detail="صلاحيات الأدمن مطلوبة")
    return user


async def require_super_admin(user: Dict[str, Any] = Depends(require_user)):
    if user.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="هذه العملية تتطلب صلاحيات المدير الأعلى")
    return user


VALID_USER_ROLES = {"super_admin", "admin", "employee"}
USER_ROLE_LABELS = {
    "super_admin": "مدير أعلى",
    "admin": "أدمن",
    "employee": "موظف",
}

# All portal keys (used in user permissions)
ALL_PORTAL_KEYS = [
    "membership",
    "financial",
    "financial.pension", "financial.resignations", "financial.dropout", "financial.letters",
    "financial.subscriptions", "financial.dues_settlements", "financial.aid", "financial.dues",
    "financial.disclosure",
    "admin", "users",
]
PORTAL_LABELS = {
    "membership": "العضوية",
    "financial": "الموقف المالي",
    "financial.pension": "الموقف المالي ← المعاش",
    "financial.resignations": "الموقف المالي ← الاستقالات",
    "financial.dropout": "الموقف المالي ← الإسقاط",
    "financial.letters": "الموقف المالي ← الخطابات",
    "financial.subscriptions": "الموقف المالي ← الاشتراكات",
    "financial.dues_settlements": "الموقف المالي ← تسوية المستحقات",
    "financial.aid": "الموقف المالي ← الإعانات",
    "financial.dues": "الموقف المالي ← مستحقات اللجان",
    "financial.disclosure": "الموقف المالي ← الكشوف التفريغية",
    "admin": "صفحة الأدمن",
    "users": "إدارة المستخدمين",
}


class UserCreate(BaseModel):
    username: str
    password: str
    display_name: str
    role: str = "employee"
    active: bool = True
    portal_permissions: List[str] = []


class UserUpdate(BaseModel):
    display_name: Optional[str] = None
    role: Optional[str] = None
    active: Optional[bool] = None
    password: Optional[str] = None  # if provided, password is reset
    portal_permissions: Optional[List[str]] = None


class UserOut(BaseModel):
    id: str
    username: str
    display_name: str
    role: str
    role_label: str = ""
    active: bool = True
    created_at: str = ""
    portal_permissions: List[str] = []


def _user_to_out(doc: Dict[str, Any]) -> UserOut:
    return UserOut(
        id=doc.get("id", ""),
        username=doc.get("username", ""),
        display_name=doc.get("display_name", ""),
        role=doc.get("role", "employee"),
        role_label=USER_ROLE_LABELS.get(doc.get("role", ""), doc.get("role", "")),
        active=doc.get("active", True),
        created_at=doc.get("created_at", ""),
        portal_permissions=doc.get("portal_permissions") or [],
    )


def _allowed_portals_for(user: Dict[str, Any]) -> List[str]:
    role = user.get("role", "")
    if role in {"super_admin", "admin"}:
        return list(ALL_PORTAL_KEYS)
    return user.get("portal_permissions") or []


class AppSettingsIn(BaseModel):
    session_timeout_hours: int = 12


SETTINGS_DOC_ID = "app_settings"


async def get_session_timeout_hours() -> int:
    doc = await db.settings.find_one({"id": SETTINGS_DOC_ID}, {"_id": 0})
    if not doc:
        return 12
    try:
        v = int(doc.get("session_timeout_hours") or 12)
        return max(1, min(v, 720))  # between 1 hour and 30 days
    except (ValueError, TypeError):
        return 12

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "نظام الأرشيف الإلكتروني يعمل"}


@api_router.post("/auth/login", response_model=LoginResponse)
async def login(input: LoginRequest):
    user = await db.users.find_one({"username": input.username.strip()}, {"_id": 0})
    if not user or not verify_password(input.password, user["salt"], user["password_hash"]):
        raise HTTPException(status_code=401, detail="بيانات الدخول غير صحيحة")
    if user.get("active") is False:
        raise HTTPException(status_code=403, detail="هذا الحساب موقوف. تواصل مع مدير النظام.")
    token = build_id() + build_id()
    hours = await get_session_timeout_hours()
    expires_at = datetime.now(timezone.utc) + timedelta(hours=hours)
    await db.sessions.insert_one({
        "id": build_id(),
        "token": token,
        "user_id": user["id"],
        "created_at": now_iso(),
        "expires_at": expires_at.isoformat(),
        # BSON Date for TTL index — auto-deleted by MongoDB at this exact moment.
        "expires_at_dt": expires_at,
    })
    return LoginResponse(
        token=token,
        user={
            "id": user["id"],
            "username": user["username"],
            "display_name": user["display_name"],
            "role": user["role"],
            "allowed_portals": _allowed_portals_for(user),
        },
    )


# -------------------- User Management (Admin) --------------------

@api_router.get("/admin/users", response_model=List[UserOut])
async def list_users(actor: Dict[str, Any] = Depends(require_admin)):
    docs = await db.users.find({}, {"_id": 0, "password_hash": 0, "salt": 0}).sort("created_at", 1).to_list(500)
    return [_user_to_out(d) for d in docs]


@api_router.post("/admin/users", response_model=UserOut)
async def create_user(payload: UserCreate, actor: Dict[str, Any] = Depends(require_admin)):
    username = (payload.username or "").strip()
    display_name = (payload.display_name or "").strip()
    role = (payload.role or "employee").strip()
    if not username:
        raise HTTPException(status_code=400, detail="اسم المستخدم مطلوب")
    if not display_name:
        raise HTTPException(status_code=400, detail="الاسم الكامل مطلوب")
    if role not in VALID_USER_ROLES:
        raise HTTPException(status_code=400, detail="دور غير مدعوم")
    if not (payload.password and len(payload.password) >= 4):
        raise HTTPException(status_code=400, detail="كلمة المرور قصيرة (4 أحرف على الأقل)")
    # Only super_admin can create another super_admin or admin
    if role in {"super_admin", "admin"} and actor.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="فقط المدير الأعلى يمكنه إضافة أدمن أو مدير أعلى")
    existing = await db.users.find_one({"username": username}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="اسم المستخدم مستخدم بالفعل")
    password_data = hash_password(payload.password)
    portal_perms = [p for p in (payload.portal_permissions or []) if p in ALL_PORTAL_KEYS]
    doc = {
        "id": build_id(),
        "username": username,
        "display_name": display_name,
        "role": role,
        "active": bool(payload.active),
        "portal_permissions": portal_perms,
        **password_data,
        "created_at": now_iso(),
    }
    await db.users.insert_one(doc)
    return _user_to_out(doc)


@api_router.put("/admin/users/{user_id}", response_model=UserOut)
async def update_user(user_id: str, payload: UserUpdate, actor: Dict[str, Any] = Depends(require_admin)):
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    update: Dict[str, Any] = {}
    if payload.display_name is not None:
        name = payload.display_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="الاسم الكامل لا يجوز أن يكون فارغًا")
        update["display_name"] = name
    if payload.role is not None:
        if payload.role not in VALID_USER_ROLES:
            raise HTTPException(status_code=400, detail="دور غير مدعوم")
        # Only super_admin can change roles to admin/super_admin or modify another super_admin
        if (payload.role in {"super_admin", "admin"} or target.get("role") == "super_admin") and actor.get("role") != "super_admin":
            raise HTTPException(status_code=403, detail="هذه العملية تتطلب صلاحيات المدير الأعلى")
        # Prevent demoting the last active super_admin
        if target.get("role") == "super_admin" and payload.role != "super_admin":
            others = await db.users.count_documents({"role": "super_admin", "active": True, "id": {"$ne": user_id}})
            if others == 0:
                raise HTTPException(status_code=400, detail="لا يمكن إزالة صلاحيات المدير الأعلى الوحيد في النظام")
        update["role"] = payload.role
    if payload.active is not None:
        # Prevent deactivating the last active super_admin
        if target.get("role") == "super_admin" and payload.active is False:
            others = await db.users.count_documents({"role": "super_admin", "active": True, "id": {"$ne": user_id}})
            if others == 0:
                raise HTTPException(status_code=400, detail="لا يمكن إيقاف المدير الأعلى الوحيد في النظام")
        update["active"] = bool(payload.active)
    if payload.password is not None and payload.password != "":
        if len(payload.password) < 4:
            raise HTTPException(status_code=400, detail="كلمة المرور قصيرة (4 أحرف على الأقل)")
        # Only super_admin can reset another super_admin's password
        if target.get("role") == "super_admin" and actor.get("id") != target.get("id") and actor.get("role") != "super_admin":
            raise HTTPException(status_code=403, detail="غير مسموح بتعديل كلمة مرور المدير الأعلى")
        update.update(hash_password(payload.password))
    if payload.portal_permissions is not None:
        update["portal_permissions"] = [p for p in payload.portal_permissions if p in ALL_PORTAL_KEYS]
    if update:
        update["updated_at"] = now_iso()
        await db.users.update_one({"id": user_id}, {"$set": update})
    new_doc = await db.users.find_one({"id": user_id}, {"_id": 0})
    return _user_to_out(new_doc)


@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, actor: Dict[str, Any] = Depends(require_admin)):
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    if target.get("id") == actor.get("id"):
        raise HTTPException(status_code=400, detail="لا يمكنك حذف حسابك")
    if target.get("role") == "super_admin" and actor.get("role") != "super_admin":
        raise HTTPException(status_code=403, detail="فقط المدير الأعلى يمكنه حذف مدير أعلى آخر")
    if target.get("role") == "super_admin":
        others = await db.users.count_documents({"role": "super_admin", "active": True, "id": {"$ne": user_id}})
        if others == 0:
            raise HTTPException(status_code=400, detail="لا يمكن حذف المدير الأعلى الوحيد في النظام")
    await db.users.delete_one({"id": user_id})
    await db.sessions.delete_many({"user_id": user_id})
    return {"deleted": True}


@api_router.get("/auth/me")
async def me(user: Dict[str, Any] = Depends(require_user)):
    return {**user, "allowed_portals": _allowed_portals_for(user)}


# =========================================================================
# Admin: detect & merge duplicate / near-duplicate union committee names.
# Real-world data often contains typos like "الاراة العامة للري" vs
# "الادارة العامة للري" — these should be unified under the most common spelling.
# =========================================================================

def _normalise_ar(s: str) -> str:
    """Normalise an Arabic name for similarity comparison: trim, collapse spaces,
    unify alef/ya forms, remove tatweel, remove diacritics. Also strips the
    'ال' (definite article) prefix on individual tokens so that
    "إدارة الزراعة" and "الإدارة الزراعية" line up better."""
    if not s:
        return ""
    s = s.strip()
    # Collapse internal whitespace (including invisible TAB / NBSP)
    s = re.sub(r"\s+", " ", s.replace("\u00A0", " ").replace("\u200F", "").replace("\u200E", ""))
    # Remove Arabic diacritics (tashkeel) U+064B..U+0652 and tatweel U+0640
    s = re.sub(r"[\u064B-\u0652\u0640]", "", s)
    # Unify alef variants
    s = s.replace("\u0623", "\u0627").replace("\u0625", "\u0627").replace("\u0622", "\u0627")
    # Unify ya / alef maqsura
    s = s.replace("\u0649", "\u064A")
    # Unify ta marbuta / ha
    s = s.replace("\u0629", "\u0647")
    return s


def _tokens_for_sim(s: str) -> List[str]:
    """Tokenise for similarity scoring. Drops short connectors ('في', 'و',
    'من', etc.) and strips leading 'ال' so 'إدارة الري' ≈ 'الإدارة للري'."""
    if not s:
        return []
    base = _normalise_ar(s)
    stopwords = {"في", "و", "من", "إلى", "الى", "علي", "علي", "عن", "هو", "هي"}
    out: List[str] = []
    for tok in base.split():
        # Strip a leading 'ال' (definite article) from each token.
        t = tok[2:] if tok.startswith("ال") and len(tok) > 3 else tok
        if t and t not in stopwords:
            out.append(t)
    return out


def _similarity(a: str, b: str) -> float:
    """Combined similarity score, robust to Arabic typos.

    Returns the MAX of three signals so we don't miss obvious matches that
    one metric would underestimate:

      1. SequenceMatcher.ratio on the normalised strings (char-level).
      2. Jaccard overlap of normalised tokens (catches reorderings).
      3. Token-set ratio: |intersection| / min(|a|, |b|) — catches the case
         where one name is a near-subset of the other ("إدارة ري" vs
         "الإدارة العامة للري").
    """
    if not a or not b:
        return 0.0
    na, nb = _normalise_ar(a), _normalise_ar(b)
    if not na or not nb:
        return 0.0
    char_ratio = SequenceMatcher(None, na, nb).ratio()
    ta = set(_tokens_for_sim(a))
    tb = set(_tokens_for_sim(b))
    jaccard = 0.0
    containment = 0.0
    if ta and tb:
        inter = len(ta & tb)
        if inter:
            jaccard = inter / len(ta | tb)
            containment = inter / min(len(ta), len(tb))
    return max(char_ratio, jaccard, containment)


@api_router.get("/admin/committees/duplicates")
async def detect_duplicate_committees(
    department_id: str,
    threshold: float = 0.72,
    cross_governorate: bool = True,
    user: Dict[str, Any] = Depends(require_admin),
):
    """Scan committees and return clusters of likely duplicates.

    Two detection passes (a cluster is the union of both):

      1) Same-governorate fuzzy match: committees inside one governorate whose
         normalised similarity ≥ `threshold` are grouped together — picks up
         typos in the committee name itself.

      2) Cross-governorate exact-name match (when `cross_governorate=True`):
         the SAME normalised committee name appearing under multiple
         governorates is flagged as a "possible governorate typo" — these
         used to be silently invisible to the operator.

    The most-frequent variant inside each cluster is suggested as the
    canonical (correct) spelling. Response also includes scan statistics so
    the operator knows exactly what was checked.
    """
    pipeline = [
        {"$match": {"department_id": department_id}},
        {"$group": {
            "_id": {"governorate": "$governorate", "committee": "$union_committee"},
            "count": {"$sum": 1},
        }},
    ]
    pairs = await db.members.aggregate(pipeline).to_list(100000)
    by_gov: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    by_normalised: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    total_committees = 0
    for p in pairs:
        gov = (p["_id"].get("governorate") or "").strip()
        com = (p["_id"].get("committee") or "").strip()
        if not gov or not com:
            continue
        total_committees += 1
        by_gov[gov].append({"name": com, "count": p["count"]})
        # Index for cross-governorate exact-match detection.
        by_normalised[_normalise_ar(com)].append({
            "name": com,
            "governorate": gov,
            "count": p["count"],
        })

    clusters: List[Dict[str, Any]] = []
    pairs_checked = 0

    # Pass 1 — same governorate, fuzzy similarity.
    for gov, items in by_gov.items():
        items_sorted = sorted(items, key=lambda x: (-x["count"], x["name"]))
        used: Set[int] = set()
        for i, anchor in enumerate(items_sorted):
            if i in used:
                continue
            group = [anchor]
            used.add(i)
            for j in range(i + 1, len(items_sorted)):
                if j in used:
                    continue
                pairs_checked += 1
                if _similarity(anchor["name"], items_sorted[j]["name"]) >= threshold:
                    group.append(items_sorted[j])
                    used.add(j)
            if len(group) >= 2:
                total = sum(g["count"] for g in group)
                clusters.append({
                    "kind": "same_governorate_fuzzy",
                    "governorate": gov,
                    "canonical": anchor["name"],
                    "variants": group,
                    "total_members": total,
                })

    # Pass 2 — cross-governorate exact (normalised) matches.
    if cross_governorate:
        for norm_key, occurrences in by_normalised.items():
            unique_govs = {o["governorate"] for o in occurrences}
            if len(unique_govs) < 2:
                continue
            # Pick the most-frequent (governorate, name) pair as the suggested
            # canonical. The variant LIST keeps every (gov, name) combination
            # so the operator can pick which governorate is correct.
            ordered = sorted(occurrences, key=lambda x: (-x["count"], x["governorate"]))
            anchor = ordered[0]
            variants = [{"name": o["name"], "count": o["count"], "governorate": o["governorate"]} for o in ordered]
            clusters.append({
                "kind": "cross_governorate_same_name",
                "governorate": anchor["governorate"],
                "canonical": anchor["name"],
                "variants": variants,
                "total_members": sum(o["count"] for o in occurrences),
            })

    clusters.sort(key=lambda c: -c["total_members"])
    return {
        "department_id": department_id,
        "threshold": threshold,
        "clusters_count": len(clusters),
        "clusters": clusters,
        "stats": {
            "distinct_committees": total_committees,
            "governorates_scanned": len(by_gov),
            "pairs_compared": pairs_checked,
        },
    }


class CommitteeMergeOp(BaseModel):
    governorate: str
    canonical: str
    aliases: List[str]


class CommitteeMergeRequest(BaseModel):
    department_id: str
    operations: List[CommitteeMergeOp]


@api_router.post("/admin/committees/merge")
async def merge_duplicate_committees(
    payload: CommitteeMergeRequest,
    user: Dict[str, Any] = Depends(require_admin),
):
    """Apply merge operations: rename `aliases` → `canonical` within a governorate.

    Updates EVERY collection that stores `union_committee` so the rename is
    consistent across the whole app — otherwise financial reports would still
    surface the obsolete names. Collections touched: members, subscriptions,
    aids, dues_settlements.
    """
    if not payload.department_id:
        raise HTTPException(status_code=400, detail="department_id مطلوب")
    results: List[Dict[str, Any]] = []
    total_updated = 0
    related_collections = (
        ("members", db.members),
        ("subscriptions", db.subscriptions),
        ("aids", db.aids),
        ("dues_settlements", db.dues_settlements),
    )
    for op in payload.operations:
        if not op.governorate or not op.canonical:
            continue
        aliases = [a for a in op.aliases if a and a != op.canonical]
        if not aliases:
            results.append({
                "governorate": op.governorate,
                "canonical": op.canonical,
                "updated": 0,
                "skipped": True,
            })
            continue
        per_collection: Dict[str, int] = {}
        op_total = 0
        for col_name, col in related_collections:
            # Aids store member_governorate/member_union_committee snapshots.
            if col_name == "aids":
                filt = {
                    "department_id": payload.department_id,
                    "member_governorate": op.governorate,
                    "member_union_committee": {"$in": aliases},
                }
                update = {"$set": {"member_union_committee": op.canonical, "updated_at": now_iso()}}
            else:
                filt = {
                    "department_id": payload.department_id,
                    "governorate": op.governorate,
                    "union_committee": {"$in": aliases},
                }
                update = {"$set": {"union_committee": op.canonical, "updated_at": now_iso()}}
            res = await col.update_many(filt, update)
            per_collection[col_name] = res.modified_count
            op_total += res.modified_count
        total_updated += op_total
        results.append({
            "governorate": op.governorate,
            "canonical": op.canonical,
            "aliases": aliases,
            "updated": op_total,
            "per_collection": per_collection,
        })
    # Tell every connected client to refresh their lists.
    if total_updated:
        for col_name in ("members", "subscriptions", "aids", "dues_settlements"):
            try:
                await _bump_change_counter(col_name)
            except Exception:  # noqa: BLE001
                pass
    return {"ok": True, "total_updated": total_updated, "results": results}


# ─── Admin: detect & merge duplicate / near-duplicate MEMBERS ────────────────
# Operates inside a (normalised governorate, normalised committee) group so
# committee typos don't split a true person-level duplicate. Returns clusters
# of members with confidence signals; the frontend lets the operator pick the
# canonical and confirm the merge — auto-merge is intentionally NOT supported.

class MemberMergeOp(BaseModel):
    canonical_id: str
    alias_ids: List[str] = Field(default_factory=list)
    # Optional canonical overrides — when the operator wants to fix a typo
    # in the surviving member's governorate / committee at the same time.
    canonical_governorate: Optional[str] = None
    canonical_union_committee: Optional[str] = None


class MemberMergeRequest(BaseModel):
    department_id: str
    operations: List[MemberMergeOp] = Field(default_factory=list)


def _normalise_digits(value: str) -> str:
    """Convert Arabic-Indic and Persian digits to ASCII, strip spaces.
    Two membership numbers like ٤٣٧ and 437 collide as the same value."""
    if not value:
        return ""
    table = str.maketrans(
        "٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹",
        "01234567890123456789",
    )
    return value.translate(table).strip()


def _is_meaningful_id(value: str, min_len: int = 4) -> bool:
    """Reject sentinel / placeholder ids that shouldn't be used for matching.

    Many imported rows have national_id like "" or "0" or "00" — treating
    those as equal would put thousands of unrelated members into one cluster.
    We require a meaningful-length numeric string with at least two distinct
    digits.
    """
    s = _normalise_digits(value)
    if not s or len(s) < min_len:
        return False
    if set(s) == {"0"}:
        return False
    return True


def _member_match_score(a: Dict[str, Any], b: Dict[str, Any]) -> Tuple[float, List[str]]:
    """Return (score, [signals]) for two members in the same gov+committee.

    Higher score = higher confidence they're the same person. Signals are
    human-readable reasons exposed in the UI so the operator can audit.

    Tiered logic (we take the BEST signal, not a sum). Each signal first
    confirms that the candidate identifier is *meaningful* — empty strings
    and sentinels like "0" or "00" are NOT treated as equal so a column of
    "missing" rows never collapses into one false-positive cluster.

      • Identical national_id (≥4 chars, not all zeros) → 1.00
      • Same membership_number + similar name (≥0.6)    → 0.95
      • Same membership_number + similar name (≥0.85)   → 0.90
      • Similar name (≥0.85) + same birth_date          → 0.92
      • Very similar name (≥0.94) alone                 → 0.80
    """
    signals: List[str] = []
    score = 0.0
    nid_a = _normalise_digits(a.get("national_id") or "")
    nid_b = _normalise_digits(b.get("national_id") or "")
    mn_a = _normalise_digits(a.get("membership_number") or "")
    mn_b = _normalise_digits(b.get("membership_number") or "")
    name_a = a.get("name") or ""
    name_b = b.get("name") or ""
    name_sim = _similarity(name_a, name_b) if name_a and name_b else 0.0
    bd_a = (a.get("birth_date") or "").strip()
    bd_b = (b.get("birth_date") or "").strip()

    if nid_a == nid_b and _is_meaningful_id(nid_a, min_len=4):
        score = max(score, 1.0)
        signals.append("نفس الرقم القومي")
    if mn_a == mn_b and _is_meaningful_id(mn_a, min_len=1):
        # Same membership_number is only useful when paired with another
        # signal — many committees re-use short membership numbers (1, 2, …)
        # across unrelated members, so we *require* a similar name.
        if name_sim >= 0.85:
            score = max(score, 0.95)
            signals.append(f"نفس رقم العضوية + اسم متشابه ({int(name_sim * 100)}%)")
        elif name_sim >= 0.6 and bd_a and bd_a == bd_b:
            score = max(score, 0.92)
            signals.append(f"نفس رقم العضوية + اسم متشابه ({int(name_sim * 100)}%) + نفس تاريخ الميلاد")
    if name_sim >= 0.85 and bd_a and bd_a == bd_b:
        score = max(score, 0.92)
        sig = f"اسم متشابه ({int(name_sim * 100)}%) + نفس تاريخ الميلاد"
        if sig not in signals:
            signals.append(sig)
    if name_sim >= 0.94 and not signals:
        score = max(score, 0.80)
        signals.append(f"اسم متطابق ({int(name_sim * 100)}%)")
    return score, signals


@api_router.get("/admin/members/duplicates")
async def detect_duplicate_members(
    department_id: str,
    threshold: float = 0.85,
    user: Dict[str, Any] = Depends(require_admin),
):
    """Scan every (governorate, committee) bucket and return clusters of
    members likely to be the same person.

    Governorate and committee are NORMALISED for grouping so that committee
    typos ("التأمين" vs "التامين") don't artificially split a real
    person-level duplicate. For each cluster we also report the most
    frequent original spellings so the operator can fix everything in one
    pass via the merge endpoint.
    """
    members = await db.members.find(
        {"department_id": department_id},
        {
            "_id": 0, "id": 1, "name": 1, "national_id": 1, "membership_number": 1,
            "governorate": 1, "union_committee": 1, "birth_date": 1, "phone": 1,
            "subscription_date": 1, "status": 1, "status_date": 1, "address": 1,
            "pdf_url": 1, "created_at": 1,
        },
    ).to_list(None)

    # Bucket by (normalised gov, normalised committee).
    buckets: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for m in members:
        gov = (m.get("governorate") or "").strip()
        com = (m.get("union_committee") or "").strip()
        if not gov or not com:
            continue
        key = (_normalise_ar(gov), _normalise_ar(com))
        buckets[key].append(m)

    clusters: List[Dict[str, Any]] = []
    pairs_checked = 0

    for (norm_gov, norm_com), bucket in buckets.items():
        if len(bucket) < 2:
            continue
        # Union-find so transitively-linked members end up in one cluster
        # (e.g. A↔B by national_id, B↔C by membership_number → {A, B, C}).
        parent = list(range(len(bucket)))
        def find(i: int) -> int:
            while parent[i] != i:
                parent[i] = parent[parent[i]]
                i = parent[i]
            return i
        def union(i: int, j: int) -> None:
            ri, rj = find(i), find(j)
            if ri != rj:
                parent[ri] = rj
        edge_signals: Dict[Tuple[int, int], List[str]] = {}
        edge_scores: Dict[Tuple[int, int], float] = {}
        for i in range(len(bucket)):
            for j in range(i + 1, len(bucket)):
                pairs_checked += 1
                s, sigs = _member_match_score(bucket[i], bucket[j])
                if s >= threshold:
                    union(i, j)
                    edge_scores[(i, j)] = s
                    edge_signals[(i, j)] = sigs

        # Group by their root.
        groups: Dict[int, List[int]] = defaultdict(list)
        for idx in range(len(bucket)):
            groups[find(idx)].append(idx)

        for root, indices in groups.items():
            if len(indices) < 2:
                continue
            members_in_cluster = [bucket[i] for i in indices]
            # Pick a canonical: the oldest subscription_date wins; fall back
            # to most-complete data, then by created_at ascending.
            def canonical_sort_key(m: Dict[str, Any]) -> Tuple[str, int, str]:
                sd = (m.get("subscription_date") or "9999-99-99").strip() or "9999-99-99"
                completeness = sum(1 for f in ("name", "national_id", "membership_number", "birth_date", "phone", "address") if (m.get(f) or "").strip())
                # Negate completeness so MORE complete → "smaller" → wins.
                return (sd, -completeness, m.get("created_at") or "")
            ranked = sorted(members_in_cluster, key=canonical_sort_key)
            canonical = ranked[0]

            # Collect per-pair signals for display ("why these are matched").
            cluster_signals: List[str] = []
            cluster_max_score = 0.0
            seen: Set[str] = set()
            for (i, j), sigs in edge_signals.items():
                if find(i) == root:
                    cluster_max_score = max(cluster_max_score, edge_scores.get((i, j), 0.0))
                    for s in sigs:
                        if s not in seen:
                            cluster_signals.append(s)
                            seen.add(s)

            # Surface the most common original spellings inside this cluster
            # so the merge endpoint can also fix the gov/committee in one shot.
            gov_counts = Counter((m.get("governorate") or "").strip() for m in members_in_cluster if (m.get("governorate") or "").strip())
            com_counts = Counter((m.get("union_committee") or "").strip() for m in members_in_cluster if (m.get("union_committee") or "").strip())
            suggested_gov = gov_counts.most_common(1)[0][0] if gov_counts else ""
            suggested_com = com_counts.most_common(1)[0][0] if com_counts else ""

            clusters.append({
                "governorate_normalised": norm_gov,
                "committee_normalised": norm_com,
                "suggested_governorate": suggested_gov,
                "suggested_committee": suggested_com,
                "governorate_variants": [{"name": n, "count": c} for n, c in gov_counts.items()] if len(gov_counts) > 1 else [],
                "committee_variants": [{"name": n, "count": c} for n, c in com_counts.items()] if len(com_counts) > 1 else [],
                "canonical_id": canonical["id"],
                "max_score": round(cluster_max_score, 2),
                "signals": cluster_signals,
                "members": members_in_cluster,
            })

    # Highest-confidence clusters first, then biggest.
    clusters.sort(key=lambda c: (-c["max_score"], -len(c["members"])))

    return {
        "department_id": department_id,
        "threshold": threshold,
        "clusters_count": len(clusters),
        "clusters": clusters,
        "stats": {
            "members_scanned": len(members),
            "buckets_scanned": len(buckets),
            "pairs_compared": pairs_checked,
        },
    }


@api_router.post("/admin/members/merge")
async def merge_duplicate_members(
    payload: MemberMergeRequest,
    user: Dict[str, Any] = Depends(require_admin),
):
    """Consolidate alias members into a canonical record.

    For each operation:
      1) Move every related record (subscriptions / aids / dues_settlements)
         that points at an alias → re-point at the canonical.
      2) Refresh the aids' denormalised member_* snapshot fields from the
         canonical (so previously stored governorate/committee labels become
         consistent across the entire archive).
      3) Optionally rewrite the canonical's governorate / committee to a
         corrected spelling that the operator picked while reviewing.
      4) Delete the alias member documents.

    Returns the count of records re-assigned and members removed.
    """
    if not payload.department_id:
        raise HTTPException(status_code=400, detail="department_id مطلوب")
    summary = {"members_deleted": 0, "subscriptions_reassigned": 0, "aids_reassigned": 0, "dues_settlements_reassigned": 0, "canonical_updated": 0}
    details: List[Dict[str, Any]] = []
    for op in payload.operations:
        canonical = await db.members.find_one({"id": op.canonical_id, "department_id": payload.department_id}, {"_id": 0})
        if not canonical:
            details.append({"canonical_id": op.canonical_id, "skipped": True, "reason": "canonical_not_found"})
            continue
        alias_ids = [aid for aid in op.alias_ids if aid and aid != op.canonical_id]
        if not alias_ids:
            details.append({"canonical_id": op.canonical_id, "skipped": True, "reason": "no_aliases"})
            continue

        # Step 3 (run first so the snapshot refresh below uses the new gov/com).
        canonical_patch: Dict[str, Any] = {}
        new_gov = (op.canonical_governorate or "").strip()
        new_com = (op.canonical_union_committee or "").strip()
        if new_gov and new_gov != canonical.get("governorate"):
            canonical_patch["governorate"] = new_gov
        if new_com and new_com != canonical.get("union_committee"):
            canonical_patch["union_committee"] = new_com
        if canonical_patch:
            canonical_patch["updated_at"] = now_iso()
            await db.members.update_one({"id": op.canonical_id}, {"$set": canonical_patch})
            canonical = {**canonical, **canonical_patch}
            summary["canonical_updated"] += 1

        # Step 1 — reassign related records to the canonical.
        sub_res = await db.subscriptions.update_many(
            {"department_id": payload.department_id, "member_id": {"$in": alias_ids}},
            {"$set": {"member_id": op.canonical_id, "updated_at": now_iso()}},
        )
        aid_res = await db.aids.update_many(
            {"department_id": payload.department_id, "member_id": {"$in": alias_ids}},
            {"$set": {
                "member_id": op.canonical_id,
                "member_name": canonical.get("name", ""),
                "member_national_id": canonical.get("national_id", ""),
                "member_membership_number": canonical.get("membership_number", ""),
                "member_governorate": canonical.get("governorate", ""),
                "member_union_committee": canonical.get("union_committee", ""),
                "member_birth_date": canonical.get("birth_date", ""),
                "member_subscription_date": canonical.get("subscription_date", ""),
                "member_address": canonical.get("address", ""),
                "member_phone": canonical.get("phone", ""),
                "member_status_date": canonical.get("status_date", ""),
                "updated_at": now_iso(),
            }},
        )
        dues_res = await db.dues_settlements.update_many(
            {"department_id": payload.department_id, "member_id": {"$in": alias_ids}},
            {"$set": {"member_id": op.canonical_id, "updated_at": now_iso()}},
        )

        # Step 4 — delete the aliases.
        del_res = await db.members.delete_many(
            {"id": {"$in": alias_ids}, "department_id": payload.department_id},
        )

        summary["subscriptions_reassigned"] += sub_res.modified_count
        summary["aids_reassigned"] += aid_res.modified_count
        summary["dues_settlements_reassigned"] += dues_res.modified_count
        summary["members_deleted"] += del_res.deleted_count
        details.append({
            "canonical_id": op.canonical_id,
            "aliases_removed": del_res.deleted_count,
            "subscriptions_reassigned": sub_res.modified_count,
            "aids_reassigned": aid_res.modified_count,
            "dues_settlements_reassigned": dues_res.modified_count,
        })

    if summary["members_deleted"] or summary["subscriptions_reassigned"] or summary["aids_reassigned"]:
        for col_name in ("members", "subscriptions", "aids", "dues_settlements"):
            try:
                await _bump_change_counter(col_name)
            except Exception:  # noqa: BLE001
                pass
    return {"ok": True, "summary": summary, "details": details}


# ─── Admin: detect members with MISSING / INCOMPLETE data ────────────────────
# Same UX pattern as the duplicate-detection feature, but instead of grouping
# similar people it flags individual records that have empty critical fields
# (national_id / membership_number / birth_date / address / phone) so the
# operator can complete them inline.

_BAD_LABEL_PATTERN = re.compile(r"^[\d/\-\.\s:]+$")


def _missing_fields_for_member(m: Dict[str, Any]) -> List[Dict[str, str]]:
    """Return a list of {key, label} for every critical field that's empty or
    looks like a placeholder (e.g. someone pasted a date into the governorate
    column). Each entry is rendered as a fixable badge in the UI."""
    checks: List[Tuple[str, str, bool]] = [
        ("name", "الاسم", bool((m.get("name") or "").strip())),
        ("national_id", "الرقم القومي", _is_meaningful_id(m.get("national_id") or "", min_len=4)),
        ("membership_number", "رقم العضوية", bool((m.get("membership_number") or "").strip())),
        ("governorate", "المحافظة", bool((m.get("governorate") or "").strip()) and not _BAD_LABEL_PATTERN.match((m.get("governorate") or "").strip())),
        ("union_committee", "اللجنة النقابية", bool((m.get("union_committee") or "").strip()) and not _BAD_LABEL_PATTERN.match((m.get("union_committee") or "").strip())),
        ("birth_date", "تاريخ الميلاد", bool((m.get("birth_date") or "").strip())),
        ("subscription_date", "تاريخ الاشتراك", bool((m.get("subscription_date") or "").strip())),
        ("address", "العنوان", bool((m.get("address") or "").strip())),
        ("phone", "رقم التليفون", bool((m.get("phone") or "").strip())),
    ]
    return [{"key": k, "label": label} for (k, label, ok) in checks if not ok]


@api_router.get("/admin/members/missing-data")
async def detect_members_missing_data(
    department_id: str,
    min_missing: int = 1,
    user: Dict[str, Any] = Depends(require_admin),
):
    """Scan every member in a department and return those with at least
    `min_missing` empty / placeholder critical fields.

    Result is grouped by (governorate, committee) for easier review and
    sorted by how many fields are missing (most-incomplete first).
    """
    cursor = db.members.find(
        {"department_id": department_id},
        {
            "_id": 0, "id": 1, "name": 1, "national_id": 1, "membership_number": 1,
            "governorate": 1, "union_committee": 1, "birth_date": 1, "subscription_date": 1,
            "phone": 1, "address": 1, "status": 1, "status_date": 1, "pdf_url": 1,
        },
    )
    members = await cursor.to_list(None)

    flagged: List[Dict[str, Any]] = []
    field_counts: Counter = Counter()
    for m in members:
        missing = _missing_fields_for_member(m)
        if len(missing) < max(1, min_missing):
            continue
        for f in missing:
            field_counts[f["label"]] += 1
        flagged.append({
            **m,
            "missing": missing,
            "missing_count": len(missing),
        })

    flagged.sort(key=lambda r: (-r["missing_count"], r.get("name", "")))
    # Field-level summary so the UI can show "X members missing birth_date".
    summary_by_field = [{"field": label, "count": cnt} for label, cnt in field_counts.most_common()]
    return {
        "department_id": department_id,
        "min_missing": min_missing,
        "total_scanned": len(members),
        "total_flagged": len(flagged),
        "summary_by_field": summary_by_field,
        "members": flagged,
    }


# ─── Admin: manually add a governorate / committee to the taxonomy ────────────
# Governorates and committees were previously DERIVED from the members
# collection only — you couldn't pre-create one. We now maintain a small
# `taxonomy` collection so operators can register a new entry BEFORE any
# member is added under it. The /classifications endpoint merges both.

class TaxonomyAddGovernorateRequest(BaseModel):
    department_id: str
    name: str


class TaxonomyAddCommitteeRequest(BaseModel):
    department_id: str
    governorate: str
    name: str


async def _existing_governorates(department_id: str) -> Set[str]:
    """All governorates currently known for a department (members + taxonomy)
    normalised for de-dup comparison."""
    out: Set[str] = set()
    async for d in db.members.aggregate([
        {"$match": {"department_id": department_id}},
        {"$group": {"_id": "$governorate"}},
    ]):
        v = (d["_id"] or "").strip()
        if v:
            out.add(_normalise_ar(v))
    async for d in db.taxonomy.find({"department_id": department_id, "kind": "governorate"}, {"_id": 0, "name": 1}):
        v = (d.get("name") or "").strip()
        if v:
            out.add(_normalise_ar(v))
    return out


async def _existing_committees(department_id: str, governorate: str) -> Set[str]:
    norm_gov = _normalise_ar(governorate)
    out: Set[str] = set()
    async for d in db.members.aggregate([
        {"$match": {"department_id": department_id}},
        {"$group": {"_id": {"g": "$governorate", "c": "$union_committee"}}},
    ]):
        g = (d["_id"].get("g") or "").strip()
        c = (d["_id"].get("c") or "").strip()
        if c and _normalise_ar(g) == norm_gov:
            out.add(_normalise_ar(c))
    async for d in db.taxonomy.find({"department_id": department_id, "kind": "committee"}, {"_id": 0, "governorate": 1, "name": 1}):
        if _normalise_ar(d.get("governorate") or "") == norm_gov:
            out.add(_normalise_ar(d.get("name") or ""))
    return out


@api_router.post("/admin/taxonomy/governorate")
async def add_governorate(payload: TaxonomyAddGovernorateRequest, user: Dict[str, Any] = Depends(require_admin)):
    name = (payload.name or "").strip()
    if not name or _BAD_LABEL_PATTERN.match(name):
        raise HTTPException(status_code=400, detail="اسم محافظة غير صالح")
    if _normalise_ar(name) in await _existing_governorates(payload.department_id):
        raise HTTPException(status_code=409, detail=f"المحافظة \"{name}\" موجودة بالفعل")
    doc = {
        "id": str(uuid.uuid4()),
        "department_id": payload.department_id,
        "kind": "governorate",
        "name": name,
        "created_at": now_iso(),
    }
    await db.taxonomy.insert_one(doc)
    await _bump_change_counter("taxonomy")
    return {"ok": True, "id": doc["id"], "name": name}


@api_router.post("/admin/taxonomy/committee")
async def add_committee(payload: TaxonomyAddCommitteeRequest, user: Dict[str, Any] = Depends(require_admin)):
    gov = (payload.governorate or "").strip()
    name = (payload.name or "").strip()
    if not gov:
        raise HTTPException(status_code=400, detail="المحافظة مطلوبة")
    if not name or _BAD_LABEL_PATTERN.match(name):
        raise HTTPException(status_code=400, detail="اسم لجنة غير صالح")
    if _normalise_ar(name) in await _existing_committees(payload.department_id, gov):
        raise HTTPException(status_code=409, detail=f"اللجنة \"{name}\" موجودة بالفعل في {gov}")
    doc = {
        "id": str(uuid.uuid4()),
        "department_id": payload.department_id,
        "kind": "committee",
        "governorate": gov,
        "name": name,
        "created_at": now_iso(),
    }
    await db.taxonomy.insert_one(doc)
    await _bump_change_counter("taxonomy")
    return {"ok": True, "id": doc["id"], "governorate": gov, "name": name}


# ─── Admin: rename a governorate or committee ─────────────────────────────────
# Cascades the new name across every collection that stores the value:
# members, subscriptions, aids (snapshot fields), dues_settlements, and the
# taxonomy collection itself. Used when an operator typed a wrong name and
# the duplicate-merge tool can't help (because the wrong name has no twin).

class TaxonomyRenameGovernorateRequest(BaseModel):
    department_id: str
    old_name: str
    new_name: str


class TaxonomyRenameCommitteeRequest(BaseModel):
    department_id: str
    governorate: str
    old_name: str
    new_name: str


@api_router.post("/admin/taxonomy/governorate/rename")
async def rename_governorate(
    payload: TaxonomyRenameGovernorateRequest,
    user: Dict[str, Any] = Depends(require_admin),
):
    old = (payload.old_name or "").strip()
    new = (payload.new_name or "").strip()
    if not payload.department_id or not old or not new:
        raise HTTPException(status_code=400, detail="department_id والاسم القديم والجديد مطلوبين")
    if _BAD_LABEL_PATTERN.match(new):
        raise HTTPException(status_code=400, detail="اسم المحافظة الجديد غير صالح")
    if _normalise_ar(old) == _normalise_ar(new):
        raise HTTPException(status_code=400, detail="الاسم الجديد مطابق للاسم القديم")
    # Block silent merge into an existing governorate.
    existing = await _existing_governorates(payload.department_id)
    if _normalise_ar(new) in existing and _normalise_ar(new) != _normalise_ar(old):
        raise HTTPException(status_code=409, detail=f"المحافظة \"{new}\" موجودة بالفعل — استخدم أداة الدمج بدلاً من إعادة التسمية")
    per: Dict[str, int] = {}
    for col_name, col, filt_field, set_field in (
        ("members", db.members, "governorate", "governorate"),
        ("subscriptions", db.subscriptions, "governorate", "governorate"),
        ("aids", db.aids, "member_governorate", "member_governorate"),
        ("dues_settlements", db.dues_settlements, "governorate", "governorate"),
    ):
        res = await col.update_many(
            {"department_id": payload.department_id, filt_field: old},
            {"$set": {set_field: new, "updated_at": now_iso()}},
        )
        per[col_name] = res.modified_count
    tax_res = await db.taxonomy.update_many(
        {"department_id": payload.department_id, "kind": "governorate", "name": old},
        {"$set": {"name": new}},
    )
    tax_com_res = await db.taxonomy.update_many(
        {"department_id": payload.department_id, "kind": "committee", "governorate": old},
        {"$set": {"governorate": new}},
    )
    per["taxonomy"] = tax_res.modified_count + tax_com_res.modified_count
    await _bump_change_counter("taxonomy")
    await _bump_change_counter("members")
    return {"ok": True, "old_name": old, "new_name": new, "per_collection": per, "total": sum(per.values())}


@api_router.post("/admin/taxonomy/committee/rename")
async def rename_committee(
    payload: TaxonomyRenameCommitteeRequest,
    user: Dict[str, Any] = Depends(require_admin),
):
    gov = (payload.governorate or "").strip()
    old = (payload.old_name or "").strip()
    new = (payload.new_name or "").strip()
    if not payload.department_id or not gov or not old or not new:
        raise HTTPException(status_code=400, detail="department_id والمحافظة والاسم القديم والجديد مطلوبين")
    if _BAD_LABEL_PATTERN.match(new):
        raise HTTPException(status_code=400, detail="اسم اللجنة الجديد غير صالح")
    if _normalise_ar(old) == _normalise_ar(new):
        raise HTTPException(status_code=400, detail="الاسم الجديد مطابق للاسم القديم")
    existing = await _existing_committees(payload.department_id, gov)
    if _normalise_ar(new) in existing and _normalise_ar(new) != _normalise_ar(old):
        raise HTTPException(status_code=409, detail=f"اللجنة \"{new}\" موجودة بالفعل في {gov} — استخدم أداة الدمج بدلاً من إعادة التسمية")
    per: Dict[str, int] = {}
    for col_name, col, gov_field, com_field in (
        ("members", db.members, "governorate", "union_committee"),
        ("subscriptions", db.subscriptions, "governorate", "union_committee"),
        ("aids", db.aids, "member_governorate", "member_union_committee"),
        ("dues_settlements", db.dues_settlements, "governorate", "union_committee"),
    ):
        res = await col.update_many(
            {"department_id": payload.department_id, gov_field: gov, com_field: old},
            {"$set": {com_field: new, "updated_at": now_iso()}},
        )
        per[col_name] = res.modified_count
    tax_res = await db.taxonomy.update_many(
        {"department_id": payload.department_id, "kind": "committee", "governorate": gov, "name": old},
        {"$set": {"name": new}},
    )
    per["taxonomy"] = tax_res.modified_count
    await _bump_change_counter("taxonomy")
    await _bump_change_counter("members")
    return {"ok": True, "governorate": gov, "old_name": old, "new_name": new, "per_collection": per, "total": sum(per.values())}




# ─── Admin: merge governorate / committee into ANOTHER existing one ───────────
# Differs from `rename` in that the target name MUST already exist. Both
# names then collapse into a single one across every collection.

class TaxonomyMergeGovernorateRequest(BaseModel):
    department_id: str
    source: str
    target: str


class TaxonomyMergeCommitteeRequest(BaseModel):
    department_id: str
    source_governorate: str
    source_committee: str
    target_governorate: str
    target_committee: str


@api_router.post("/admin/taxonomy/governorate/merge")
async def merge_governorate(
    payload: TaxonomyMergeGovernorateRequest,
    user: Dict[str, Any] = Depends(require_admin),
):
    src = (payload.source or "").strip()
    dst = (payload.target or "").strip()
    if not payload.department_id or not src or not dst:
        raise HTTPException(status_code=400, detail="department_id والمصدر والهدف مطلوبين")
    if _normalise_ar(src) == _normalise_ar(dst):
        raise HTTPException(status_code=400, detail="المصدر والهدف متطابقان")
    per: Dict[str, int] = {}
    for col_name, col, filt_field, set_field in (
        ("members", db.members, "governorate", "governorate"),
        ("subscriptions", db.subscriptions, "governorate", "governorate"),
        ("aids", db.aids, "member_governorate", "member_governorate"),
        ("dues_settlements", db.dues_settlements, "governorate", "governorate"),
    ):
        res = await col.update_many(
            {"department_id": payload.department_id, filt_field: src},
            {"$set": {set_field: dst, "updated_at": now_iso()}},
        )
        per[col_name] = res.modified_count
    # Delete the source governorate from the taxonomy collection (it is being
    # merged away), and move any committees registered under it.
    del_res = await db.taxonomy.delete_many(
        {"department_id": payload.department_id, "kind": "governorate", "name": src}
    )
    com_move = await db.taxonomy.update_many(
        {"department_id": payload.department_id, "kind": "committee", "governorate": src},
        {"$set": {"governorate": dst}},
    )
    per["taxonomy"] = del_res.deleted_count + com_move.modified_count
    await _bump_change_counter("taxonomy")
    await _bump_change_counter("members")
    return {"ok": True, "source": src, "target": dst, "per_collection": per, "total": sum(per.values())}


@api_router.post("/admin/taxonomy/committee/merge")
async def merge_committee_into(
    payload: TaxonomyMergeCommitteeRequest,
    user: Dict[str, Any] = Depends(require_admin),
):
    sg = (payload.source_governorate or "").strip()
    sc = (payload.source_committee or "").strip()
    dg = (payload.target_governorate or "").strip()
    dc = (payload.target_committee or "").strip()
    if not (payload.department_id and sg and sc and dg and dc):
        raise HTTPException(status_code=400, detail="كل الحقول مطلوبة")
    if _normalise_ar(sg) == _normalise_ar(dg) and _normalise_ar(sc) == _normalise_ar(dc):
        raise HTTPException(status_code=400, detail="المصدر والهدف متطابقان")
    per: Dict[str, int] = {}
    for col_name, col, gov_field, com_field in (
        ("members", db.members, "governorate", "union_committee"),
        ("subscriptions", db.subscriptions, "governorate", "union_committee"),
        ("aids", db.aids, "member_governorate", "member_union_committee"),
        ("dues_settlements", db.dues_settlements, "governorate", "union_committee"),
    ):
        res = await col.update_many(
            {"department_id": payload.department_id, gov_field: sg, com_field: sc},
            {"$set": {gov_field: dg, com_field: dc, "updated_at": now_iso()}},
        )
        per[col_name] = res.modified_count
    del_res = await db.taxonomy.delete_many(
        {"department_id": payload.department_id, "kind": "committee", "governorate": sg, "name": sc}
    )
    per["taxonomy"] = del_res.deleted_count
    await _bump_change_counter("taxonomy")
    await _bump_change_counter("members")
    return {
        "ok": True,
        "source": {"governorate": sg, "committee": sc},
        "target": {"governorate": dg, "committee": dc},
        "per_collection": per,
        "total": sum(per.values()),
    }


@api_router.get("/health")
async def health_check():
    """Lightweight liveness probe used by the local Windows health-loop runner.

    Returns 200 if the API process is alive AND can reach MongoDB.
    """
    try:
        await db.command("ping")
        return {"ok": True, "db": "up", "ts": now_iso()}
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=503, detail=f"db_unreachable: {e}")


@api_router.get("/changes/state")
async def changes_state(user: Dict[str, Any] = Depends(require_user)):
    """Per-collection version counters used by the live-sync poller.

    The frontend polls this every few seconds; any counter change tells it the
    corresponding list view should be refetched silently. Mutating endpoints
    bump these counters via _bump_change_counter().
    """
    doc = await db["_change_state"].find_one({"key": "counters"}, {"_id": 0}) or {}
    return {"counters": doc.get("counters", {}), "updated_at": doc.get("updated_at", "")}


# ─── Presence: "who is online" indicator ─────────────────────────────────────
# Lightweight heartbeat system so users on the LAN can see who else is
# currently using the app and on which page. MongoDB TTL index on
# `expires_at_dt` auto-prunes stale rows. Each row keyed by user_id.
PRESENCE_TTL_SECONDS = 45  # client beats every 15s; we keep ~3 missed beats.


class PresenceBeat(BaseModel):
    model_config = ConfigDict(extra="ignore")
    path: str = ""
    page_title: str = ""


@api_router.post("/presence/heartbeat")
async def presence_heartbeat(payload: PresenceBeat, user: Dict[str, Any] = Depends(require_user)):
    """Record that the calling user is currently active.

    Called every ~15s from the frontend while a tab is visible. Stores the
    current page so other users can see "Ahmad is editing Subscriptions".
    """
    now = datetime.now(timezone.utc)
    expires = now + timedelta(seconds=PRESENCE_TTL_SECONDS)
    await db.presence.update_one(
        {"user_id": user.get("id")},
        {"$set": {
            "user_id": user.get("id"),
            "username": user.get("username", ""),
            "display_name": user.get("display_name") or user.get("username", ""),
            "role": user.get("role", ""),
            "path": (payload.path or "").strip()[:200],
            "page_title": (payload.page_title or "").strip()[:120],
            "last_seen": now.isoformat(),
            "last_seen_dt": now,
            "expires_at_dt": expires,
        }},
        upsert=True,
    )
    return {"ok": True, "ts": now.isoformat()}


@api_router.post("/presence/logout")
async def presence_logout(user: Dict[str, Any] = Depends(require_user)):
    """Mark the current user as offline immediately (called on logout / tab close)."""
    await db.presence.delete_one({"user_id": user.get("id")})
    return {"ok": True}


@api_router.get("/presence/online")
async def presence_online(user: Dict[str, Any] = Depends(require_user)):
    """Return everyone seen within the active window — including self.

    The frontend filters out the current user when rendering avatars.
    TTL index handles stale row eviction automatically, but we also filter
    by `last_seen_dt` as a defensive measure in case the TTL purge lags.
    """
    cutoff = datetime.now(timezone.utc) - timedelta(seconds=PRESENCE_TTL_SECONDS)
    docs = await db.presence.find(
        {"last_seen_dt": {"$gte": cutoff}},
        {"_id": 0, "expires_at_dt": 0, "last_seen_dt": 0},
    ).to_list(200)
    return {"users": docs, "count": len(docs)}


@api_router.get("/admin/export-all-data")
async def export_all_data(actor: Dict[str, Any] = Depends(require_admin)):
    """Full database snapshot for migrating cloud data to a local install.

    Returns a single JSON document with every collection serialised. The receiving
    side can stream it back into a fresh MongoDB via the companion `restore_data.py`
    script shipped in the deploy/ folder.
    """
    collections = [
        "members", "subscriptions", "aids", "departments", "users",
        "retirement_schedules", "classifications", "letters", "documents",
        "dues_settlements", "financial_letters", "aid_disbursed",
        "user_settings", "scan_jobs",
    ]
    snapshot: Dict[str, Any] = {
        "version": 1,
        "exported_at": now_iso(),
        "collections": {},
    }
    for name in collections:
        try:
            docs = await db[name].find({}, {"_id": 0}).to_list(100000)
            snapshot["collections"][name] = docs
        except Exception as e:  # noqa: BLE001
            snapshot["collections"][name] = {"_error": str(e)}
    return snapshot


# ─── Local Backup / Restore — saved on user's PC, no passwords included ───
#
# Collections that we back up. We intentionally EXCLUDE password material when
# serialising the `users` collection (see _sanitised_user below).
_BACKUP_COLLECTIONS = [
    "members", "subscriptions", "aids", "departments", "users",
    "retirement_schedules", "classifications", "letters", "documents",
    "dues_settlements", "financial_letters", "aid_disbursed",
    "user_settings", "scan_jobs", "audit_log",
]
_PASSWORD_FIELDS = {"password_hash", "salt", "password"}


def _sanitise_doc_for_backup(coll_name: str, doc: Dict[str, Any]) -> Dict[str, Any]:
    """Strip password material before writing a backup."""
    if coll_name != "users":
        return doc
    return {k: v for k, v in doc.items() if k not in _PASSWORD_FIELDS}


@api_router.get("/admin/backup/export")
async def admin_backup_export(actor: Dict[str, Any] = Depends(require_admin)):
    """Return a full JSON snapshot ready to be saved to the user's computer.

    NOTE: password hashes and salts are stripped from the `users` collection so
    the resulting file is safe to keep on a USB stick or share with a colleague
    for disaster-recovery purposes.
    """
    payload: Dict[str, Any] = {
        "kind": "electronic-archive-backup",
        "version": 1,
        "exported_at": now_iso(),
        "exported_by": actor.get("username", ""),
        "collections": {},
    }
    for name in _BACKUP_COLLECTIONS:
        try:
            docs = await db[name].find({}, {"_id": 0}).to_list(200000)
            payload["collections"][name] = [_sanitise_doc_for_backup(name, d) for d in docs]
        except Exception:
            payload["collections"][name] = []
    # Remember the moment we issued a backup so the schedule check can disable
    # the reminder until next year.
    await db.user_settings.update_one(
        {"key": "backup_status"},
        {"$set": {"key": "backup_status", "last_backup_at": payload["exported_at"]}},
        upsert=True,
    )
    return payload


@api_router.post("/admin/backup/restore")
async def admin_backup_restore(
    payload: Dict[str, Any] = Body(...),
    mode: str = "merge",  # merge | replace
    actor: Dict[str, Any] = Depends(require_admin),
):
    """Restore a previously-saved backup JSON.

    - `mode=merge` (default): cloud-side rows from the file are upserted by `id`
      so existing records get updated and missing ones get inserted. Local rows
      that are not in the file are kept.
    - `mode=replace`: each collection is wiped first, then re-populated from the
      file. Use with care.

    Password material is NEVER restored; existing admin/user credentials remain
    untouched. If a user from the backup doesn't exist locally, they are
    inserted with `active: False` so the operator must reset their password
    before they can log in.
    """
    if not isinstance(payload, dict) or payload.get("kind") != "electronic-archive-backup":
        raise HTTPException(status_code=400, detail="ملف النسخة الاحتياطية غير صالح أو تالف.")
    if mode not in ("merge", "replace"):
        raise HTTPException(status_code=400, detail="القيمة المسموح بها لـ mode هي merge أو replace فقط.")

    cols = payload.get("collections") or {}
    summary: Dict[str, Any] = {"mode": mode, "collections": {}}
    for name in _BACKUP_COLLECTIONS:
        docs = cols.get(name) or []
        if not isinstance(docs, list):
            continue
        # Strip any password fields that may have been added externally.
        cleaned = [_sanitise_doc_for_backup(name, d) for d in docs if isinstance(d, dict)]
        if mode == "replace":
            try:
                await db[name].delete_many({})
            except Exception:
                pass
        inserted = updated = skipped = 0
        for doc in cleaned:
            if not doc.get("id"):
                skipped += 1
                continue
            try:
                if name == "users":
                    existing = await db.users.find_one({"id": doc["id"]}, {"_id": 0})
                    if existing:
                        # Keep existing salt/password_hash intact; only update profile.
                        await db.users.update_one(
                            {"id": doc["id"]},
                            {"$set": {k: v for k, v in doc.items() if k not in _PASSWORD_FIELDS}},
                        )
                        updated += 1
                    else:
                        doc["active"] = False  # force a manual password reset
                        await db.users.insert_one({**doc})
                        inserted += 1
                else:
                    existing = await db[name].find_one({"id": doc["id"]}, {"_id": 0})
                    if existing:
                        await db[name].update_one({"id": doc["id"]}, {"$set": doc})
                        updated += 1
                    else:
                        await db[name].insert_one({**doc})
                        inserted += 1
            except Exception as exc:  # noqa: BLE001
                skipped += 1
        summary["collections"][name] = {"inserted": inserted, "updated": updated, "skipped": skipped}
    return summary


@api_router.get("/admin/backup/schedule")
async def admin_backup_schedule_get(actor: Dict[str, Any] = Depends(require_admin)):
    """Read the user's annual backup preference + status of last backup."""
    schedule = await db.user_settings.find_one({"key": "backup_schedule"}, {"_id": 0}) or {}
    status = await db.user_settings.find_one({"key": "backup_status"}, {"_id": 0}) or {}

    last_backup_at = status.get("last_backup_at", "")
    enabled = bool(schedule.get("enabled"))
    month = int(schedule.get("month") or 1)
    day = int(schedule.get("day") or 1)

    # Compute whether a backup is currently due.
    due = False
    if enabled:
        today = datetime.now(timezone.utc).date()
        target_this_year = today.replace(month=month, day=min(day, 28))
        last_dt = None
        if last_backup_at:
            try:
                last_dt = datetime.fromisoformat(last_backup_at).date()
            except Exception:
                last_dt = None
        # Due if we've passed (or hit) this year's anniversary AND last backup is
        # either missing or older than the anniversary.
        if today >= target_this_year and (last_dt is None or last_dt < target_this_year):
            due = True
    return {
        "enabled": enabled,
        "month": month,
        "day": day,
        "last_backup_at": last_backup_at,
        "due": due,
    }


@api_router.put("/admin/backup/schedule")
async def admin_backup_schedule_set(
    payload: Dict[str, Any] = Body(...),
    actor: Dict[str, Any] = Depends(require_admin),
):
    """Save the user's annual backup preference.

    Body: { enabled: bool, month: 1-12, day: 1-31 }
    """
    enabled = bool(payload.get("enabled"))
    month = int(payload.get("month") or 1)
    day = int(payload.get("day") or 1)
    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="الشهر يجب أن يكون بين 1 و 12.")
    if not (1 <= day <= 31):
        raise HTTPException(status_code=400, detail="اليوم يجب أن يكون بين 1 و 31.")
    await db.user_settings.update_one(
        {"key": "backup_schedule"},
        {"$set": {"key": "backup_schedule", "enabled": enabled, "month": month, "day": day,
                  "updated_at": now_iso(), "updated_by": actor.get("username", "")}},
        upsert=True,
    )
    return {"ok": True, "enabled": enabled, "month": month, "day": day}



@api_router.get("/installer/download")
async def installer_download():
    """Public endpoint: returns a ZIP with backend + frontend source + deploy scripts.
    Used by setup.bat on the user's local Windows machine to install/update the app.
    """
    import io, zipfile
    repo_root = Path(__file__).resolve().parent.parent  # /app
    buf = io.BytesIO()
    # Folders to include from /app
    INCLUDE_DIRS = ["backend", "frontend", "deploy"]
    # Path patterns to skip (substring match on relative path)
    SKIP_PATTERNS = (
        "__pycache__", ".pytest_cache", "node_modules",
        ".next", ".cache", ".turbo", ".git",
        "backend/storage", "backend/.env",
        "frontend/.env", ".log",
        "frontend/src", "frontend/public", "frontend/.gitignore",
        "frontend/yarn.lock", "frontend/package-lock.json",
        "frontend/postcss.config", "frontend/tailwind.config",
        "frontend/components.json", "frontend/jsconfig.json",
        "frontend/craco.config",
    )
    def _skip(rel: str) -> bool:
        rel_norm = rel.replace("\\", "/")
        return any(p in rel_norm for p in SKIP_PATTERNS)

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for d in INCLUDE_DIRS:
            base = repo_root / d
            if not base.exists():
                continue
            for p in base.rglob("*"):
                if not p.is_file():
                    continue
                rel = p.relative_to(repo_root).as_posix()
                if _skip(rel):
                    continue
                try:
                    zf.write(p, arcname=rel)
                except Exception:
                    continue
        # README at top of zip
        readme_path = repo_root / "deploy" / "README_AR.md"
        if readme_path.exists():
            zf.write(readme_path, arcname="README.md")
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=ElectronicArchive.zip"},
    )


# =========================================================================
# Programmer credits / IP stamp — visible only to super_admin via UI.
# Backend stamp is hardcoded for legal-grade traceability. The signature image
# lives in backend/storage/assets/signature.jpeg (preserved across updates).
# =========================================================================

PROGRAMMER_CREDITS: Dict[str, Any] = {
    "name_en": "Youssef Abdel Ghane Ahmed",
    "name_ar": "يوسف عبد الغني أحمد",
    "national_id": "28611250103535",
    "mobile": "01151177324",
    "email": "zouma.cloner@gmail.com",
    "signature_url": "/api/credits/signature",
    "copyright_year": datetime.now(timezone.utc).year,
}


@api_router.get("/credits")
async def get_programmer_credits(user: Dict[str, Any] = Depends(require_super_admin)):
    """Returns programmer copyright stamp. Visible to super_admin only."""
    return PROGRAMMER_CREDITS


@api_router.get("/credits/signature")
async def get_programmer_signature(user: Dict[str, Any] = Depends(require_super_admin)):
    """Serves the programmer's personal signature image (super_admin only)."""
    sig_path = STORAGE_DIR / "assets" / "signature.jpeg"
    if not sig_path.exists():
        raise HTTPException(status_code=404, detail="صورة التوقيع غير موجودة")
    return FileResponse(str(sig_path), media_type="image/jpeg")




@api_router.get("/installer/install.ps1", response_class=HTMLResponse)
async def installer_install_ps1():
    """Public endpoint: returns the PowerShell one-shot installer/updater.
    Usage from PowerShell (Admin):
       iwr "<server>/api/installer/install.ps1" | iex
    """
    public_url = os.environ.get("PUBLIC_BACKEND_URL", "").rstrip("/")
    if not public_url:
        try:
            env_path = Path(__file__).resolve().parent.parent / "frontend" / ".env"
            for line in env_path.read_text().splitlines():
                if line.startswith("REACT_APP_BACKEND_URL="):
                    public_url = line.split("=", 1)[1].strip().rstrip("/")
                    break
        except Exception:
            public_url = ""
    public_url = public_url or "https://member-scan-test.preview.emergentagent.com"
    ps1_path = Path(__file__).resolve().parent.parent / "deploy" / "install.ps1"
    if not ps1_path.exists():
        raise HTTPException(status_code=500, detail="install.ps1 not found on server")
    script = ps1_path.read_text(encoding="utf-8").replace("__SERVER_URL__", public_url)
    return Response(
        content=script.encode("utf-8"),
        media_type="text/plain; charset=utf-8",
        headers={"Cache-Control": "no-store"},
    )


@api_router.get("/installer/setup.bat", response_class=HTMLResponse)
@api_router.get("/installer/quick.bat", response_class=HTMLResponse)
async def installer_setup_bat():
    """Public endpoint: returns the one-shot Windows installer.
    Reads from /app/deploy/quick_start.bat for easier maintenance.
    """
    public_url = os.environ.get("PUBLIC_BACKEND_URL", "").rstrip("/")
    if not public_url:
        try:
            env_path = Path(__file__).resolve().parent.parent / "frontend" / ".env"
            for line in env_path.read_text().splitlines():
                if line.startswith("REACT_APP_BACKEND_URL="):
                    public_url = line.split("=", 1)[1].strip().rstrip("/")
                    break
        except Exception:
            public_url = ""
    public_url = public_url or "https://member-scan-test.preview.emergentagent.com"

    # Prefer the file-based quick_start.bat (easier to edit), fallback to inline template.
    quick_path = Path(__file__).resolve().parent.parent / "deploy" / "quick_start.bat"
    if quick_path.exists():
        script = quick_path.read_text(encoding="utf-8").replace("__SERVER_URL__", public_url)
    else:
        script = SETUP_BAT_TEMPLATE.replace("__SERVER_URL__", public_url)
    script_crlf = script.replace("\r\n", "\n").replace("\n", "\r\n")
    return Response(
        content=script_crlf.encode("utf-8"),
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": "attachment; filename=\"electronic-archive-setup.bat\"",
            "Cache-Control": "no-store",
        },
    )


@api_router.get("/admin/settings")
async def get_settings(user: Dict[str, Any] = Depends(require_admin)):
    """Get current admin settings like session timeout."""
    doc = await db.settings.find_one({"id": SETTINGS_DOC_ID}, {"_id": 0})
    if not doc:
        return {"session_timeout_hours": 12}
    return {"session_timeout_hours": doc.get("session_timeout_hours", 12)}


# ═══════════════════════════════════════════════════════════════════════════
# LOCAL UPDATE MANAGER - Super Admin Only
# ═══════════════════════════════════════════════════════════════════════════

class UpdateStatus(BaseModel):
    """Status of update operation"""
    status: str  # "running", "success", "failed"
    step: str
    progress: int  # 0-100
    message: str
    timestamp: str = Field(default_factory=now_iso)

# Global update status
_update_status: Optional[UpdateStatus] = None

@api_router.get("/admin/update/status")
async def get_update_status(user: Dict[str, Any] = Depends(require_super_admin)):
    """Get current update status"""
    return _update_status or UpdateStatus(
        status="idle",
        step="",
        progress=0,
        message="No update in progress"
    )

@api_router.post("/admin/update/trigger")
async def trigger_local_update(user: Dict[str, Any] = Depends(require_super_admin)):
    """
    Trigger local update process
    
    This will:
    1. Create backup
    2. Rebuild frontend
    3. Update Service Worker version
    4. Notify all clients
    5. Trigger reload on all connected browsers
    """
    global _update_status
    
    if _update_status and _update_status.status == "running":
        raise HTTPException(
            status_code=409,
            detail="Update already in progress"
        )
    
    # Start update in background
    asyncio.create_task(perform_local_update())
    
    return {"message": "Update started", "check_status": "/api/admin/update/status"}

async def perform_local_update():
    """Perform the actual update process"""
    global _update_status
    
    try:
        # Step 1: Initialize
        _update_status = UpdateStatus(
            status="running",
            step="Initializing",
            progress=5,
            message="Starting update process..."
        )
        await asyncio.sleep(1)
        
        # Step 2: Backup
        _update_status.step = "Backup"
        _update_status.progress = 15
        _update_status.message = "Creating backup..."
        # Backup is already handled by installer
        await asyncio.sleep(1)
        
        # Step 3: Frontend Build
        _update_status.step = "Frontend Build"
        _update_status.progress = 30
        _update_status.message = "Building React frontend..."
        
        frontend_path = Path(__file__).resolve().parent.parent / "frontend"
        
        # Run yarn build
        import subprocess
        result = subprocess.run(
            ["yarn", "build"],
            cwd=str(frontend_path),
            capture_output=True,
            text=True,
            timeout=300
        )
        
        if result.returncode != 0:
            raise Exception(f"Build failed: {result.stderr}")
        
        _update_status.progress = 60
        _update_status.message = "Build completed"
        await asyncio.sleep(1)
        
        # Step 4: Update Service Worker Version
        _update_status.step = "Service Worker"
        _update_status.progress = 70
        _update_status.message = "Updating Service Worker version..."
        
        sw_path = frontend_path / "build" / "service-worker.js"
        if sw_path.exists():
            sw_content = sw_path.read_text()
            build_time = str(int(datetime.now(timezone.utc).timestamp() * 1000))
            
            # Replace BUILD_TIMESTAMP
            sw_content = sw_content.replace("__BUILD_TIMESTAMP__", build_time)
            sw_path.write_text(sw_content)
            
            # Create version.json
            version_json = {
                "version": build_time,
                "buildDate": now_iso(),
                "updatedBy": "Local Update Manager"
            }
            version_path = frontend_path / "build" / "version.json"
            version_path.write_text(json.dumps(version_json, indent=2))
            
            _update_status.message = f"Service Worker updated (v{build_time})"
        
        await asyncio.sleep(1)
        
        # Step 5: Notify clients (via presence/broadcast)
        _update_status.step = "Notify"
        _update_status.progress = 85
        _update_status.message = "Notifying connected clients..."
        
        # Update a flag in MongoDB that clients can check
        await db["_system_state"].update_one(
            {"key": "update_available"},
            {"$set": {
                "key": "update_available",
                "value": True,
                "version": build_time if sw_path.exists() else str(int(time.time())),
                "timestamp": now_iso()
            }},
            upsert=True
        )
        
        await asyncio.sleep(1)
        
        # Step 6: Complete
        _update_status.step = "Complete"
        _update_status.progress = 100
        _update_status.status = "success"
        _update_status.message = "Update completed successfully!"
        
    except Exception as e:
        _update_status = UpdateStatus(
            status="failed",
            step="Error",
            progress=0,
            message=f"Update failed: {str(e)}"
        )

@api_router.get("/admin/update/check")
async def check_for_updates():
    """Check if update is available (public endpoint)"""
    update_flag = await db["_system_state"].find_one(
        {"key": "update_available"},
        {"_id": 0}
    )
    
    if update_flag and update_flag.get("value"):
        return {
            "updateAvailable": True,
            "version": update_flag.get("version"),
            "timestamp": update_flag.get("timestamp")
        }
    
    return {"updateAvailable": False}

@api_router.post("/admin/update/acknowledge")
async def acknowledge_update(user: Dict[str, Any] = Depends(require_user)):
    """Mark update as acknowledged (client reloaded)"""
    # Clear the update flag after client reloads
    await db["_system_state"].update_one(
        {"key": "update_available"},
        {"$set": {"value": False}}
    )
    return {"message": "Update acknowledged"}

# ═══════════════════════════════════════════════════════════════════════════
# END LOCAL UPDATE MANAGER
# ═══════════════════════════════════════════════════════════════════════════

async def get_settings(user: Dict[str, Any] = Depends(require_user)):
    """Read app settings - any logged-in user can read (UI may need timeout info), only admin can update."""
    hours = await get_session_timeout_hours()
    return {
        "session_timeout_hours": hours,
        "all_portals": [{"key": k, "label": PORTAL_LABELS.get(k, k)} for k in ALL_PORTAL_KEYS],
    }


@api_router.put("/admin/settings")
async def update_settings(payload: AppSettingsIn, actor: Dict[str, Any] = Depends(require_admin)):
    hours = max(1, min(int(payload.session_timeout_hours or 12), 720))
    await db.settings.update_one(
        {"id": SETTINGS_DOC_ID},
        {"$set": {"id": SETTINGS_DOC_ID, "session_timeout_hours": hours, "updated_at": now_iso()}},
        upsert=True,
    )
    return {"session_timeout_hours": hours}


@api_router.get("/departments", response_model=List[Department])
async def list_departments(user: Dict[str, Any] = Depends(require_user)):
    departments = await db.departments.find({}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    return departments


@api_router.post("/departments", response_model=Department)
async def create_department(input: DepartmentIn, user: Dict[str, Any] = Depends(require_admin)):
    # التحقق من الحقول المطلوبة
    if not input.name or not input.name.strip():
        raise HTTPException(status_code=422, detail="اسم الإدارة مطلوب")
    if not input.code or not input.code.strip():
        raise HTTPException(status_code=422, detail="الكود المختصر مطلوب")
    if not input.description or not input.description.strip():
        raise HTTPException(status_code=422, detail="وصف الإدارة مطلوب")
    
    department = Department(**input.model_dump())
    doc = department.model_dump()
    await db.departments.insert_one(doc)
    return department


@api_router.put("/departments/{department_id}", response_model=Department)
async def update_department(department_id: str, input: DepartmentIn, user: Dict[str, Any] = Depends(require_admin)):
    update_doc = input.model_dump()
    update_doc["updated_at"] = now_iso()
    result = await db.departments.find_one_and_update(
        {"id": department_id},
        {"$set": update_doc},
        projection={"_id": 0},
        return_document=True,
    )
    if not result:
        raise HTTPException(status_code=404, detail="الإدارة غير موجودة")
    return result


@api_router.delete("/departments/{department_id}")
async def delete_department(department_id: str, user: Dict[str, Any] = Depends(require_super_admin)):
    """
    حذف إدارة (super_admin فقط).
    يتحقق أولاً من عدم وجود أعضاء في هذه الإدارة.
    """
    # التحقق من وجود أعضاء في هذه الإدارة
    members_count = await db.members.count_documents({"department_id": department_id})
    if members_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"لا يمكن حذف الإدارة. يوجد {members_count} عضو مسجل في هذه الإدارة. يجب حذف أو نقل الأعضاء أولاً."
        )
    
    # التحقق من وجود اشتراكات
    subscriptions_count = await db.subscriptions.count_documents({"department_id": department_id})
    if subscriptions_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"لا يمكن حذف الإدارة. يوجد {subscriptions_count} اشتراك مسجل. يجب حذفها أولاً."
        )
    
    # التحقق من وجود مساعدات
    aids_count = await db.aids.count_documents({"department_id": department_id})
    if aids_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"لا يمكن حذف الإدارة. يوجد {aids_count} مساعدة مسجلة. يجب حذفها أولاً."
        )
    
    # حذف الإدارة
    result = await db.departments.delete_one({"id": department_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الإدارة غير موجودة")
    
    return {"deleted": True, "message": "تم حذف الإدارة بنجاح"}


@api_router.get("/admin/retirement-schedule", response_model=List[RetirementSchedule])
async def get_retirement_schedule(user: Dict[str, Any] = Depends(require_admin)):
    return await get_retirement_schedule_docs()


@api_router.put("/admin/retirement-schedule", response_model=List[RetirementSchedule])
async def save_retirement_schedule(input: List[RetirementScheduleIn], user: Dict[str, Any] = Depends(require_admin)):
    if not input:
        raise HTTPException(status_code=400, detail="يجب إضافة بند واحد على الأقل في جدول المعاش")
    
    # التحقق من الحقول المطلوبة لكل صف
    for idx, item in enumerate(input, 1):
        if not item.effective_date or not item.effective_date.strip():
            raise HTTPException(status_code=422, detail=f"تاريخ السريان مطلوب في الصف {idx}")
        if not item.retirement_age or item.retirement_age <= 0:
            raise HTTPException(status_code=422, detail=f"سن المعاش مطلوب ويجب أن يكون أكبر من صفر في الصف {idx}")
    
    docs = [RetirementSchedule(**item.model_dump()).model_dump() for item in input]
    await db.retirement_schedule.delete_many({})
    await db.retirement_schedule.insert_many(docs)
    return docs


@api_router.post("/documents/upload", response_model=DocumentRecord)
async def upload_document(file: UploadFile = File(...), user: Dict[str, Any] = Depends(require_user)):
    file_name = safe_filename(file.filename or "document.pdf")
    target_path = TMP_DIR / file_name
    with target_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    content_type = file.content_type or mimetypes.guess_type(file.filename or "")[0] or "application/octet-stream"
    extracted_text = extract_text_for_document(target_path, content_type)
    record = DocumentRecord(
        file_name=file_name,
        original_name=file.filename or file_name,
        content_type=content_type,
        path=str(target_path),
        source="upload",
        extracted_text=extracted_text[:8000],
        extracted_fields=parse_membership_fields(extracted_text),
    )
    await db.documents.insert_one(record.model_dump())
    return record


# ─── Member duplicate-detection rule (single source of truth) ──────────────
#
# قواعد التكرار:
# 1. رقم العضوية: يُعتبر مكرر فقط في نفس اللجنة النقابية (في نفس الإدارة)
# 2. الكشف الذكي: إذا تطابق (الاسم + الرقم القومي + تاريخ الميلاد) = نفس الشخص (في أي مكان!)
#
async def _find_member_duplicate(record: Dict[str, Any], exclude_id: Optional[str] = None) -> Optional[Dict[str, Any]]:
    """
    قاعدة التكرار الوحيدة:
    غير مسموح بتسجيل بيانات عضو في أكثر من لجنة أو أكثر من محافظة
    يتم الفحص التدريجي: الاسم → الرقم القومي → تاريخ الميلاد
    إذا تطابقت الثلاثة معاً، يُرفض التسجيل نهائياً
    """
    national_id = (record.get("national_id") or "").strip()
    name = (record.get("name") or "").strip()
    birth_date = record.get("birth_date")

    # يجب توفر الثلاثة للفحص
    if not (name and national_id and birth_date):
        return None

    query: Dict[str, Any] = {
        "name": name,
        "national_id": national_id,
        "birth_date": birth_date
    }
    
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    
    return await db.members.find_one(query, {"_id": 0})


# ─── البحث عن جميع التكرارات في كل اللجان ──────────────
async def _find_all_member_duplicates(record: Dict[str, Any], exclude_id: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    البحث عن جميع التكرارات لنفس العضو في كل اللجان والمحافظات.
    القاعدة: غير مسموح بتسجيل عضو في أكثر من لجنة أو محافظة
    يتم البحث باستخدام: الاسم + الرقم القومي + تاريخ الميلاد
    """
    national_id = (record.get("national_id") or "").strip()
    name = (record.get("name") or "").strip()
    birth_date = record.get("birth_date")

    # يجب توفر الثلاثة للبحث
    if not (name and national_id and birth_date):
        return []

    query: Dict[str, Any] = {
        "name": name,
        "national_id": national_id,
        "birth_date": birth_date
    }
    
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    
    # البحث عن جميع التكرارات
    all_duplicates = await db.members.find(query, {"_id": 0}).to_list(100)
    return all_duplicates


def _duplicate_reason(found: Dict[str, Any], record: Dict[str, Any], all_duplicates: List[Dict[str, Any]] = None) -> str:
    """
    رسالة التكرار التي تظهر للمستخدم.
    توضح أن العضو مسجل بالفعل وتذكر اسم اللجنة/اللجان.
    """
    name = (record.get("name") or "").strip()
    nid = (record.get("national_id") or "").strip()
    
    # عدد اللجان المكررة
    num_duplicates = len(all_duplicates) if all_duplicates else 1
    
    if num_duplicates > 1:
        # العضو مسجل في أكثر من لجنة
        committees_list = []
        for dup in (all_duplicates or [found])[:3]:  # أول 3 لجان
            uc = dup.get("union_committee", "غير محدد")
            gov = dup.get("governorate", "غير محدد")
            committees_list.append(f"\"{uc}\" - {gov}")
        
        committees_text = "، ".join(committees_list)
        if num_duplicates > 3:
            committees_text += f" و{num_duplicates - 3} لجنة أخرى"
        
        return f"العضو \"{name}\" (رقم قومي: {nid}) مسجل بالفعل في {num_duplicates} لجنة: {committees_text}"
    else:
        # العضو مسجل في لجنة واحدة
        uc = found.get("union_committee", "غير محدد")
        gov = found.get("governorate", "غير محدد")
        return f"العضو \"{name}\" (رقم قومي: {nid}) مسجل بالفعل في لجنة \"{uc}\" - محافظة {gov}"


def _raise_duplicate(found: Dict[str, Any], record: Dict[str, Any], all_duplicates: List[Dict[str, Any]] = None) -> None:
    """Raise the standard 409 HTTPException carrying the existing member doc.

    The frontend uses ``existing_member`` to render the popup with all the
    existing member's fields + a close button.
    """
    
    # إعداد قائمة اللجان المكررة
    committees_info = []
    if all_duplicates:
        for dup in all_duplicates:
            committees_info.append({
                "governorate": dup.get("governorate", ""),
                "union_committee": dup.get("union_committee", ""),
                "name": dup.get("name", ""),
                "membership_number": dup.get("membership_number", ""),
                "national_id": dup.get("national_id", ""),
                "status": dup.get("status", ""),
                "id": dup.get("id", "")
            })
    
    raise HTTPException(
        status_code=409,
        detail={
            "code": "duplicate_member",
            "message": _duplicate_reason(found, record, all_duplicates),
            "existing_member": found,
            "all_duplicates": all_duplicates if all_duplicates else [found],
            "committees_info": committees_info,
            "duplicate_count": len(all_duplicates) if all_duplicates else 1
        },
    )


@api_router.post("/members", response_model=Member)
async def create_member(input: MemberIn, user: Dict[str, Any] = Depends(require_user)):
    normalized_input = input.model_dump()
    for key, value in normalized_input.items():
        if isinstance(value, str):
            normalized_input[key] = re.sub(r"\s+", " ", value.strip())

    department = await db.departments.find_one({"id": normalized_input["department_id"]}, {"_id": 0})
    if not department:
        raise HTTPException(status_code=404, detail="الإدارة غير موجودة")
    document_file_name = ""
    document_original_name = ""
    document_url = ""
    if normalized_input.get("document_id"):
        document = await db.documents.find_one({"id": normalized_input["document_id"]}, {"_id": 0})
        if not document:
            raise HTTPException(status_code=404, detail="ملف الاستمارة غير موجود")
        document_file_name = document["file_name"]
        document_original_name = document["original_name"]
        document_url = f"/api/documents/{document['id']}/download"
    duplicate = await _find_member_duplicate(normalized_input)
    if duplicate:
        # البحث عن جميع التكرارات في كل اللجان
        all_duplicates = await _find_all_member_duplicates(normalized_input)
        _raise_duplicate(duplicate, normalized_input, all_duplicates)
    member = Member(
        **normalized_input,
        document_file_name=document_file_name,
        document_original_name=document_original_name,
        document_url=document_url,
    )
    member_doc = member.model_dump(exclude={"retirement_age", "retirement_date", "retirement_due", "retirement_label"})
    await db.members.insert_one(member_doc)
    return await enrich_member_retirement(member_doc)


@api_router.put("/members/{member_id}", response_model=Member)
async def update_member(member_id: str, input: MemberIn, user: Dict[str, Any] = Depends(require_user)):
    """Full edit of a member record (all fields)."""
    existing = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    normalized = input.model_dump()
    for key, value in normalized.items():
        if isinstance(value, str):
            normalized[key] = re.sub(r"\s+", " ", value.strip())
    # Duplicate-check (exclude self) — same unified rule as create.
    duplicate = await _find_member_duplicate(normalized, exclude_id=member_id)
    if duplicate:
        # البحث عن جميع التكرارات في كل اللجان
        all_duplicates = await _find_all_member_duplicates(normalized, exclude_id=member_id)
        _raise_duplicate(duplicate, normalized, all_duplicates)
    document_file_name = existing.get("document_file_name", "")
    document_original_name = existing.get("document_original_name", "")
    document_url = existing.get("document_url", "")
    if normalized.get("document_id") and normalized["document_id"] != existing.get("document_id"):
        document = await db.documents.find_one({"id": normalized["document_id"]}, {"_id": 0})
        if document:
            document_file_name = document["file_name"]
            document_original_name = document["original_name"]
            document_url = f"/api/documents/{document['id']}/download"
    update = {
        **normalized,
        "document_file_name": document_file_name,
        "document_original_name": document_original_name,
        "document_url": document_url,
        "updated_at": now_iso(),
    }
    await db.members.update_one({"id": member_id}, {"$set": update})
    updated = await db.members.find_one({"id": member_id}, {"_id": 0})
    return await enrich_member_retirement(updated)



VALID_MEMBER_STATUSES = {"فعال", "متوفي", "استقالة", "إسقاط", "عجز كلي أو جزئي منهي للخدمة"}


@api_router.patch("/members/{member_id}/status", response_model=Member)
async def update_member_status(member_id: str, payload: MemberStatusUpdate, user: Dict[str, Any] = Depends(require_user)):
    if payload.status not in VALID_MEMBER_STATUSES:
        raise HTTPException(status_code=400, detail="حالة غير مدعومة")
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    status_date = payload.status_date or today_ar()
    update = {"status": payload.status, "status_date": status_date, "updated_at": now_iso()}
    await db.members.update_one({"id": member_id}, {"$set": update})
    # Auto-create a pending aid record when status triggers aid eligibility
    aid_type = STATUS_TO_AID_TYPE.get(payload.status)
    if aid_type:
        existing = await db.aids.find_one(
            {"department_id": member.get("department_id"), "member_id": member_id, "aid_type": aid_type},
            {"_id": 0},
        )
        if not existing:
            aid = AidRecord(
                department_id=member.get("department_id", ""),
                member_id=member_id,
                aid_type=aid_type,
                status=AID_STATUS_PENDING,
                triggered_by_status=payload.status,
                member_name=member.get("name", ""),
                member_national_id=member.get("national_id", ""),
                member_membership_number=member.get("membership_number", ""),
                member_governorate=member.get("governorate", ""),
                member_union_committee=member.get("union_committee", ""),
                member_birth_date=member.get("birth_date", ""),
                member_subscription_date=member.get("subscription_date", ""),
                member_address=member.get("address", ""),
                member_phone=member.get("phone", ""),
                member_beneficiary_name=member.get("beneficiary_name", ""),
                member_status_date=status_date,
            )
            await db.aids.insert_one(aid.model_dump())
    updated = await db.members.find_one({"id": member_id}, {"_id": 0})
    return await enrich_member_retirement(updated)


@api_router.delete("/members/{member_id}")
async def delete_member(member_id: str, user: Dict[str, Any] = Depends(require_user)):
    result = await db.members.delete_one({"id": member_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    return {"deleted": True}


@api_router.delete("/members")
async def delete_all_members(
    department_id: str,
    confirm: str,
    user: Dict[str, Any] = Depends(require_admin),
):
    """Admin-only: wipe all members in a department. Requires confirm=DELETE."""
    if confirm != "DELETE":
        raise HTTPException(status_code=400, detail="تأكيد الحذف غير صحيح")
    if not department_id:
        raise HTTPException(status_code=400, detail="department_id مطلوب")
    result = await db.members.delete_many({"department_id": department_id})
    return {"deleted": True, "count": result.deleted_count}



@api_router.get("/members", response_model=List[Member])
async def list_members(
    department_id: Optional[str] = None,
    search: Optional[str] = None,
    governorate: Optional[str] = None,
    union_committee: Optional[str] = None,
    retirement_due: Optional[bool] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    query: Dict[str, Any] = {}
    if department_id:
        query["department_id"] = department_id
    if governorate:
        query["governorate"] = governorate
    if union_committee:
        query["union_committee"] = union_committee
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"national_id": {"$regex": search, "$options": "i"}},
        ]
    members = await db.members.find(query, {"_id": 0}).sort("created_at", -1).to_list(100000)
    enriched = await enrich_members_retirement(members)
    if retirement_due is not None:
        enriched = [member for member in enriched if member.get("retirement_due") is retirement_due]
    return enriched


@api_router.get("/members/paginated")
async def list_members_paginated(
    department_id: Optional[str] = None,
    search: Optional[str] = None,
    governorate: Optional[str] = None,
    union_committee: Optional[str] = None,
    status: Optional[str] = None,
    retirement_due: Optional[bool] = None,
    missing: Optional[str] = None,  # "governorate" | "committee" | "both" | "any"
    page: int = 1,
    page_size: int = 50,
    user: Dict[str, Any] = Depends(require_user),
):
    """Paginated, indexed listing — fast for large datasets (>1000 members).

    `missing` filters for members with missing classification data:
      • governorate → empty/null governorate only
      • committee   → empty/null union_committee only
      • both        → both governorate AND committee empty
      • any         → at least one of them empty
    """
    page = max(page, 1)
    page_size = max(min(page_size, 500), 1)
    query: Dict[str, Any] = {}
    if department_id:
        query["department_id"] = department_id
    if governorate:
        query["governorate"] = governorate
    if union_committee:
        query["union_committee"] = union_committee
    if status:
        query["status"] = status
    if missing:
        empty_gov = {"$or": [{"governorate": ""}, {"governorate": None}, {"governorate": {"$exists": False}}]}
        empty_com = {"$or": [{"union_committee": ""}, {"union_committee": None}, {"union_committee": {"$exists": False}}]}
        if missing == "governorate":
            query["$and"] = (query.get("$and") or []) + [empty_gov]
        elif missing == "committee":
            query["$and"] = (query.get("$and") or []) + [empty_com]
        elif missing == "both":
            query["$and"] = (query.get("$and") or []) + [empty_gov, empty_com]
        elif missing == "any":
            query["$and"] = (query.get("$and") or []) + [{"$or": [empty_gov, empty_com]}]
    if search:
        s = search.strip()
        if s:
            query["$or"] = [
                {"name": {"$regex": s, "$options": "i"}},
                {"national_id": {"$regex": s, "$options": "i"}},
                {"membership_number": {"$regex": s, "$options": "i"}},
            ]
    skip = (page - 1) * page_size
    # When retirement_due filter is active we need to enrich the full set then filter
    # because retirement is computed, not stored. Otherwise we paginate at DB level.
    if retirement_due is not None:
        all_members = await db.members.find(query, {"_id": 0}).sort("created_at", -1).to_list(100000)
        enriched_all = await enrich_members_retirement(all_members)
        filtered = [m for m in enriched_all if m.get("retirement_due") is retirement_due]
        total = len(filtered)
        items = filtered[skip:skip + page_size]
    else:
        total = await db.members.count_documents(query)
        cursor = db.members.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size)
        page_docs = await cursor.to_list(page_size)
        items = await enrich_members_retirement(page_docs)
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if page_size else 1,
    }


@api_router.get("/classifications")
async def get_classifications(department_id: Optional[str] = None, user: Dict[str, Any] = Depends(require_user)):
    query: Dict[str, Any] = {}
    if department_id:
        query["department_id"] = department_id
    # Use a single aggregation to build the governorate -> committees mapping.
    # This is dramatically faster than streaming all member docs to Python.
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": {"governorate": "$governorate", "union_committee": "$union_committee"},
        }},
    ]
    pairs = await db.members.aggregate(pipeline).to_list(100000)
    governorates_set = set()
    union_committees_set = set()
    gov_map: Dict[str, set] = {}

    # Drop obviously-bad values that crept in from import errors (e.g. someone
    # pasted a birth-date into the governorate column). A real governorate /
    # committee name is text-only — anything that's pure digits + separators
    # like "02/17/1954" or "1965-11-03" is filtered out of the picker list.
    bad_pattern = re.compile(r"^[\d/\-\.\s:]+$")

    def is_real_label(value: str) -> bool:
        return bool(value) and not bad_pattern.match(value)

    for p in pairs:
        g = (p["_id"].get("governorate") or "").strip()
        c = (p["_id"].get("union_committee") or "").strip()
        if is_real_label(g):
            governorates_set.add(g)
        if is_real_label(c):
            union_committees_set.add(c)
        if is_real_label(g) and is_real_label(c):
            gov_map.setdefault(g, set()).add(c)
    # Also surface any entries that an admin added manually via the taxonomy
    # endpoints — those may not have a single member yet.
    tax_query: Dict[str, Any] = {}
    if department_id:
        tax_query["department_id"] = department_id
    async for d in db.taxonomy.find(tax_query, {"_id": 0, "kind": 1, "name": 1, "governorate": 1}):
        if d.get("kind") == "governorate":
            v = (d.get("name") or "").strip()
            if is_real_label(v):
                governorates_set.add(v)
        elif d.get("kind") == "committee":
            cg = (d.get("governorate") or "").strip()
            cn = (d.get("name") or "").strip()
            if is_real_label(cg) and is_real_label(cn):
                governorates_set.add(cg)
                union_committees_set.add(cn)
                gov_map.setdefault(cg, set()).add(cn)
    committees_by_governorate = {g: sorted(list(coms)) for g, coms in gov_map.items()}
    return {
        "governorates": sorted(governorates_set),
        "union_committees": sorted(union_committees_set),
        "committees_by_governorate": committees_by_governorate,
    }


@api_router.get("/reports/membership")
async def membership_reports(department_id: Optional[str] = None, user: Dict[str, Any] = Depends(require_user)):
    query: Dict[str, Any] = {}
    if department_id:
        query["department_id"] = department_id
    # Only fetch the fields we actually need for reporting — drastically reduces
    # memory + serialization time when there are thousands of members.
    projection = {
        "_id": 0,
        "governorate": 1,
        "union_committee": 1,
        "status": 1,
        "birth_date": 1,
    }
    members = await db.members.find(query, projection).to_list(100000)
    enriched = await enrich_members_retirement(members)
    inactive_statuses = {"متوفي", "استقالة", "إسقاط", "عجز كلي أو جزئي منهي للخدمة"}

    by_governorate_active: Dict[str, int] = defaultdict(int)
    by_governorate_all: Dict[str, int] = defaultdict(int)
    by_committee_per_governorate: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    active_count = 0
    retirement_due_count = 0
    deceased_count = 0
    resignation_count = 0
    dropout_count = 0
    disability_count = 0
    missing_governorate_count = 0
    missing_committee_count = 0
    missing_both_count = 0
    for member in enriched:
        raw_gov = (member.get("governorate") or "").strip()
        raw_com = (member.get("union_committee") or "").strip()
        if not raw_gov and not raw_com:
            missing_both_count += 1
        elif not raw_gov:
            missing_governorate_count += 1
        elif not raw_com:
            missing_committee_count += 1
        gov = raw_gov or "غير محدد"
        com = raw_com or "غير محدد"
        status = member.get("status") or ""
        by_governorate_all[gov] += 1
        if member.get("retirement_due"):
            retirement_due_count += 1
        if status == "متوفي":
            deceased_count += 1
        elif status == "استقالة":
            resignation_count += 1
        elif status == "إسقاط":
            dropout_count += 1
        elif status == "عجز كلي أو جزئي منهي للخدمة":
            disability_count += 1
        is_active = status not in inactive_statuses and not member.get("retirement_due")
        if is_active:
            active_count += 1
            by_governorate_active[gov] += 1
            by_committee_per_governorate[gov][com] += 1

    return {
        "total_members": len(enriched),
        "active_count": active_count,
        "retirement_due_count": retirement_due_count,
        "deceased_count": deceased_count,
        "disability_count": disability_count,
        "resignation_count": resignation_count,
        "dropout_count": dropout_count,
        "missing_governorate_count": missing_governorate_count,
        "missing_committee_count": missing_committee_count,
        "missing_both_count": missing_both_count,
        "missing_any_count": missing_governorate_count + missing_committee_count + missing_both_count,
        "by_governorate": [{"name": key, "count": value} for key, value in sorted(by_governorate_all.items())],
        "by_governorate_active": [{"name": key, "count": value} for key, value in sorted(by_governorate_active.items())],
        "by_committee_per_governorate": {
            gov: [{"name": com, "count": cnt} for com, cnt in sorted(committees.items())]
            for gov, committees in sorted(by_committee_per_governorate.items())
        },
        "by_union_committee": [
            {"name": com, "count": sum(committees.get(com, 0) for committees in by_committee_per_governorate.values())}
            for com in sorted({c for committees in by_committee_per_governorate.values() for c in committees})
        ],
    }


@api_router.get("/reports/pension")
async def pension_report(
    department_id: Optional[str] = None,
    governorate: Optional[str] = None,
    union_committee: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    user: Dict[str, Any] = Depends(require_user),
):
    """تقرير المعاش - الأعضاء الذين بلغوا سن المعاش (retirement_due = true)"""
    page = max(page, 1)
    page_size = max(min(page_size, 100), 1)
    
    query: Dict[str, Any] = {}
    
    if department_id:
        query["department_id"] = department_id
    if governorate:
        query["governorate"] = governorate
    if union_committee:
        query["union_committee"] = union_committee
    
    # البحث
    if search and search.strip():
        search_normalized = search.strip()
        # تحويل الأرقام العربية/الهندية إلى إنجليزية
        arabic_to_english = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
        search_en = search_normalized.translate(arabic_to_english)
        
        query["$or"] = [
            {"name": {"$regex": search_normalized, "$options": "i"}},
            {"national_id": {"$regex": search_en, "$options": "i"}},
            {"membership_number": {"$regex": search_en, "$options": "i"}},
            {"union_committee": {"$regex": search_normalized, "$options": "i"}},
        ]
    
    # جلب جميع الأعضاء وحساب retirement_due
    all_members = await db.members.find(query, {"_id": 0}).to_list(100000)
    enriched_members = await enrich_members_retirement(all_members)
    
    # فلترة الأعضاء الذين بلغوا سن المعاش
    pension_members = [m for m in enriched_members if m.get("retirement_due") is True]
    
    # فلترة حسب تاريخ المعاش (retirement_date)
    if from_date or to_date:
        filtered = []
        for m in pension_members:
            ret_date = m.get("retirement_date", "")
            if ret_date:
                if from_date and ret_date < from_date:
                    continue
                if to_date and ret_date > to_date:
                    continue
                filtered.append(m)
        pension_members = filtered
    
    # ترتيب حسب تاريخ المعاش (الأحدث أولاً)
    pension_members.sort(key=lambda x: x.get("retirement_date", ""), reverse=True)
    
    # Pagination
    total_count = len(pension_members)
    skip = (page - 1) * page_size
    paginated_members = pension_members[skip:skip + page_size]
    
    return {
        "members": paginated_members,
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": (total_count + page_size - 1) // page_size if total_count > 0 else 0,
    }


@api_router.get("/members/{member_id}", response_model=Member)
async def get_member(member_id: str, user: Dict[str, Any] = Depends(require_user)):
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    return await enrich_member_retirement(member)


# ─── Disclosure-statement reports (الكشوف التفريغية) ──────────────────────────
# Four read-only reports the operator can print on A4 with auto page numbers.
#   1) Detailed financial report — one governorate, every collected
#      subscription row between two dates.
#   2) Summary financial report — one row per (governorate, committee) for
#      the chosen period, showing membership size, collected, expected dues.
#   3) Overall membership report — count per governorate + grand total.
#   4) Detailed membership report — committees inside each governorate,
#      subtotals + grand total at the bottom.
#
# All endpoints respect the program-wide FINANCIAL_START_MONTH floor.


def _months_between(from_ym: str, to_ym: str) -> int:
    """Inclusive count of months between two YYYY-MM markers."""
    try:
        fy, fm = (int(x) for x in from_ym.split("-")[:2])
        ty, tm = (int(x) for x in to_ym.split("-")[:2])
    except Exception:  # noqa: BLE001
        return 0
    return max(0, (ty - fy) * 12 + (tm - fm) + 1)


def _ym(date_str: str) -> str:
    return (date_str or "")[:7]


def _floor_ym(ym: str) -> str:
    return ym if (ym and ym >= FINANCIAL_START_MONTH) else FINANCIAL_START_MONTH


@api_router.get("/reports/disclosure/governorate-detailed")
async def report_governorate_detailed(
    department_id: str,
    governorate: str,
    from_date: str,
    to_date: str,
    user: Dict[str, Any] = Depends(require_user),
):
    """Row-level dump of every subscription COLLECTED inside `governorate`
    between `from_date` and `to_date` (both inclusive, ISO-8601 strings)."""
    if not (department_id and governorate and from_date and to_date):
        raise HTTPException(status_code=400, detail="department_id / governorate / from_date / to_date مطلوبين")
    query = {
        "department_id": department_id,
        "governorate": governorate,
        "status": "تم التحصيل",
        "issued_at": {"$gte": from_date, "$lte": to_date},
    }
    rows = await db.subscriptions.find(query, {"_id": 0}).sort("issued_at", 1).to_list(None)
    total = sum(float(r.get("amount") or 0) for r in rows)
    return {
        "department_id": department_id, "governorate": governorate,
        "from_date": from_date, "to_date": to_date,
        "rows": rows, "total": total, "count": len(rows),
    }


@api_router.get("/reports/disclosure/committee-detailed")
async def report_committee_detailed(
    department_id: str,
    governorate: str,
    union_committee: str,
    from_date: str,
    to_date: str,
    user: Dict[str, Any] = Depends(require_user),
):
    """Same row-level dump as `governorate-detailed`, scoped to ONE
    committee inside the chosen governorate."""
    if not (department_id and governorate and union_committee and from_date and to_date):
        raise HTTPException(status_code=400, detail="department_id / governorate / union_committee / from_date / to_date مطلوبين")
    query = {
        "department_id": department_id,
        "governorate": governorate,
        "union_committee": union_committee,
        "status": "تم التحصيل",
        "issued_at": {"$gte": from_date, "$lte": to_date},
    }
    rows = await db.subscriptions.find(query, {"_id": 0}).sort("issued_at", 1).to_list(None)
    total = sum(float(r.get("amount") or 0) for r in rows)
    return {
        "department_id": department_id, "governorate": governorate, "union_committee": union_committee,
        "from_date": from_date, "to_date": to_date,
        "rows": rows, "total": total, "count": len(rows),
    }


def _owed_months(owed: float, members: int) -> float:
    """Number-of-months equivalent of an owed amount, given an active size."""
    if not members or members <= 0:
        return 0.0
    denom = members * COMMITTEE_MONTHLY_RATE
    if denom <= 0:
        return 0.0
    return round(max(0.0, owed / denom), 2)


@api_router.get("/reports/disclosure/all-governorates-overdue")
async def report_all_governorates_overdue(
    department_id: str,
    from_date: str,
    to_date: str,
    only_overdue: bool = True,
    user: Dict[str, Any] = Depends(require_user),
):
    """Overdue report — every (governorate, committee) pair with active
    membership size, collected amount, expected amount, outstanding
    dues, and equivalent owed-months count for the period. Includes
    governorate subtotals and a grand total. Pass `only_overdue=false`
    to also include committees with `owed == 0`."""
    if not (department_id and from_date and to_date):
        raise HTTPException(status_code=400, detail="department_id / from_date / to_date مطلوبين")
    members = await db.members.find(
        {"department_id": department_id, "status": "فعال"},
        {"_id": 0, "governorate": 1, "union_committee": 1},
    ).to_list(None)
    subs = await db.subscriptions.find(
        {"department_id": department_id, "status": "تم التحصيل",
         "issued_at": {"$gte": from_date, "$lte": to_date}},
        {"_id": 0, "governorate": 1, "union_committee": 1, "amount": 1},
    ).to_list(None)
    from_ym = _floor_ym(_ym(from_date))
    to_ym = _ym(to_date)
    months = _months_between(from_ym, to_ym)
    bucket: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(lambda: {"members": 0, "collected": 0.0})
    for m in members:
        gov = (m.get("governorate") or "").strip()
        com = (m.get("union_committee") or "").strip()
        if _is_invalid_taxonomy_value(gov) or _is_invalid_taxonomy_value(com):
            continue
        bucket[(gov, com)]["members"] += 1
    for s in subs:
        gov = (s.get("governorate") or "").strip()
        com = (s.get("union_committee") or "").strip()
        if _is_invalid_taxonomy_value(gov) or _is_invalid_taxonomy_value(com):
            continue
        bucket[(gov, com)]["collected"] += float(s.get("amount") or 0)
    by_gov: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for (gov, com), agg in bucket.items():
        if not gov or not com:
            continue
        expected = agg["members"] * COMMITTEE_MONTHLY_RATE * months
        owed = max(0.0, expected - agg["collected"])
        if only_overdue and owed <= 0:
            continue
        by_gov[gov].append({
            "union_committee": com, "members": agg["members"],
            "collected": round(agg["collected"], 2),
            "expected": round(expected, 2),
            "owed": round(owed, 2),
            "owed_months": _owed_months(owed, agg["members"]),
        })
    governorates: List[Dict[str, Any]] = []
    for gov in sorted(by_gov.keys()):
        coms = sorted(by_gov[gov], key=lambda x: -x["owed"])
        if not coms:
            continue
        sub_members = int(sum(c["members"] for c in coms))
        sub = {
            "members": sub_members,
            "expected": round(sum(c["expected"] for c in coms), 2),
            "collected": round(sum(c["collected"] for c in coms), 2),
            "owed": round(sum(c["owed"] for c in coms), 2),
        }
        sub["owed_months"] = _owed_months(sub["owed"], sub_members)
        governorates.append({"governorate": gov, "committees": coms, "subtotal": sub})
    grand_members = int(sum(g["subtotal"]["members"] for g in governorates))
    grand = {
        "members": grand_members,
        "expected": round(sum(g["subtotal"]["expected"] for g in governorates), 2),
        "collected": round(sum(g["subtotal"]["collected"] for g in governorates), 2),
        "owed": round(sum(g["subtotal"]["owed"] for g in governorates), 2),
    }
    grand["owed_months"] = _owed_months(grand["owed"], grand_members)
    return {
        "department_id": department_id, "from_date": from_date, "to_date": to_date,
        "from_month": from_ym, "to_month": to_ym, "months": months,
        "monthly_rate": COMMITTEE_MONTHLY_RATE,
        "governorates": governorates, "grand_total": grand,
    }


@api_router.get("/reports/disclosure/governorate-overdue")
async def report_governorate_overdue(
    department_id: str,
    governorate: str,
    from_date: str,
    to_date: str,
    only_overdue: bool = True,
    user: Dict[str, Any] = Depends(require_user),
):
    """Per-governorate overdue report — one row per committee with active
    size, collected, expected, owed amount and owed-months equivalent,
    plus a governorate-level total row. Pass `only_overdue=false` to also
    include committees with zero overdue."""
    if not (department_id and governorate and from_date and to_date):
        raise HTTPException(status_code=400, detail="department_id / governorate / from_date / to_date مطلوبين")
    members = await db.members.find(
        {"department_id": department_id, "governorate": governorate, "status": "فعال"},
        {"_id": 0, "union_committee": 1},
    ).to_list(None)
    sub_rows = await db.subscriptions.find(
        {
            "department_id": department_id, "governorate": governorate,
            "status": "تم التحصيل",
            "issued_at": {"$gte": from_date, "$lte": to_date},
        },
        {"_id": 0, "union_committee": 1, "amount": 1},
    ).to_list(None)
    from_ym = _floor_ym(_ym(from_date))
    to_ym = _ym(to_date)
    months = _months_between(from_ym, to_ym)

    by_committee: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"members": 0, "collected": 0.0})
    for m in members:
        by_committee[(m.get("union_committee") or "").strip()]["members"] += 1
    for s in sub_rows:
        by_committee[(s.get("union_committee") or "").strip()]["collected"] += float(s.get("amount") or 0)

    rows: List[Dict[str, Any]] = []
    for committee, agg in by_committee.items():
        if not committee or _is_invalid_taxonomy_value(committee):
            continue
        expected = agg["members"] * COMMITTEE_MONTHLY_RATE * months
        owed = max(0.0, expected - agg["collected"])
        if only_overdue and owed <= 0:
            continue
        rows.append({
            "union_committee": committee,
            "members": agg["members"],
            "collected": round(agg["collected"], 2),
            "expected": round(expected, 2),
            "owed": round(owed, 2),
            "owed_months": _owed_months(owed, agg["members"]),
        })
    rows.sort(key=lambda x: -x["owed"])
    tot_members = int(sum(r["members"] for r in rows))
    totals = {
        "members": tot_members,
        "expected": round(sum(r["expected"] for r in rows), 2),
        "collected": round(sum(r["collected"] for r in rows), 2),
        "owed": round(sum(r["owed"] for r in rows), 2),
    }
    totals["owed_months"] = _owed_months(totals["owed"], tot_members)
    return {
        "department_id": department_id, "governorate": governorate,
        "from_date": from_date, "to_date": to_date,
        "from_month": from_ym, "to_month": to_ym, "months": months,
        "monthly_rate": COMMITTEE_MONTHLY_RATE,
        "rows": rows, "totals": totals,
    }


@api_router.get("/reports/disclosure/committee-overdue")
async def report_committee_overdue(
    department_id: str,
    governorate: str,
    union_committee: str,
    from_date: str,
    to_date: str,
    user: Dict[str, Any] = Depends(require_user),
):
    """Single-committee overdue summary with the same metric set."""
    if not (department_id and governorate and union_committee and from_date and to_date):
        raise HTTPException(status_code=400, detail="department_id / governorate / union_committee / from_date / to_date مطلوبين")
    members = await db.members.count_documents({"department_id": department_id, "status": "فعال",
                                                 "governorate": governorate, "union_committee": union_committee})
    sub_rows = await db.subscriptions.find(
        {"department_id": department_id, "status": "تم التحصيل",
         "governorate": governorate, "union_committee": union_committee,
         "issued_at": {"$gte": from_date, "$lte": to_date}},
        {"_id": 0, "amount": 1},
    ).to_list(None)
    collected = round(sum(float(s.get("amount") or 0) for s in sub_rows), 2)
    from_ym = _floor_ym(_ym(from_date))
    to_ym = _ym(to_date)
    months = _months_between(from_ym, to_ym)
    expected = round(members * COMMITTEE_MONTHLY_RATE * months, 2)
    owed = round(max(0, expected - collected), 2)
    return {
        "department_id": department_id, "governorate": governorate, "union_committee": union_committee,
        "from_date": from_date, "to_date": to_date, "from_month": from_ym, "to_month": to_ym, "months": months,
        "monthly_rate": COMMITTEE_MONTHLY_RATE,
        "members": members, "collected": collected, "expected": expected,
        "owed": owed, "owed_months": _owed_months(owed, members),
    }


@api_router.get("/reports/disclosure/governorate-summary")
async def report_governorate_summary(
    department_id: str,
    governorate: str,
    from_date: str,
    to_date: str,
    user: Dict[str, Any] = Depends(require_user),
):
    """One row per committee in `governorate`: active-members, collected
    amount, dues expected over the period. Totals at the bottom."""
    if not (department_id and governorate and from_date and to_date):
        raise HTTPException(status_code=400, detail="department_id / governorate / from_date / to_date مطلوبين")
    members = await db.members.find(
        {"department_id": department_id, "governorate": governorate, "status": "فعال"},
        {"_id": 0, "union_committee": 1},
    ).to_list(None)
    sub_rows = await db.subscriptions.find(
        {
            "department_id": department_id, "governorate": governorate,
            "status": "تم التحصيل",
            "issued_at": {"$gte": from_date, "$lte": to_date},
        },
        {"_id": 0, "union_committee": 1, "amount": 1},
    ).to_list(None)

    from_ym = _floor_ym(_ym(from_date))
    to_ym = _ym(to_date)
    months = _months_between(from_ym, to_ym)

    by_committee: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"members": 0, "collected": 0.0})
    for m in members:
        by_committee[(m.get("union_committee") or "").strip()]["members"] += 1
    for s in sub_rows:
        by_committee[(s.get("union_committee") or "").strip()]["collected"] += float(s.get("amount") or 0)

    rows: List[Dict[str, Any]] = []
    for committee, agg in sorted(by_committee.items()):
        if _is_invalid_taxonomy_value(committee):
            continue
        expected = agg["members"] * COMMITTEE_MONTHLY_RATE * months
        rows.append({
            "union_committee": committee or "(بدون لجنة)",
            "members": agg["members"],
            "collected": round(agg["collected"], 2),
            "expected": round(expected, 2),
            "owed": round(max(0, expected - agg["collected"]), 2),
        })
    totals = {
        "members": sum(r["members"] for r in rows),
        "collected": round(sum(r["collected"] for r in rows), 2),
        "expected": round(sum(r["expected"] for r in rows), 2),
        "owed": round(sum(r["owed"] for r in rows), 2),
    }
    return {
        "department_id": department_id, "governorate": governorate,
        "from_date": from_date, "to_date": to_date,
        "from_month": from_ym, "to_month": to_ym, "months": months,
        "monthly_rate": COMMITTEE_MONTHLY_RATE,
        "rows": rows, "totals": totals,
    }


@api_router.get("/reports/disclosure/membership-overall")
async def report_membership_overall(
    department_id: str,
    user: Dict[str, Any] = Depends(require_user),
):
    """Governorate × count, with a grand total. Includes inactive members
    explicitly so the operator can see the full picture."""
    pipeline = [
        {"$match": {"department_id": department_id}},
        {"$group": {"_id": "$governorate", "total": {"$sum": 1},
                    "active": {"$sum": {"$cond": [{"$eq": ["$status", "فعال"]}, 1, 0]}},
                    "inactive": {"$sum": {"$cond": [{"$eq": ["$status", "فعال"]}, 0, 1]}}}},
    ]
    raw = await db.members.aggregate(pipeline).to_list(None)
    rows: List[Dict[str, Any]] = []
    for r in raw:
        gov = (r["_id"] or "").strip()
        if not gov or _is_invalid_taxonomy_value(gov):
            continue
        rows.append({"governorate": gov, "total": r["total"], "active": r["active"], "inactive": r["inactive"]})
    rows.sort(key=lambda x: -x["total"])
    grand = {
        "total": sum(r["total"] for r in rows),
        "active": sum(r["active"] for r in rows),
        "inactive": sum(r["inactive"] for r in rows),
    }
    return {"department_id": department_id, "rows": rows, "grand_total": grand}


@api_router.get("/reports/disclosure/membership-detailed")
async def report_membership_detailed(
    department_id: str,
    user: Dict[str, Any] = Depends(require_user),
):
    """Governorates → committees → counts. Each governorate gets a subtotal,
    and there's a grand total at the very bottom."""
    pipeline = [
        {"$match": {"department_id": department_id}},
        {"$group": {
            "_id": {"g": "$governorate", "c": "$union_committee"},
            "total": {"$sum": 1},
            "active": {"$sum": {"$cond": [{"$eq": ["$status", "فعال"]}, 1, 0]}},
            "inactive": {"$sum": {"$cond": [{"$eq": ["$status", "فعال"]}, 0, 1]}},
        }},
    ]
    raw = await db.members.aggregate(pipeline).to_list(None)
    by_gov: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in raw:
        gov = (r["_id"].get("g") or "").strip()
        com = (r["_id"].get("c") or "").strip()
        if not gov or not com:
            continue
        if _is_invalid_taxonomy_value(gov) or _is_invalid_taxonomy_value(com):
            continue
        by_gov[gov].append({"union_committee": com, "total": r["total"], "active": r["active"], "inactive": r["inactive"]})

    governorates: List[Dict[str, Any]] = []
    for gov in sorted(by_gov.keys()):
        committees = sorted(by_gov[gov], key=lambda x: -x["total"])
        subtotal = {
            "total": sum(c["total"] for c in committees),
            "active": sum(c["active"] for c in committees),
            "inactive": sum(c["inactive"] for c in committees),
        }
        governorates.append({"governorate": gov, "committees": committees, "subtotal": subtotal})
    grand = {
        "total": sum(g["subtotal"]["total"] for g in governorates),
        "active": sum(g["subtotal"]["active"] for g in governorates),
        "inactive": sum(g["subtotal"]["inactive"] for g in governorates),
    }
    return {"department_id": department_id, "governorates": governorates, "grand_total": grand}


def _format_date_for_form(value: str) -> str:
    parsed = parse_date_value(value or "")
    if not parsed:
        return ""
    return parsed.strftime("%d / %m / %Y")


# ─── Live-sync change counter ────────────────────────────────────────────────
# A tiny per-collection "version counter" stored in a single Mongo document.
# Every write endpoint that mutates user-visible data calls _bump_change_counter
# so peer clients can detect change via the cheap /api/changes/state poll.
async def _bump_change_counter(*collections: str) -> None:
    if not collections:
        return
    try:
        inc = {f"counters.{c}": 1 for c in collections}
        await db["_change_state"].update_one(
            {"key": "counters"},
            {"$inc": inc, "$set": {"updated_at": now_iso()}},
            upsert=True,
        )
    except Exception:
        # Never let live-sync break a real operation.
        pass


async def _last_collected_subscription_month(department_id: str, governorate: str, union_committee: str) -> str:
    """Return the latest YYYY-MM where this committee has a 'تم التحصيل' subscription.

    Used to auto-fill 'آخر شهر سداد للاشتراكات' on the printed case research form.
    Returns '' if no collected subscription exists.
    """
    if not (department_id and governorate and union_committee):
        return ""
    pipeline = [
        {"$match": {
            "department_id": department_id,
            "governorate": governorate,
            "union_committee": union_committee,
            "status": "تم التحصيل",
            "is_dues_settlement": {"$ne": True},
            "subscription_month": {"$ne": ""},
        }},
        {"$group": {"_id": None, "max_month": {"$max": "$subscription_month"}}},
    ]
    try:
        result = await db.subscriptions.aggregate(pipeline).to_list(1)
    except Exception:
        return ""
    if result and result[0].get("max_month"):
        return str(result[0]["max_month"])
    return ""


def render_case_research_form_html(member: Dict[str, Any], mode: str = "print", dues_note: str = "", dues_note_clean: bool = False, last_paid_month: str = "", beneficiaries: list = None) -> str:
    """Thin wrapper around services.caseform_html.render_case_research_form_html.

    Passes the local _format_date_for_form helper so dates display as DD / MM / YYYY.
    """
    return _render_case_form_external(
        member, mode=mode, dues_note=dues_note, dues_note_clean=dues_note_clean,
        date_formatter=_format_date_for_form,
        last_paid_month=last_paid_month,
        beneficiaries=beneficiaries,
    )


@api_router.get("/members/{member_id}/case-form", response_class=HTMLResponse)
async def member_case_research_form(
    member_id: str,
    mode: str = "print",
    user: Dict[str, Any] = Depends(require_user),
):
    if mode not in {"print", "view"}:
        raise HTTPException(status_code=400, detail="وضع غير مدعوم")
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    enriched = await enrich_member_retirement(member)
    # Decide whether to show the "financial statement" note in the printed form.
    # Rule:
    #   - If member has a related aid record with print_dues_note = False → hide the note.
    #   - Otherwise → always include one of the two messages (owed or clean).
    aid = await db.aids.find_one(
        {"department_id": member.get("department_id"), "member_id": member_id},
        {"_id": 0},
        sort=[("triggered_at", -1)],
    )
    dues_note = ""
    dues_note_clean = False
    show_note = True
    if aid and aid.get("print_dues_note") is False:
        show_note = False
    if show_note:
        # Determine the period for the dues calculation.
        snapshot = aid or {
            "member_subscription_date": member.get("subscription_date", ""),
            "member_status_date": member.get("status_date", ""),
        }
        from_ym, to_ym = _aid_period_for_member(snapshot)
        dues = await _compute_committee_dues_for_period(
            department_id=member.get("department_id", ""),
            governorate=member.get("governorate", ""),
            union_committee=member.get("union_committee", ""),
            from_month=from_ym,
            to_month=to_ym,
        )
        owed = dues.get("owed_amount") or 0
        if owed > 0:
            owed_fmt = f"{owed:,.2f}".replace(",", "٬")
            if aid and aid.get("aid_type") == AID_TYPE_DEATH:
                label = "المتوفي"
            elif aid and aid.get("aid_type") == AID_TYPE_DISABILITY:
                label = "العضو (عجز كلي/جزئي منهي للخدمة)"
            else:
                label = "العضو"
            dues_note = (
                f"يوجد على {label} مستحقات مالية على لجنته النقابية بمبلغ "
                f"<strong>{owed_fmt} ج.م</strong> "
                f"عن الفترة من <strong>{from_ym}</strong> إلى <strong>{to_ym}</strong>."
            )
        else:
            dues_note = "المذكور مشترك بمشروع التكافل الاجتماعي ومسدد لكافة الاشتراكات حتى تاريخ الوفاة."
            dues_note_clean = True
    # Last collected subscription month for the member's committee (YYYY-MM)
    last_paid_month = await _last_collected_subscription_month(
        department_id=member.get("department_id", ""),
        governorate=member.get("governorate", ""),
        union_committee=member.get("union_committee", ""),
    )
    
    # جلب بيانات المستحقين إذا كانت موجودة
    beneficiaries_list = []
    if aid:
        beneficiaries_doc = await db.aid_beneficiaries.find_one({"aid_id": aid["id"]}, {"_id": 0})
        if beneficiaries_doc:
            beneficiaries_list = beneficiaries_doc.get("beneficiaries", [])
    
    return HTMLResponse(render_case_research_form_html(
        enriched, 
        mode=mode, 
        dues_note=dues_note, 
        dues_note_clean=dues_note_clean, 
        last_paid_month=last_paid_month,
        beneficiaries=beneficiaries_list
    ))


@api_router.get("/documents/{document_id}/download")
async def download_document(document_id: str):
    document = await db.documents.find_one({"id": document_id}, {"_id": 0})
    if not document:
        raise HTTPException(status_code=404, detail="الملف غير موجود")
    path = Path(document["path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="الملف غير موجود على الخادم")
    return FileResponse(path, media_type=document.get("content_type") or "application/pdf", filename=document.get("original_name") or path.name)


@api_router.get("/category-records", response_model=List[CategoryRecord])
async def list_category_records(
    department_id: str,
    category: str,
    search: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    if category not in FINANCIAL_CATEGORIES:
        raise HTTPException(status_code=400, detail="بوابة غير صالحة")
    query: Dict[str, Any] = {"department_id": department_id, "category": category}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"national_id": {"$regex": search, "$options": "i"}},
            {"membership_number": {"$regex": search, "$options": "i"}},
            {"reference_number": {"$regex": search, "$options": "i"}},
            {"subject": {"$regex": search, "$options": "i"}},
        ]
    records = await db.category_records.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return records


@api_router.post("/category-records", response_model=CategoryRecord)
async def create_category_record(input: CategoryRecordIn, user: Dict[str, Any] = Depends(require_user)):
    if input.category not in FINANCIAL_CATEGORIES:
        raise HTTPException(status_code=400, detail="بوابة غير صالحة")
    department = await db.departments.find_one({"id": input.department_id}, {"_id": 0})
    if not department:
        raise HTTPException(status_code=404, detail="الإدارة غير موجودة")
    normalized = input.model_dump()
    for key, value in normalized.items():
        if isinstance(value, str):
            normalized[key] = re.sub(r"\s+", " ", value.strip())
    document_file_name = ""
    document_original_name = ""
    document_url = ""
    if normalized.get("document_id"):
        document = await db.documents.find_one({"id": normalized["document_id"]}, {"_id": 0})
        if not document:
            raise HTTPException(status_code=404, detail="الملف المرفق غير موجود")
        document_file_name = document.get("file_name", "")
        document_original_name = document.get("original_name", "")
        document_url = f"/api/documents/{document['id']}/download"
    record = CategoryRecord(
        **normalized,
        document_file_name=document_file_name,
        document_original_name=document_original_name,
        document_url=document_url,
    )
    await db.category_records.insert_one(record.model_dump())
    return record


@api_router.delete("/category-records/{record_id}")
async def delete_category_record(record_id: str, user: Dict[str, Any] = Depends(require_user)):
    result = await db.category_records.delete_one({"id": record_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    return {"deleted": True}


@api_router.get("/category-records/summary")
async def category_records_summary(department_id: str, user: Dict[str, Any] = Depends(require_user)):
    counts: Dict[str, int] = {}
    for key in FINANCIAL_CATEGORIES:
        counts[key] = await db.category_records.count_documents({"department_id": department_id, "category": key})
    counts["subscriptions"] = await db.subscriptions.count_documents({"department_id": department_id, "is_dues_settlement": {"$ne": True}})
    counts["dues_settlements"] = await db.subscriptions.count_documents({"department_id": department_id, "is_dues_settlement": True})
    # Aids come from a dedicated collection (auto-created from member status changes)
    counts["aid_pending"] = await db.aids.count_documents({"department_id": department_id, "status": AID_STATUS_PENDING})
    counts["aid_disbursed"] = await db.aids.count_documents({"department_id": department_id, "status": AID_STATUS_DISBURSED})
    return {"counts": counts, "labels": FINANCIAL_CATEGORIES}


# -------------------- Letters (Generated) --------------------

LETTER_FOOTER_EMAIL_HTML = "gtuwa&#64;gtuwa.org.eg"


@api_router.get("/letters/generate")
async def generate_committee_letters(
    department_id: str,
    from_month: str,                              # YYYY-MM
    to_month: str,                                # YYYY-MM
    year_label: Optional[str] = None,             # e.g. "2025" (displayed in body)
    governorate: Optional[str] = None,
    committees: Optional[str] = None,             # comma-separated list of committees in selected governorate
    issue_date: Optional[str] = None,             # display date for "القاهرة في"
    user: Dict[str, Any] = Depends(require_user),
):
    """Generate one or more committee letters as a single A4 print-ready HTML."""
    department = await db.departments.find_one({"id": department_id}, {"_id": 0})
    if not department:
        raise HTTPException(status_code=404, detail="الإدارة غير موجودة")
    a = _ym_to_index(from_month)
    b = _ym_to_index(to_month)
    if a is None or b is None or a > b:
        raise HTTPException(status_code=400, detail="فترة غير صالحة")
    # Determine target committees: from explicit list or all in governorate(s)
    selected: List[tuple] = []
    if committees:
        if not governorate:
            raise HTTPException(status_code=400, detail="يجب تحديد المحافظة عند اختيار لجان")
        for c in [c.strip() for c in committees.split(",") if c.strip()]:
            selected.append((governorate, c))
    else:
        # All committees in selected governorate (or all governorates if none)
        member_query: Dict[str, Any] = {"department_id": department_id}
        if governorate:
            member_query["governorate"] = governorate
        pairs = set()
        async for m in db.members.find(member_query, {"_id": 0, "governorate": 1, "union_committee": 1}):
            g = (m.get("governorate") or "").strip()
            c = (m.get("union_committee") or "").strip()
            if g and c:
                pairs.add((g, c))
        selected = sorted(pairs)
    if not selected:
        raise HTTPException(status_code=404, detail="لا توجد لجان مطابقة")
    # Compute dues per committee
    letters: List[Dict[str, Any]] = []
    for (gov, com) in selected:
        dues = await _compute_committee_dues_for_period(department_id, gov, com, from_month, to_month)
        letters.append({
            "governorate": gov,
            "union_committee": com,
            "membership_size": dues.get("membership_size", 0),
            "owed_amount": dues.get("owed_amount", 0.0),
            "from_month": from_month,
            "to_month": to_month,
            "monthly_rate": dues.get("monthly_rate", COMMITTEE_MONTHLY_RATE),
        })
    html = render_letters_html(department.get("name") or "المشروع", letters, year_label or "", issue_date or today_ar())
    return HTMLResponse(html)


# -------------------- Excel Exports --------------------
# Heavy lifting moved to services/excel.py — keep thin aliases for backwards compat.
_build_xlsx = _build_xlsx_external
_xlsx_response = _xlsx_response_external



@api_router.get("/imports/members/template")
async def members_import_template(user: Dict[str, Any] = Depends(require_user)):
    """Download an empty Excel template with the supported columns for member import."""
    headers = ["الاسم", "الرقم القومي", "رقم العضوية", "المحافظة", "اللجنة النقابية",
               "تاريخ الميلاد", "تاريخ الاشتراك", "العنوان", "الهاتف", "اسم المستفيد",
               "الحالة", "تاريخ الحالة"]
    sample = [
        ["محمد أحمد علي", "29812345678901", "00123", "القاهرة", "اللجنة النقابية للزراعة",
         "1978-05-12", "2010-03-01", "شارع 9 - مدينة نصر", "01001234567", "فاطمة محمد",
         "فعال", "2010-03-01"],
    ]
    data = _build_xlsx([{"name": "قالب-أعضاء", "headers": headers, "rows": sample}])
    return _xlsx_response(data, "members-import-template.xlsx")


@api_router.post("/imports/members")
async def import_members(
    department_id: str,
    file: UploadFile = File(...),
    user: Dict[str, Any] = Depends(require_user),
):
    """Import members from an uploaded .xlsx file. Returns counts + per-row errors."""
    from openpyxl import load_workbook
    import io as _io

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="الرجاء رفع ملف Excel بصيغة .xlsx")
    dept = await db.departments.find_one({"id": department_id}, {"_id": 0})
    if not dept:
        raise HTTPException(status_code=404, detail="القسم غير موجود")

    raw = await file.read()
    try:
        wb = load_workbook(_io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"تعذر قراءة الملف: {exc}")
    ws = wb.active

    HEADER_MAP = {
        "الاسم": "name", "اسم العضو": "name", "name": "name",
        "الرقم القومي": "national_id", "national id": "national_id", "national_id": "national_id",
        "رقم العضوية": "membership_number", "رقم العضويه": "membership_number",
        "المحافظة": "governorate", "المحافظه": "governorate",
        "اللجنة النقابية": "union_committee", "اللجنه النقابيه": "union_committee", "اللجنة": "union_committee",
        "تاريخ الميلاد": "birth_date", "الميلاد": "birth_date",
        "تاريخ الاشتراك": "subscription_date", "الاشتراك": "subscription_date",
        "منضم بتاريخ": "subscription_date", "تاريخ الانضمام": "subscription_date", "الانضمام": "subscription_date",
        "العنوان": "address",
        "الهاتف": "phone", "التليفون": "phone", "الموبايل": "phone",
        "اسم المستفيد": "beneficiary_name", "المستفيد": "beneficiary_name",
        "تسليم قيمة مبلغ التكافل الي": "beneficiary_name",
        "تسليم قيمة مبلغ التكافل إلى": "beneficiary_name",
        "تسليم قيمة مبلغ التكافل": "beneficiary_name",
        "الحالة": "status",
        "تاريخ الحالة": "status_date",
    }

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="الملف فارغ")
    raw_headers = [str(h or "").strip() for h in rows[0]]
    col_to_field = {}
    for i, h in enumerate(raw_headers):
        key = h.strip().lower() if h else ""
        if key in HEADER_MAP:
            col_to_field[i] = HEADER_MAP[key]
        elif h in HEADER_MAP:
            col_to_field[i] = HEADER_MAP[h]
    if "name" not in col_to_field.values():
        raise HTTPException(status_code=400, detail="الملف يجب أن يحتوي على عمود 'الاسم'")

    created, updated, skipped = 0, 0, 0
    errors: List[Dict[str, Any]] = []
    now = now_iso()
    for r_idx, row in enumerate(rows[1:], start=2):
        if not row or all((cell is None or str(cell).strip() == "") for cell in row):
            continue
        rec: Dict[str, Any] = {}
        for col_idx, field in col_to_field.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is None:
                    continue
                if hasattr(val, "strftime"):
                    val = val.strftime("%Y-%m-%d")
                sval = str(val).strip()
                if sval:
                    rec[field] = sval
        # Required fields: name + governorate + union_committee.
        # Other fields (national_id, membership_number, ...) are optional.
        missing: List[str] = []
        if not rec.get("name"):
            missing.append("الاسم")
        if not rec.get("governorate"):
            missing.append("المحافظة")
        if not rec.get("union_committee"):
            missing.append("اللجنة النقابية")
        if missing:
            skipped += 1
            errors.append({"row": r_idx, "error": f"بيانات إلزامية مفقودة: {', '.join(missing)}"})
            continue
        rec.setdefault("status", "فعال")
        rec.setdefault("status_date", today_ar())
        rec["department_id"] = department_id
        rec["updated_at"] = now

        # Unified duplicate detection — same rule as the manual add form:
        #   - membership_number duplicated WITHIN the same committee blocks insert
        #   - 4-part name match + any disambiguator match blocks insert
        # On a match → UPDATE the existing record (Excel import is an upsert).
        normalised_name = " ".join(rec["name"].split())
        rec["name"] = normalised_name
        existing = await _find_member_duplicate(rec)
        try:
            if existing:
                await db.members.update_one({"id": existing["id"]}, {"$set": rec})
                updated += 1
            else:
                rec["id"] = build_id()
                rec["created_at"] = now
                await db.members.insert_one(rec)
                created += 1
        except Exception as exc:  # noqa: BLE001
            skipped += 1
            errors.append({"row": r_idx, "error": str(exc)[:200]})
    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": created + updated + skipped,
        "errors": errors[:50],
    }


@api_router.get("/exports/members")
async def export_members(department_id: str, user: Dict[str, Any] = Depends(require_user)):
    docs = await db.members.find({"department_id": department_id}, {"_id": 0}).to_list(100000)
    enriched = await enrich_members_retirement(docs)
    headers = ["م", "الاسم", "الرقم القومي", "رقم العضوية", "المحافظة", "اللجنة النقابية",
               "تاريخ الميلاد", "تاريخ الاشتراك", "العنوان", "الهاتف", "اسم المستفيد",
               "الحالة", "تاريخ الحالة", "بلغ سن المعاش؟"]
    rows = []
    for i, m in enumerate(enriched, 1):
        rows.append([
            i, m.get("name", ""), m.get("national_id", ""), m.get("membership_number", ""),
            m.get("governorate", ""), m.get("union_committee", ""), m.get("birth_date", ""),
            m.get("subscription_date", ""), m.get("address", ""), m.get("phone", ""),
            m.get("beneficiary_name", ""), m.get("status", "فعال"), m.get("status_date", ""),
            "نعم" if m.get("retirement_due") else "لا",
        ])
    data = _build_xlsx([{"name": "العضوية", "headers": headers, "rows": rows}])
    return _xlsx_response(data, f"members-{department_id[:8]}.xlsx")


@api_router.get("/exports/subscriptions")
async def export_subscriptions(department_id: str, user: Dict[str, Any] = Depends(require_user)):
    docs = await db.subscriptions.find({"department_id": department_id}, {"_id": 0}).sort("date", -1).to_list(20000)
    headers = ["م", "رقم الإيصال", "المبلغ", "المحافظة", "اللجنة", "طريقة الدفع",
               "شهر الاشتراك", "تاريخ الاشتراك", "الحالة", "المُسجِّل"]
    rows = []
    for i, s in enumerate(docs, 1):
        rows.append([
            i, s.get("receipt_number", ""), float(s.get("amount") or 0.0),
            s.get("governorate", ""), s.get("union_committee", ""), s.get("payment_method", ""),
            s.get("subscription_month", ""), s.get("date", ""), s.get("status", ""),
            s.get("recorded_by_name", ""),
        ])
    data = _build_xlsx([{"name": "الاشتراكات", "headers": headers, "rows": rows}])
    return _xlsx_response(data, f"subscriptions-{department_id[:8]}.xlsx")


@api_router.get("/exports/aids")
async def export_aids(department_id: str, status: Optional[str] = None, user: Dict[str, Any] = Depends(require_user)):
    query: Dict[str, Any] = {"department_id": department_id}
    if status:
        query["status"] = status
    docs = await db.aids.find(query, {"_id": 0}).sort("triggered_at", -1).to_list(20000)
    headers = ["م", "نوع الإعانة", "حالة الإعانة", "اسم العضو", "الرقم القومي", "رقم العضوية",
               "المحافظة", "اللجنة", "تاريخ الحالة", "رقم الشيك", "تاريخ الشيك", "البنك",
               "المبلغ", "المستفيدون", "تاريخ الصرف"]
    rows = []
    status_label = {AID_STATUS_PENDING: "في انتظار الموافقة", AID_STATUS_DISBURSED: "تم الصرف"}
    for i, a in enumerate(docs, 1):
        rows.append([
            i, a.get("aid_type", ""), status_label.get(a.get("status", ""), a.get("status", "")),
            a.get("member_name", ""), a.get("member_national_id", ""),
            a.get("member_membership_number", ""), a.get("member_governorate", ""),
            a.get("member_union_committee", ""), a.get("member_status_date", ""),
            a.get("cheque_number", ""), a.get("cheque_date", ""), a.get("cheque_bank", ""),
            float(a.get("amount") or 0.0),
            "، ".join(a.get("beneficiaries") or []),
            (a.get("disbursed_at") or "")[:10],
        ])
    data = _build_xlsx([{"name": "الإعانات", "headers": headers, "rows": rows}])
    return _xlsx_response(data, f"aids-{department_id[:8]}.xlsx")


@api_router.get("/exports/dues")
async def export_dues(
    department_id: str,
    from_month: Optional[str] = None,
    to_month: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    result = await committees_dues(department_id=department_id, from_month=from_month, to_month=to_month, governorate=None, user=user)
    headers = ["م", "المحافظة", "اللجنة", "حجم العضوية", "الفعّال", "المُلحَق", "الأشهر",
               "القيمة الشهرية", "المستحق (ج.م)", "المُحصَّل (ج.م)", "المتبقي (ج.م)", "الحالة"]
    rows = []
    for i, r in enumerate(result.get("rows") or [], 1):
        rows.append([
            i, r.get("governorate", ""), r.get("union_committee", ""),
            r.get("membership_size", 0), r.get("active_size", 0), r.get("held_back_size", 0),
            r.get("months", 0), r.get("monthly_rate", 0),
            float(r.get("expected_amount") or 0), float(r.get("paid_amount") or 0),
            float(r.get("owed_amount") or 0),
            "مسددة" if r.get("settled") else ("دفع زائد" if (r.get("credit_amount") or 0) > 0 else "عليها متأخرات"),
        ])
    gov_headers = ["م", "المحافظة", "عدد اللجان", "حجم العضوية", "المستحق", "المُحصَّل", "المتبقي"]
    gov_rows = []
    for i, g in enumerate(result.get("governorate_totals") or [], 1):
        gov_rows.append([
            i, g.get("governorate", ""), g.get("committees_count", 0), g.get("membership_size", 0),
            float(g.get("expected_amount") or 0), float(g.get("paid_amount") or 0), float(g.get("owed_amount") or 0),
        ])
    sheets = [
        {"name": "تفصيل اللجان", "headers": headers, "rows": rows},
        {"name": "إجمالي المحافظات", "headers": gov_headers, "rows": gov_rows},
    ]
    data = _build_xlsx(sheets)
    return _xlsx_response(data, f"dues-{department_id[:8]}.xlsx")


@api_router.get("/exports/aids-report")
async def export_aids_report(
    department_id: str,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    governorate: Optional[str] = None,
    aid_type: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    report = await aids_report(department_id, from_date, to_date, governorate, aid_type, user)
    gov_headers = ["م", "المحافظة", "عدد الإعانات", "المبلغ (ج.م)"]
    gov_rows = [[i, g["name"], g["count"], g["amount"]] for i, g in enumerate(report.get("by_governorate") or [], 1)]
    type_headers = ["م", "نوع الإعانة", "العدد", "المبلغ (ج.م)"]
    type_rows = [[i, t["name"], t["count"], t["amount"]] for i, t in enumerate(report.get("by_aid_type") or [], 1)]
    month_headers = ["م", "الشهر", "العدد", "المبلغ (ج.م)"]
    month_rows = [[i, m["name"], m["count"], m["amount"]] for i, m in enumerate(report.get("by_month") or [], 1)]
    detail_headers = ["م", "نوع الإعانة", "اسم العضو", "الرقم القومي", "المحافظة", "اللجنة",
                      "رقم الشيك", "تاريخ الشيك", "البنك", "المبلغ", "المستفيدون"]
    detail_rows = [[i, r["aid_type"], r["member_name"], r["member_national_id"],
                    r["member_governorate"], r["member_union_committee"], r["cheque_number"],
                    r["cheque_date"], r["cheque_bank"], float(r["amount"] or 0),
                    "، ".join(r.get("beneficiaries") or [])] for i, r in enumerate(report.get("rows") or [], 1)]
    sheets = [
        {"name": "حسب المحافظة", "headers": gov_headers, "rows": gov_rows},
        {"name": "حسب نوع الإعانة", "headers": type_headers, "rows": type_rows},
        {"name": "حسب الشهر", "headers": month_headers, "rows": month_rows},
        {"name": "التفاصيل", "headers": detail_headers, "rows": detail_rows},
    ]
    data = _build_xlsx(sheets)
    return _xlsx_response(data, f"aids-report-{from_date or 'all'}-{to_date or 'all'}.xlsx")




@api_router.get("/aids")
async def list_aids(
    department_id: str,
    status: Optional[str] = None,  # pending | disbursed
    aid_type: Optional[str] = None,
    search: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    query: Dict[str, Any] = {"department_id": department_id}
    if status:
        if status not in {AID_STATUS_PENDING, AID_STATUS_DISBURSED}:
            raise HTTPException(status_code=400, detail="حالة إعانة غير صالحة")
        query["status"] = status
    if aid_type:
        query["aid_type"] = aid_type
    if search:
        s = search.strip()
        query["$or"] = [
            {"member_name": {"$regex": s, "$options": "i"}},
            {"member_national_id": {"$regex": s, "$options": "i"}},
            {"member_membership_number": {"$regex": s, "$options": "i"}},
        ]
    docs = await db.aids.find(query, {"_id": 0}).sort("triggered_at", -1).to_list(5000)
    results: List[Dict[str, Any]] = []
    for d in docs:
        record = AidRecord(**d).model_dump()
        from_ym, to_ym = _aid_period_for_member(d)
        dues = await _compute_committee_dues_for_period(
            department_id=d.get("department_id", ""),
            governorate=d.get("member_governorate", ""),
            union_committee=d.get("member_union_committee", ""),
            from_month=from_ym,
            to_month=to_ym,
        )
        record["committee_dues"] = dues
        results.append(record)
    return results


class AidPrintNoteToggle(BaseModel):
    enabled: bool


@api_router.patch("/aids/{aid_id}/print-note", response_model=AidRecord)
async def toggle_aid_print_note(aid_id: str, payload: AidPrintNoteToggle, user: Dict[str, Any] = Depends(require_user)):
    aid = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    if not aid:
        raise HTTPException(status_code=404, detail="الإعانة غير موجودة")
    await db.aids.update_one({"id": aid_id}, {"$set": {"print_dues_note": bool(payload.enabled), "updated_at": now_iso()}})
    updated = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    return AidRecord(**updated)


@api_router.post("/aids/{aid_id}/disburse", response_model=AidRecord)
async def disburse_aid(aid_id: str, payload: AidDisburseIn, user: Dict[str, Any] = Depends(require_user)):
    aid = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    if not aid:
        raise HTTPException(status_code=404, detail="الإعانة غير موجودة")
    if aid.get("status") == AID_STATUS_DISBURSED:
        raise HTTPException(status_code=400, detail="تم صرف هذه الإعانة بالفعل")
    cheque_number = (payload.cheque_number or "").strip()
    cheque_bank = (payload.cheque_bank or "").strip()
    cheque_date = (payload.cheque_date or "").strip()
    if not cheque_number or not cheque_bank or not cheque_date:
        raise HTTPException(status_code=400, detail="رقم الشيك وتاريخه واسم البنك مطلوبة")
    if payload.amount is None or payload.amount <= 0:
        raise HTTPException(status_code=400, detail="المبلغ يجب أن يكون أكبر من صفر")
    beneficiaries = [b.strip() for b in (payload.beneficiaries or []) if (b or "").strip()]
    if not beneficiaries:
        raise HTTPException(status_code=400, detail="يجب إضافة مستفيد واحد على الأقل")
    update = {
        "status": AID_STATUS_DISBURSED,
        "cheque_number": cheque_number,
        "cheque_date": cheque_date,
        "cheque_bank": cheque_bank,
        "amount": float(payload.amount),
        "beneficiaries": beneficiaries,
        "disbursed_at": now_iso(),
        "disbursed_by_id": user.get("id", ""),
        "disbursed_by_name": user.get("name") or user.get("username") or "",
        "updated_at": now_iso(),
    }
    await db.aids.update_one({"id": aid_id}, {"$set": update})
    updated = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    return AidRecord(**updated)


@api_router.post("/aids/{aid_id}/restore-member", response_model=Dict[str, Any])
async def restore_member_to_active(aid_id: str, user: Dict[str, Any] = Depends(require_user)):
    """Restore a member back to active status and remove the aid (pending or disbursed).

    For disbursed aids, requires admin privileges since it impacts financial records.
    """
    aid = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    if not aid:
        raise HTTPException(status_code=404, detail="الإعانة غير موجودة")
    # Disbursed aids require admin (extra audit gate)
    if aid.get("status") == AID_STATUS_DISBURSED:
        if user.get("role") not in ("super_admin", "admin"):
            raise HTTPException(status_code=403, detail="يلزم صلاحية أدمن لاسترجاع عضو لإعانة مصروفة")
    member_id = aid.get("member_id")
    if not member_id:
        raise HTTPException(status_code=400, detail="السجل لا يحتوي على رقم عضو")
    member = await db.members.find_one({"id": member_id}, {"_id": 0})
    if not member:
        await db.aids.delete_one({"id": aid_id})
        return {"ok": True, "deleted_aid": True, "member_restored": False, "message": "العضو غير موجود — تم حذف الإعانة فقط."}
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    await db.members.update_one(
        {"id": member_id},
        {"$set": {"status": "فعال", "status_date": today, "updated_at": now_iso()}},
    )
    await db.aids.delete_one({"id": aid_id})
    return {
        "ok": True,
        "deleted_aid": True,
        "member_restored": True,
        "was_disbursed": aid.get("status") == AID_STATUS_DISBURSED,
        "member_id": member_id,
        "member_name": member.get("name"),
        "message": f"تم رد العضو {member.get('name')} إلى الحالة فعّال وحذف الإعانة.",
    }


@api_router.delete("/aids/{aid_id}")
async def delete_aid(aid_id: str, user: Dict[str, Any] = Depends(require_admin)):
    """Delete an aid record (admin only). Works for both pending and disbursed aids."""
    aid = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    if not aid:
        raise HTTPException(status_code=404, detail="الإعانة غير موجودة")
    await db.aids.delete_one({"id": aid_id})
    return {"deleted": True, "id": aid_id, "was_disbursed": aid.get("status") == AID_STATUS_DISBURSED}


@api_router.post("/aids/{aid_id}/calculate-beneficiaries")
async def calculate_aid_beneficiaries(
    aid_id: str,
    payload: Dict[str, Any] = Body(...),
    user: Dict[str, Any] = Depends(require_user),
):
    """حساب توزيع الإعانة على المستحقين"""
    aid = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    if not aid:
        raise HTTPException(status_code=404, detail="الإعانة غير موجودة")
    
    total_amount = payload.get("total_amount", 0)
    beneficiaries_input = payload.get("beneficiaries", [])
    
    if total_amount <= 0:
        raise HTTPException(status_code=400, detail="أصل المبلغ يجب أن يكون أكبر من صفر")
    
    if not beneficiaries_input:
        raise HTTPException(status_code=400, detail="يجب إدخال مستحق واحد على الأقل")
    
    # حساب التوزيع
    result = calculate_inheritance(total_amount, beneficiaries_input)
    
    return result


@api_router.post("/aids/{aid_id}/save-beneficiaries")
async def save_aid_beneficiaries(
    aid_id: str,
    payload: AidBeneficiariesData = Body(...),
    user: Dict[str, Any] = Depends(require_user),
):
    """حفظ بيانات المستحقين للإعانة"""
    aid = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    if not aid:
        raise HTTPException(status_code=404, detail="الإعانة غير موجودة")
    
    # حفظ في collection منفصلة
    beneficiaries_doc = {
        "aid_id": aid_id,
        "member_id": aid.get("member_id"),
        "total_amount": payload.total_amount,
        "beneficiaries": [b.model_dump() for b in payload.beneficiaries],
        "summary_explanation": payload.summary_explanation,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    
    # حذف السجل القديم إن وجد
    await db.aid_beneficiaries.delete_many({"aid_id": aid_id})
    
    # إدراج السجل الجديد
    await db.aid_beneficiaries.insert_one(beneficiaries_doc)
    
    return {"saved": True, "aid_id": aid_id, "beneficiaries_count": len(payload.beneficiaries)}


@api_router.get("/aids/{aid_id}/beneficiaries")
async def get_aid_beneficiaries(aid_id: str, user: Dict[str, Any] = Depends(require_user)):
    """جلب بيانات المستحقين للإعانة"""
    beneficiaries_doc = await db.aid_beneficiaries.find_one({"aid_id": aid_id}, {"_id": 0})
    
    if not beneficiaries_doc:
        return {"found": False, "beneficiaries": [], "summary_explanation": ""}
    
    return {
        "found": True,
        "total_amount": beneficiaries_doc.get("total_amount", 0),
        "beneficiaries": beneficiaries_doc.get("beneficiaries", []),
        "summary_explanation": beneficiaries_doc.get("summary_explanation", ""),
    }


@api_router.post("/aids/{aid_id}/recalculate")
async def recalculate_aid_dues(aid_id: str, user: Dict[str, Any] = Depends(require_user)):
    """Force a fresh recalculation of an aid's committee dues + refresh its member snapshot.

    Why this exists: aid records were created with a copy of the member's data at
    the moment of creation. Two things can drift afterwards:
      1) The member's own profile (governorate, committee, subscription_date,
         status_date, etc.) may have been edited.
      2) The global financial floor rule (FINANCIAL_START_MONTH = 2025-01) was
         introduced AFTER some legacy aid records were created.

    This endpoint pulls the latest member data, syncs the aid's snapshot fields,
    bumps `updated_at`, and returns the freshly computed dues — same shape as
    the `committee_dues` field used in the list view.
    """
    aid = await db.aids.find_one({"id": aid_id}, {"_id": 0})
    if not aid:
        raise HTTPException(status_code=404, detail="الإعانة غير موجودة")
    member = await db.members.find_one({"id": aid.get("member_id")}, {"_id": 0})
    snapshot_update: Dict[str, Any] = {"updated_at": now_iso()}
    if member:
        # Refresh denormalized member fields so the aid reflects current truth.
        snapshot_update.update({
            "member_name": member.get("name", ""),
            "member_national_id": member.get("national_id", ""),
            "member_membership_number": member.get("membership_number", ""),
            "member_governorate": member.get("governorate", ""),
            "member_union_committee": member.get("union_committee", ""),
            "member_birth_date": member.get("birth_date", ""),
            "member_subscription_date": member.get("subscription_date", ""),
            "member_address": member.get("address", ""),
            "member_phone": member.get("phone", ""),
            "member_beneficiary_name": member.get("beneficiary_name", ""),
            "member_status_date": member.get("status_date", ""),
        })
    await db.aids.update_one({"id": aid_id}, {"$set": snapshot_update})
    refreshed = await db.aids.find_one({"id": aid_id}, {"_id": 0}) or aid
    from_ym, to_ym = _aid_period_for_member(refreshed)
    dues = await _compute_committee_dues_for_period(
        department_id=refreshed.get("department_id", ""),
        governorate=refreshed.get("member_governorate", ""),
        union_committee=refreshed.get("member_union_committee", ""),
        from_month=from_ym,
        to_month=to_ym,
    )
    # Note: the global _live_sync_middleware already bumps the 'aids' counter
    # for any successful POST /api/aids/*, so we deliberately do NOT call
    # _bump_change_counter('aids') here (it would cause peer clients to
    # double-refetch on every recalc).
    return {
        "ok": True,
        "aid": AidRecord(**refreshed).model_dump(),
        "committee_dues": dues,
        "member_refreshed": bool(member),
    }


@api_router.get("/aids/summary")
async def aids_summary(department_id: str, user: Dict[str, Any] = Depends(require_user)):
    pending = await db.aids.count_documents({"department_id": department_id, "status": AID_STATUS_PENDING})
    disbursed = await db.aids.count_documents({"department_id": department_id, "status": AID_STATUS_DISBURSED})
    pipeline = [
        {"$match": {"department_id": department_id, "status": AID_STATUS_DISBURSED}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}},
    ]
    total_disbursed_amount = 0.0
    async for doc in db.aids.aggregate(pipeline):
        total_disbursed_amount = float(doc.get("total") or 0.0)
    return {
        "pending": pending,
        "disbursed": disbursed,
        "total_disbursed_amount": round(total_disbursed_amount, 2),
    }


@api_router.get("/aids/report")
async def aids_report(
    department_id: str,
    from_date: Optional[str] = None,    # YYYY-MM-DD (uses cheque_date or disbursed_at)
    to_date: Optional[str] = None,
    governorate: Optional[str] = None,
    aid_type: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    """Periodic report on disbursed aids grouped by governorate, aid type, and month."""
    query: Dict[str, Any] = {"department_id": department_id, "status": AID_STATUS_DISBURSED}
    if governorate:
        query["member_governorate"] = governorate
    if aid_type:
        query["aid_type"] = aid_type
    docs = await db.aids.find(query, {"_id": 0}).sort("cheque_date", -1).to_list(20000)

    # Filter by date in Python (cheque_date is YYYY-MM-DD)
    def _in_range(d: str) -> bool:
        if not from_date and not to_date:
            return True
        if not d:
            return False
        if from_date and d < from_date:
            return False
        if to_date and d > to_date:
            return False
        return True

    rows = [d for d in docs if _in_range(d.get("cheque_date") or d.get("disbursed_at", "")[:10])]

    by_gov: Dict[str, Dict[str, float]] = {}
    by_type: Dict[str, Dict[str, float]] = {}
    by_month: Dict[str, Dict[str, float]] = {}
    total_amount = 0.0
    total_beneficiaries = 0
    for r in rows:
        amount = float(r.get("amount") or 0.0)
        total_amount += amount
        total_beneficiaries += len(r.get("beneficiaries") or [])
        gov = (r.get("member_governorate") or "غير محدد").strip()
        t = (r.get("aid_type") or "غير محدد").strip()
        cd = (r.get("cheque_date") or r.get("disbursed_at", "")[:10] or "")[:7] or "غير محدد"
        for bucket, key in ((by_gov, gov), (by_type, t), (by_month, cd)):
            acc = bucket.setdefault(key, {"name": key, "count": 0, "amount": 0.0})
            acc["count"] += 1
            acc["amount"] += amount

    def _list(d: Dict[str, Dict[str, float]]):
        out = []
        for k, v in d.items():
            out.append({"name": v["name"], "count": int(v["count"]), "amount": round(v["amount"], 2)})
        out.sort(key=lambda x: (-x["amount"], x["name"]))
        return out

    return {
        "from_date": from_date or "",
        "to_date": to_date or "",
        "governorate": governorate or "",
        "aid_type": aid_type or "",
        "totals": {
            "count": len(rows),
            "amount": round(total_amount, 2),
            "beneficiaries": total_beneficiaries,
        },
        "by_governorate": _list(by_gov),
        "by_aid_type": _list(by_type),
        "by_month": sorted(_list(by_month), key=lambda x: x["name"]),
        "rows": [
            {
                "id": r.get("id"),
                "aid_type": r.get("aid_type"),
                "member_name": r.get("member_name"),
                "member_national_id": r.get("member_national_id"),
                "member_governorate": r.get("member_governorate"),
                "member_union_committee": r.get("member_union_committee"),
                "cheque_number": r.get("cheque_number"),
                "cheque_date": r.get("cheque_date"),
                "cheque_bank": r.get("cheque_bank"),
                "amount": float(r.get("amount") or 0.0),
                "beneficiaries": r.get("beneficiaries") or [],
                "disbursed_at": (r.get("disbursed_at") or "")[:10],
            }
            for r in rows
        ],
    }


# -------------------- Subscriptions --------------------

@api_router.get("/subscriptions")
async def list_subscriptions(
    department_id: str,
    status: Optional[str] = None,
    search: Optional[str] = None,
    is_dues_settlement: Optional[bool] = None,
    page: int = 1,
    page_size: int = 20,
    user: Dict[str, Any] = Depends(require_user),
):
    query: Dict[str, Any] = {"department_id": department_id}
    if status:
        if status not in SUBSCRIPTION_STATUSES:
            raise HTTPException(status_code=400, detail="حالة غير صالحة")
        query["status"] = status
    if is_dues_settlement is not None:
        if is_dues_settlement:
            query["is_dues_settlement"] = True
        else:
            query["is_dues_settlement"] = {"$ne": True}
    if search:
        # Normalize Arabic and English numbers for better search
        search_normalized = search.strip()
        # Convert Arabic numbers to English
        arabic_to_english = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')
        search_english = search_normalized.translate(arabic_to_english)
        # Convert English numbers to Arabic
        english_to_arabic = str.maketrans('0123456789', '٠١٢٣٤٥٦٧٨٩')
        search_arabic = search_normalized.translate(english_to_arabic)
        
        # Search in both formats (original, English numbers, Arabic numbers)
        query["$or"] = [
            {"permit_number": {"$regex": search_normalized, "$options": "i"}},
            {"governorate": {"$regex": search_normalized, "$options": "i"}},
            {"union_committee": {"$regex": search_normalized, "$options": "i"}},
            {"cheque_number": {"$regex": search_normalized, "$options": "i"}},
            {"electronic_reference": {"$regex": search_normalized, "$options": "i"}},
            {"payment_details": {"$regex": search_normalized, "$options": "i"}},
            {"notes": {"$regex": search_normalized, "$options": "i"}},
        ]
        
        # Add searches with converted numbers if different from original
        if search_english != search_normalized:
            query["$or"].extend([
                {"permit_number": {"$regex": search_english, "$options": "i"}},
                {"cheque_number": {"$regex": search_english, "$options": "i"}},
                {"electronic_reference": {"$regex": search_english, "$options": "i"}},
            ])
        
        if search_arabic != search_normalized:
            query["$or"].extend([
                {"permit_number": {"$regex": search_arabic, "$options": "i"}},
                {"cheque_number": {"$regex": search_arabic, "$options": "i"}},
                {"electronic_reference": {"$regex": search_arabic, "$options": "i"}},
            ])
    
    # Get total count
    total = await db.subscriptions.count_documents(query)
    
    # Fetch items sorted by creation date (newest first) with pagination
    skip = (page - 1) * page_size
    items = await db.subscriptions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    return {
        "data": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if page_size else 1,
    }


@api_router.get("/subscriptions/lookup-reference")
async def lookup_electronic_reference(
    department_id: str,
    electronic_reference: str,
    exclude_id: Optional[str] = None,
    kind: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    """Return the existing subscription record (if any) that already uses the
    given electronic payment reference within the department. Used by the
    frontend to alert about duplicates BEFORE submission.

    `kind` is "number" or "text" — when "text", we skip the lookup entirely
    because text labels are explicitly allowed to repeat.
    """
    if _normalise_reference_kind(kind) == ELECTRONIC_REFERENCE_KIND_TEXT:
        return {"found": False}
    ref = (electronic_reference or "").strip()
    if not ref:
        return {"found": False}
    query: Dict[str, Any] = {
        "department_id": department_id,
        "electronic_reference": ref,
        "payment_method": "دفع الكتروني",
        # Only collide with rows whose stored kind is "number" — never with a
        # row that the operator explicitly marked as a manual text label.
        "$or": [
            {"electronic_reference_kind": ELECTRONIC_REFERENCE_KIND_NUMBER},
            {"electronic_reference_kind": {"$exists": False}},  # legacy rows
        ],
    }
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    existing = await db.subscriptions.find_one(query, {"_id": 0})
    return {"found": bool(existing), "record": existing}


def _normalize_subscription_input(payload: SubscriptionIn) -> Dict[str, Any]:
    data = payload.model_dump()
    for key, value in list(data.items()):
        if isinstance(value, str):
            data[key] = re.sub(r"\s+", " ", value.strip())
    if data.get("payment_method") not in SUBSCRIPTION_PAYMENT_METHODS:
        raise HTTPException(status_code=400, detail="طريقة الدفع غير مدعومة")
    if data.get("payment_method") == "دفع الكتروني":
        data["cheque_number"] = ""
        data["cheque_bank"] = ""
        data["cheque_date"] = ""
        data["electronic_reference_kind"] = _normalise_reference_kind(data.get("electronic_reference_kind"))
    else:
        data["electronic_reference"] = ""
        data["electronic_reference_kind"] = ELECTRONIC_REFERENCE_KIND_NUMBER
    if data.get("status") not in SUBSCRIPTION_STATUSES:
        raise HTTPException(status_code=400, detail="حالة غير مدعومة")
    if data.get("status") == "لا يخص مشروع التكافل":
        target = data.get("not_takaful_target")
        if target not in SUBSCRIPTION_NOT_TAKAFUL_TARGETS:
            raise HTTPException(status_code=400, detail="حدد الجهة (النقابة العامة أو جهة أخرى)")
        if target == "جهة أخرى" and not (data.get("not_takaful_other") or "").strip():
            raise HTTPException(status_code=400, detail="حدد اسم الجهة الأخرى")
    else:
        data["not_takaful_target"] = ""
        data["not_takaful_other"] = ""
    try:
        data["amount"] = float(data.get("amount") or 0)
    except (TypeError, ValueError):
        data["amount"] = 0.0
    return data


# Two flavours for the "electronic_reference" cell:
#   - "number" → real transaction id issued by a bank or payment gateway;
#               must be unique within the department.
#   - "text"   → a manual descriptive label entered by the operator
#               (e.g. "طبقا لكشف الحساب - بنك التنمية الصناعية"); MAY repeat
#               across rows because it is not a transactional identifier.
# The kind is set explicitly by the UI via a radio next to the field, so we
# do NOT auto-detect from the content — operators stay in control.
ELECTRONIC_REFERENCE_KIND_NUMBER = "number"
ELECTRONIC_REFERENCE_KIND_TEXT = "text"
ELECTRONIC_REFERENCE_KINDS = {
    ELECTRONIC_REFERENCE_KIND_NUMBER,
    ELECTRONIC_REFERENCE_KIND_TEXT,
}


def _normalise_reference_kind(value: Optional[str]) -> str:
    """Coerce missing / unknown values to the safe default ('number')."""
    v = (value or "").strip().lower()
    if v in ELECTRONIC_REFERENCE_KINDS:
        return v
    return ELECTRONIC_REFERENCE_KIND_NUMBER


async def _ensure_unique_electronic_reference(data: Dict[str, Any], exclude_id: Optional[str] = None):
    """Raise 409 if the electronic_reference already exists in the same
    department for another subscription. The error detail includes the
    conflicting record so the UI can show its details.

    Whitelist: a few values are *manual labels* (not real transaction IDs)
    that legitimately repeat across many subscription rows — e.g. payments
    made via bank statement reconciliation. We skip the uniqueness check for
    those instead of forcing operators to invent fake unique strings.
    """
    if data.get("payment_method") != "دفع الكتروني":
        return
    # Manual text labels are explicitly allowed to repeat — only enforce
    # uniqueness on real transaction numbers.
    if _normalise_reference_kind(data.get("electronic_reference_kind")) == ELECTRONIC_REFERENCE_KIND_TEXT:
        return
    ref = (data.get("electronic_reference") or "").strip()
    if not ref:
        return
    query: Dict[str, Any] = {
        "department_id": data.get("department_id"),
        "electronic_reference": ref,
        "payment_method": "دفع الكتروني",
        # Same nuance as the lookup endpoint: never collide with rows that
        # were stored as a manual text label.
        "$or": [
            {"electronic_reference_kind": ELECTRONIC_REFERENCE_KIND_NUMBER},
            {"electronic_reference_kind": {"$exists": False}},  # legacy rows
        ],
    }
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    existing = await db.subscriptions.find_one(query, {"_id": 0})
    if existing:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "رقم الدفع الإلكتروني مستخدم مسبقاً في سجل آخر",
                "code": "duplicate_electronic_reference",
                "existing": existing,
            },
        )


@api_router.post("/subscriptions", response_model=Subscription)
async def create_subscription(payload: SubscriptionIn, user: Dict[str, Any] = Depends(require_user)):
    department = await db.departments.find_one({"id": payload.department_id}, {"_id": 0})
    if not department:
        raise HTTPException(status_code=404, detail="الإدارة غير موجودة")
    data = _normalize_subscription_input(payload)
    await _ensure_unique_electronic_reference(data)
    record = Subscription(
        **data,
        recorded_by_id=user.get("id", ""),
        recorded_by_name=user.get("display_name") or user.get("username") or "",
    )
    await db.subscriptions.insert_one(record.model_dump())
    return record


@api_router.patch("/subscriptions/{subscription_id}/status", response_model=Subscription)
async def update_subscription_status(subscription_id: str, payload: SubscriptionStatusUpdate, user: Dict[str, Any] = Depends(require_user)):
    if payload.status not in SUBSCRIPTION_STATUSES:
        raise HTTPException(status_code=400, detail="حالة غير صالحة")
    record = await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    update = {"status": payload.status, "updated_at": now_iso()}
    if payload.status == "لا يخص مشروع التكافل":
        target = (payload.not_takaful_target or "").strip()
        if target not in SUBSCRIPTION_NOT_TAKAFUL_TARGETS:
            raise HTTPException(status_code=400, detail="حدد الجهة (النقابة العامة أو جهة أخرى)")
        if target == "جهة أخرى" and not (payload.not_takaful_other or "").strip():
            raise HTTPException(status_code=400, detail="حدد اسم الجهة الأخرى")
        update["not_takaful_target"] = target
        update["not_takaful_other"] = (payload.not_takaful_other or "").strip()
    else:
        update["not_takaful_target"] = ""
        update["not_takaful_other"] = ""
    await db.subscriptions.update_one({"id": subscription_id}, {"$set": update})
    return await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})


@api_router.put("/subscriptions/{subscription_id}", response_model=Subscription)
async def update_subscription(subscription_id: str, payload: SubscriptionIn, user: Dict[str, Any] = Depends(require_user)):
    """Full edit of a subscription record (all fields)."""
    existing = await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    data = _normalize_subscription_input(payload)
    await _ensure_unique_electronic_reference(data, exclude_id=subscription_id)
    update = {**data, "updated_at": now_iso()}
    # Preserve original metadata
    update["id"] = subscription_id
    update["created_at"] = existing.get("created_at")
    update["recorded_by_id"] = existing.get("recorded_by_id", "")
    update["recorded_by_name"] = existing.get("recorded_by_name", "")
    await db.subscriptions.update_one({"id": subscription_id}, {"$set": update})
    return await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})


@api_router.get("/subscriptions/{subscription_id}/print", response_class=HTMLResponse)
async def print_subscription(subscription_id: str, user: Dict[str, Any] = Depends(require_user)):
    """Render an A4-print HTML view of a single subscription voucher."""
    record = await db.subscriptions.find_one({"id": subscription_id}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    is_settlement = bool(record.get("is_dues_settlement"))
    title = "بيان تسوية مستحقات" if is_settlement else "بيان اشتراك"
    method = record.get("payment_method") or "-"
    method_details_html = ""
    if method == "شيك":
        method_details_html = (
            f"<tr><th>رقم الشيك</th><td>{record.get('cheque_number') or '-'}</td>"
            f"<th>البنك</th><td>{record.get('cheque_bank') or '-'}</td></tr>"
            f"<tr><th>تاريخ الشيك</th><td colspan='3'>{record.get('cheque_date') or '-'}</td></tr>"
        )
    else:
        method_details_html = (
            f"<tr><th>رقم/مرجع الدفع الإلكتروني</th><td colspan='3'>{record.get('electronic_reference') or '-'}</td></tr>"
        )
    not_takaful_html = ""
    if record.get("status") == "لا يخص مشروع التكافل":
        target = record.get("not_takaful_target") or "-"
        if target == "جهة أخرى" and record.get("not_takaful_other"):
            target = f"جهة أخرى: {record.get('not_takaful_other')}"
        not_takaful_html = (
            f"<tr><th>الجهة التابع لها</th><td colspan='3'>{target}</td></tr>"
        )
    amount_ar = f"{float(record.get('amount') or 0):,.2f}"
    notes_html = ""
    if record.get("notes"):
        notes_html = (
            f"<tr><th>ملاحظات</th><td colspan='3'>{record.get('notes')}</td></tr>"
        )
    issued_at = record.get("issued_at") or "-"
    sub_month = record.get("subscription_month") or "-"
    recorded_by = record.get("recorded_by_name") or "-"
    today = today_ar()
    html = f"""<!DOCTYPE html>
<html dir=\"rtl\" lang=\"ar\">
<head>
<meta charset=\"utf-8\">
<title>{title} — {record.get('permit_number') or ''}</title>
<style>
  @page {{ size: A4 portrait; margin: 18mm 16mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: \"Cairo\", \"Tajawal\", \"Segoe UI\", Arial, sans-serif; color: #111827; margin: 0; padding: 0; }}
  .sheet {{ width: 100%; }}
  .header {{ text-align: center; border-bottom: 3px double #0f3a73; padding-bottom: 12px; margin-bottom: 16px; }}
  .header h1 {{ margin: 0; color: #0f3a73; font-size: 22pt; }}
  .header .sub {{ margin-top: 4px; font-size: 11pt; color: #475569; }}
  .info {{ width: 100%; border-collapse: collapse; font-size: 12pt; }}
  .info th, .info td {{ border: 1px solid #cbd5e1; padding: 8px 10px; text-align: start; }}
  .info th {{ background: #f1f5f9; color: #0f3a73; font-weight: 700; width: 22%; }}
  .amount-row td {{ background: #ecfdf5; color: #065f46; font-weight: 800; font-size: 14pt; }}
  .status-pill {{ display: inline-block; padding: 2px 10px; border-radius: 999px; font-weight: 700; font-size: 11pt; background: #e0e7ff; color: #1e3a8a; }}
  .footer {{ margin-top: 28px; display: grid; grid-template-columns: 1fr 1fr; gap: 24px; font-size: 11pt; }}
  .signbox {{ border-top: 1px solid #94a3b8; padding-top: 8px; text-align: center; color: #334155; }}
  .meta {{ margin-top: 18px; font-size: 9pt; color: #64748b; text-align: center; }}
  @media print {{ .no-print {{ display: none !important; }} }}
</style>
</head>
<body>
<div class=\"sheet\">
  <div class=\"header\">
    <h1>{title}</h1>
    <div class=\"sub\">مشروع التكافل الاجتماعي</div>
  </div>
  <table class=\"info\">
    <tr><th>رقم الإذن</th><td>{record.get('permit_number') or '-'}</td><th>تحريراً في</th><td>{issued_at}</td></tr>
    <tr><th>المحافظة</th><td>{record.get('governorate') or '-'}</td><th>اللجنة النقابية</th><td>{record.get('union_committee') or '-'}</td></tr>
    <tr><th>شهر الاشتراك</th><td>{sub_month}</td><th>طريقة الدفع</th><td>{method}</td></tr>
    {method_details_html}
    <tr class=\"amount-row\"><th>المبلغ</th><td colspan=\"3\">{amount_ar} ج.م</td></tr>
    <tr><th>الحالة</th><td colspan=\"3\"><span class=\"status-pill\">{record.get('status') or '-'}</span></td></tr>
    {not_takaful_html}
    {notes_html}
  </table>
  <div class=\"footer\">
    <div class=\"signbox\">الموظف المختص: <strong>{recorded_by}</strong></div>
    <div class=\"signbox\">المحاسب / المراجع</div>
  </div>
  <div class=\"meta\">طُبع في: {today}</div>
</div>
<script>setTimeout(function(){{window.print();}}, 250);</script>
</body>
</html>"""
    return HTMLResponse(content=html)


@api_router.delete("/subscriptions/{subscription_id}")
async def delete_subscription(subscription_id: str, user: Dict[str, Any] = Depends(require_user)):
    result = await db.subscriptions.delete_one({"id": subscription_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    return {"deleted": True}


@api_router.get("/subscriptions/summary")
async def subscriptions_summary(department_id: str, user: Dict[str, Any] = Depends(require_user)):
    counts = {}
    totals = {}
    for status in SUBSCRIPTION_STATUSES:
        items = await db.subscriptions.find({"department_id": department_id, "status": status}, {"_id": 0, "amount": 1}).to_list(5000)
        counts[status] = len(items)
        totals[status] = sum(float(item.get("amount") or 0) for item in items)
    counts["total"] = sum(counts.values())
    totals["total"] = sum(totals.values())
    return {"counts": counts, "totals": totals}


def _ym_to_index(ym: str) -> Optional[int]:
    """تحويل YYYY-MM إلى عدد شهور للترتيب/الفرق."""
    if not ym or len(ym) < 7:
        return None
    try:
        year = int(ym[:4])
        month = int(ym[5:7])
        return year * 12 + (month - 1)
    except (ValueError, TypeError):
        return None


def _date_to_ym(date_str: str) -> str:
    """Extract YYYY-MM from a date string supporting common formats."""
    if not date_str:
        return ""
    s = str(date_str).strip()
    # ISO YYYY-MM-DD or YYYY-MM already
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    # DD/MM/YYYY or DD-MM-YYYY
    parsed = parse_date_value(s)
    if parsed:
        return parsed.strftime("%Y-%m")
    return ""


async def _compute_committee_dues_for_period(
    department_id: str,
    governorate: str,
    union_committee: str,
    from_month: str,
    to_month: str,
) -> Dict[str, Any]:
    """Computes dues for a single (governorate, union_committee) over a YYYY-MM period.
    Uses ACTIVE-ONLY membership (no hold-back) to determine the raw amount the committee owes.

    Program-wide guarantee: nothing earlier than FINANCIAL_START_MONTH is ever
    counted, even if `from_month` was passed as an older value. This honours
    the user's "all finance starts at 2025-01" requirement.
    """
    from_month = _floor_to_financial_start(from_month)
    if _ym_to_index(from_month) and _ym_to_index(to_month) and _ym_to_index(from_month) > _ym_to_index(to_month):
        # to_month is before the cutoff -> empty result
        return {"expected_amount": 0.0, "paid_amount": 0.0, "owed_amount": 0.0, "credit_amount": 0.0,
                "membership_size": 0, "months": 0, "monthly_rate": COMMITTEE_MONTHLY_RATE,
                "from_month": from_month, "to_month": to_month}
    months = _months_in_range(from_month, to_month)
    if months == 0:
        return {"expected_amount": 0.0, "paid_amount": 0.0, "owed_amount": 0.0, "credit_amount": 0.0,
                "membership_size": 0, "months": 0, "monthly_rate": COMMITTEE_MONTHLY_RATE,
                "from_month": from_month, "to_month": to_month}
    gov = (governorate or "غير محدد").strip()
    com = (union_committee or "غير محدد").strip()
    # Active members in this committee (exclude all inactive + retirement-due)
    inactive_statuses = {"متوفي", "استقالة", "إسقاط", "عجز كلي أو جزئي منهي للخدمة"}
    members = await db.members.find(
        {"department_id": department_id, "governorate": gov, "union_committee": com},
        {"_id": 0},
    ).to_list(20000)
    enriched_members = await enrich_members_retirement(members)
    size = 0
    for m in enriched_members:
        if (m.get("status") or "") in inactive_statuses or m.get("retirement_due"):
            continue
        size += 1
    expected = size * COMMITTEE_MONTHLY_RATE * months
    # Sum collected subscriptions for the committee in period
    subs = await db.subscriptions.find(
        {"department_id": department_id, "status": "تم التحصيل", "governorate": gov, "union_committee": com},
        {"_id": 0, "amount": 1, "subscription_month": 1},
    ).to_list(20000)
    paid = 0.0
    a = _ym_to_index(from_month)
    b = _ym_to_index(to_month)
    for s in subs:
        idx = _ym_to_index(s.get("subscription_month") or "")
        if idx is None or a is None or b is None:
            continue
        if a <= idx <= b:
            paid += float(s.get("amount") or 0)
    balance = round(expected - paid, 2)
    return {
        "expected_amount": round(expected, 2),
        "paid_amount": round(paid, 2),
        "owed_amount": balance if balance > 0 else 0.0,
        "credit_amount": -balance if balance < 0 else 0.0,
        "membership_size": size,
        "months": months,
        "monthly_rate": COMMITTEE_MONTHLY_RATE,
        "from_month": from_month,
        "to_month": to_month,
    }


def _aid_period_for_member(member_snapshot: Dict[str, Any], today: Optional[datetime] = None) -> tuple:
    """Determine the period (from_month, to_month) for a member's committee dues calculation.
    Defaults: from subscription_date (or 24 months back) to status_date (or current month).
    """
    today = today or datetime.now(timezone.utc)
    to_ym = _date_to_ym(member_snapshot.get("member_status_date") or "") or f"{today.year:04d}-{today.month:02d}"
    from_ym = _date_to_ym(member_snapshot.get("member_subscription_date") or "")
    if not from_ym:
        # Fallback: 24 months back from to_ym
        idx = _ym_to_index(to_ym)
        if idx is not None:
            back = idx - 23
            from_ym = f"{back // 12:04d}-{(back % 12) + 1:02d}"
        else:
            from_ym = to_ym
    if _ym_to_index(from_ym) and _ym_to_index(to_ym) and _ym_to_index(from_ym) > _ym_to_index(to_ym):
        from_ym = to_ym
    # Apply the program-wide financial cutoff: nothing earlier than FINANCIAL_START_MONTH.
    from_ym = _floor_to_financial_start(from_ym)
    if _ym_to_index(from_ym) and _ym_to_index(to_ym) and _ym_to_index(from_ym) > _ym_to_index(to_ym):
        # Member's status_date is before the financial cutoff -> empty period.
        from_ym = to_ym = FINANCIAL_START_MONTH
    return from_ym, to_ym


# ─── Program-wide financial floor ───────────────────────────────────────────
# All committee-dues / arrears calculations must NEVER look at months earlier
# than this. The user requested 2025-01 as the program's "start of accounting".
FINANCIAL_START_MONTH = "2025-01"


def _floor_to_financial_start(ym: str) -> str:
    """Clamp a YYYY-MM string to be >= FINANCIAL_START_MONTH.

    Returns FINANCIAL_START_MONTH if `ym` is empty, malformed, or earlier than
    the configured floor. Otherwise returns `ym` unchanged.
    """
    if not ym or _ym_to_index(ym) is None:
        return FINANCIAL_START_MONTH
    if _ym_to_index(ym) < _ym_to_index(FINANCIAL_START_MONTH):
        return FINANCIAL_START_MONTH
    return ym



def _months_in_range(from_month: str, to_month: str) -> int:
    a = _ym_to_index(from_month)
    b = _ym_to_index(to_month)
    if a is None or b is None or b < a:
        return 0
    return b - a + 1


# -------------------- Committee Prior Arrears --------------------

class PriorArrearIn(BaseModel):
    department_id: str
    governorate: str
    union_committee: str
    period_label: str = ""   # free text like "2020" or "2020-05" or "ما قبل النظام"
    amount: float = 0.0
    note: str = ""


@api_router.get("/committees/prior-arrears")
async def list_prior_arrears(
    department_id: str,
    governorate: Optional[str] = None,
    union_committee: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    q: Dict[str, Any] = {"department_id": department_id}
    if governorate:
        q["governorate"] = governorate
    if union_committee:
        q["union_committee"] = union_committee
    docs = await db.committee_prior_arrears.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return docs


@api_router.post("/committees/prior-arrears")
async def add_prior_arrear(payload: PriorArrearIn, user: Dict[str, Any] = Depends(require_admin)):
    doc = payload.model_dump()
    doc["id"] = build_id()
    doc["created_at"] = now_iso()
    doc["created_by_id"] = user.get("id", "")
    doc["created_by_name"] = user.get("display_name") or user.get("username", "")
    await db.committee_prior_arrears.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.delete("/committees/prior-arrears/{arrear_id}")
async def delete_prior_arrear(arrear_id: str, user: Dict[str, Any] = Depends(require_admin)):
    res = await db.committee_prior_arrears.delete_one({"id": arrear_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    return {"ok": True}


@api_router.get("/committees/dues")
async def committees_dues(
    department_id: str,
    from_month: Optional[str] = None,  # YYYY-MM
    to_month: Optional[str] = None,    # YYYY-MM
    governorate: Optional[str] = None,
    user: Dict[str, Any] = Depends(require_user),
):
    if not from_month or not to_month:
        # Default = current month only
        today = datetime.now(timezone.utc)
        from_month = from_month or f"{today.year:04d}-{today.month:02d}"
        to_month = to_month or f"{today.year:04d}-{today.month:02d}"
    # Honour the program-wide "all finance starts at 2025-01" cutoff.
    from_month = _floor_to_financial_start(from_month)
    months = _months_in_range(from_month, to_month)
    if months == 0:
        raise HTTPException(status_code=400, detail="فترة غير صالحة (تأكد من from_month <= to_month بصيغة YYYY-MM)")

    # 1. Active membership size per (governorate, committee)
    member_query: Dict[str, Any] = {"department_id": department_id}
    members = await db.members.find(member_query, {"_id": 0}).to_list(100000)
    enriched_members = await enrich_members_retirement(members)
    # "Hard" inactive: voluntarily left → always excluded
    hard_inactive = {"استقالة", "إسقاط"}
    # "Soft" inactive: deceased / disability → excluded ONLY if their committee owes nothing
    soft_inactive = {"متوفي", "عجز كلي أو جزئي منهي للخدمة"}
    sizes: Dict[tuple, int] = defaultdict(int)              # active members (excludes soft + hard + retirement)
    soft_counts: Dict[tuple, int] = defaultdict(int)        # deceased/disabled members per (gov, com)
    for m in enriched_members:
        status = (m.get("status") or "")
        gov = (m.get("governorate") or "غير محدد").strip()
        com = (m.get("union_committee") or "غير محدد").strip()
        if status in hard_inactive:
            continue
        if status in soft_inactive:
            soft_counts[(gov, com)] += 1
            continue
        if m.get("retirement_due"):
            continue
        sizes[(gov, com)] += 1

    # 2. Collected subscriptions per (gov, committee) within period
    sub_query: Dict[str, Any] = {"department_id": department_id, "status": "تم التحصيل"}
    subs = await db.subscriptions.find(sub_query, {"_id": 0}).to_list(20000)
    paid: Dict[tuple, float] = defaultdict(float)
    for s in subs:
        idx = _ym_to_index(s.get("subscription_month") or "")
        if idx is None:
            continue
        if not (_ym_to_index(from_month) <= idx <= _ym_to_index(to_month)):
            continue
        gov = (s.get("governorate") or "غير محدد").strip()
        com = (s.get("union_committee") or "غير محدد").strip()
        paid[(gov, com)] += float(s.get("amount") or 0)

    # 3. Build rows: every key from sizes union paid union soft_counts
    keys = set(sizes.keys()) | set(paid.keys()) | set(soft_counts.keys())
    rows = []
    for (gov, com) in sorted(keys):
        if governorate and gov != governorate:
            continue
        if _is_invalid_taxonomy_value(gov) or _is_invalid_taxonomy_value(com):
            continue
        base_size = sizes.get((gov, com), 0)
        soft_size = soft_counts.get((gov, com), 0)
        paid_amount = round(paid.get((gov, com), 0.0), 2)
        # First pass: assume soft-inactive (deceased/disability) are excluded
        expected_base = base_size * COMMITTEE_MONTHLY_RATE * months
        # If committee still owes after excluding them, they are HELD BACK and re-included
        held_back = (expected_base - paid_amount) > 0 and soft_size > 0
        size = base_size + (soft_size if held_back else 0)
        expected = size * COMMITTEE_MONTHLY_RATE * months
        balance = round(expected - paid_amount, 2)
        rows.append({
            "governorate": gov,
            "union_committee": com,
            "membership_size": size,
            "active_size": base_size,
            "held_back_size": soft_size if held_back else 0,
            "months": months,
            "monthly_rate": COMMITTEE_MONTHLY_RATE,
            "expected_amount": round(expected, 2),
            "paid_amount": paid_amount,
            "owed_amount": balance if balance > 0 else 0.0,
            "credit_amount": -balance if balance < 0 else 0.0,
            "settled": balance == 0,
        })

    # 4. Aggregate by governorate (overall summary)
    # First, fetch all prior arrears for this department and add them to relevant rows.
    prior_docs = await db.committee_prior_arrears.find(
        {"department_id": department_id}, {"_id": 0}
    ).to_list(5000)
    prior_by_key: Dict[tuple, float] = defaultdict(float)
    for p in prior_docs:
        key = ((p.get("governorate") or "غير محدد").strip(), (p.get("union_committee") or "غير محدد").strip())
        prior_by_key[key] += float(p.get("amount") or 0.0)
    for r in rows:
        key = (r["governorate"], r["union_committee"])
        prior = round(prior_by_key.get(key, 0.0), 2)
        r["prior_arrears"] = prior
        if prior > 0:
            r["owed_amount"] = round(r["owed_amount"] + prior, 2)
    # Also include committees that have only prior arrears (no active members in this department).
    existing_keys = {(r["governorate"], r["union_committee"]) for r in rows}
    for key, prior in prior_by_key.items():
        if key in existing_keys or prior <= 0:
            continue
        rows.append({
            "governorate": key[0],
            "union_committee": key[1],
            "membership_size": 0,
            "active_size": 0,
            "held_back_size": 0,
            "months": months,
            "monthly_rate": COMMITTEE_MONTHLY_RATE,
            "expected_amount": 0.0,
            "paid_amount": 0.0,
            "owed_amount": round(prior, 2),
            "credit_amount": 0.0,
            "prior_arrears": round(prior, 2),
            "settled": False,
        })

    totals = {
        "membership_size": sum(r["membership_size"] for r in rows),
        "expected_amount": round(sum(r["expected_amount"] for r in rows), 2),
        "paid_amount": round(sum(r["paid_amount"] for r in rows), 2),
        "owed_amount": round(sum(r["owed_amount"] for r in rows), 2),
        "credit_amount": round(sum(r["credit_amount"] for r in rows), 2),
        "prior_arrears": round(sum(r.get("prior_arrears", 0.0) for r in rows), 2),
    }

    gov_acc: Dict[str, Dict[str, float]] = {}
    for r in rows:
        g = r["governorate"]
        acc = gov_acc.setdefault(g, {
            "governorate": g,
            "committees_count": 0,
            "membership_size": 0,
            "expected_amount": 0.0,
            "paid_amount": 0.0,
            "owed_amount": 0.0,
            "credit_amount": 0.0,
        })
        acc["committees_count"] += 1
        acc["membership_size"] += r["membership_size"]
        acc["expected_amount"] += r["expected_amount"]
        acc["paid_amount"] += r["paid_amount"]
        acc["owed_amount"] += r["owed_amount"]
        acc["credit_amount"] += r["credit_amount"]

    governorate_totals = []
    for g in sorted(gov_acc.keys()):
        acc = gov_acc[g]
        governorate_totals.append({
            "governorate": acc["governorate"],
            "committees_count": acc["committees_count"],
            "membership_size": acc["membership_size"],
            "expected_amount": round(acc["expected_amount"], 2),
            "paid_amount": round(acc["paid_amount"], 2),
            "owed_amount": round(acc["owed_amount"], 2),
            "credit_amount": round(acc["credit_amount"], 2),
            "settled": round(acc["expected_amount"] - acc["paid_amount"], 2) == 0,
        })

    return {
        "from_month": from_month,
        "to_month": to_month,
        "months": months,
        "monthly_rate": COMMITTEE_MONTHLY_RATE,
        "rows": rows,
        "totals": totals,
        "governorate_totals": governorate_totals,
    }


@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    # Exclude MongoDB's _id field from the query results
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Live-sync auto-bumper middleware ──────────────────────────────────────
# Detects successful write operations by URL prefix and bumps the per-collection
# change counter so peer clients refresh silently within their poll interval.
_LIVE_SYNC_PATHS = (
    ("/api/members", "members"),
    ("/api/subscriptions", "subscriptions"),
    ("/api/aids", "aids"),
    ("/api/departments", "departments"),
    ("/api/letters", "letters"),
    ("/api/dues-settlements", "dues_settlements"),
    ("/api/retirement-schedules", "retirement_schedules"),
)


@app.middleware("http")
async def _live_sync_middleware(request, call_next):
    response = await call_next(request)
    try:
        method = (request.method or "GET").upper()
        if method in ("POST", "PUT", "PATCH", "DELETE") and 200 <= response.status_code < 300:
            path = request.url.path or ""
            for prefix, coll in _LIVE_SYNC_PATHS:
                if path.startswith(prefix):
                    await _bump_change_counter(coll)
                    break
    except Exception:
        pass
    return response

# -------- Serve built frontend (LAN deployment) --------
# When the frontend is built to /app/frontend/build, this mount makes the
# whole app self-contained: open http://[server-ip]:PORT and the SPA + APIs work.
# IMPORTANT: We use a 404 exception handler instead of a catch-all GET route to
# avoid shadowing API routes (a catch-all GET would cause FastAPI to return 405
# "Method Not Allowed" for POST/PUT/DELETE to unknown /api/* paths instead of 404).
try:
    from fastapi.staticfiles import StaticFiles
    from fastapi.responses import FileResponse as _FileResponse
    from fastapi import Request as _Request
    from starlette.exceptions import HTTPException as _StarletteHTTPException
    import hashlib as _hashlib

    _FRONTEND_BUILD = Path(__file__).resolve().parent.parent / "frontend" / "build"

    def _compute_build_version() -> str:
        """Return a short hash of the current frontend bundle filenames.

        The frontend's CRA build uses content-hashed filenames (e.g. main.abc123.js),
        so concatenating them gives a stable identifier that changes on every build.
        """
        try:
            static_dir = _FRONTEND_BUILD / "static"
            if not static_dir.exists():
                return "dev"
            names = sorted(p.name for p in static_dir.rglob("*") if p.is_file())
            return _hashlib.sha1("\n".join(names).encode("utf-8")).hexdigest()[:12]
        except Exception:
            return "unknown"

    _BUILD_VERSION = _compute_build_version()

    @app.get("/api/version")
    async def app_version():
        """Public liveness + version endpoint. The frontend polls this so it can
        auto-reload itself when a new build is deployed."""
        return {"version": _BUILD_VERSION, "ts": now_iso()}

    def _file_response_no_cache(path: Path) -> "_FileResponse":
        """Serve a file with strict 'no-cache' so browsers always re-check.

        Used for index.html (and similar entry points) — the rest of /static is
        already content-hashed and can be cached aggressively below.
        """
        resp = _FileResponse(str(path))
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        resp.headers["X-App-Version"] = _BUILD_VERSION
        return resp

    if _FRONTEND_BUILD.exists() and (_FRONTEND_BUILD / "index.html").exists():
        # /static/* files have content-hashed names, so they can be cached forever.
        # We wrap StaticFiles to add an immutable cache header.
        class _ImmutableStaticFiles(StaticFiles):
            async def get_response(self, path: str, scope):
                response = await super().get_response(path, scope)
                # Only cache successful asset responses; let 404s pass through.
                if getattr(response, "status_code", 0) == 200:
                    response.headers["Cache-Control"] = "public, max-age=31536000, immutable"
                return response

        if (_FRONTEND_BUILD / "static").exists():
            app.mount(
                "/static",
                _ImmutableStaticFiles(directory=str(_FRONTEND_BUILD / "static")),
                name="frontend-static",
            )

        @app.exception_handler(404)
        async def _spa_fallback(request: _Request, exc: _StarletteHTTPException):
            path = request.url.path.lstrip("/")
            # Never intercept API / docs / openapi routes — return JSON 404 as usual
            if path.startswith("api") or path.startswith("docs") or path.startswith("openapi") or path.startswith("redoc"):
                from fastapi.responses import JSONResponse
                return JSONResponse({"detail": exc.detail or "Not Found"}, status_code=404)
            # Try to serve a static asset directly (favicon, manifest, etc.)
            target = _FRONTEND_BUILD / path
            if path and target.is_file():
                # Non-/static root assets (favicon, manifest.json, robots.txt) —
                # short cache so the user picks up changes within minutes.
                resp = _FileResponse(str(target))
                resp.headers["Cache-Control"] = "public, max-age=300"
                resp.headers["X-App-Version"] = _BUILD_VERSION
                return resp
            # Otherwise serve the SPA's index.html so React Router can handle it
            return _file_response_no_cache(_FRONTEND_BUILD / "index.html")
except Exception as _e:  # noqa: BLE001
    # Static frontend serving is optional; in cloud preview the frontend is served separately.
    pass

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


@app.on_event("startup")
async def startup_seed():
    await seed_defaults()